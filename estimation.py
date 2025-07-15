import openpyxl
from openpyxl.utils import get_column_letter

def generate_excel_report(
    system_input: str,
    elevation_type: str,
    total_count: int,
    bays_wide: int,
    bays_tall: int,
    opening_width: float,
    opening_height: float,
    sqft_per_type: float,
    total_sqft: float,
    perimeter_ft: float,
    total_perimeter_ft: float,
    completion_callback # A callback function to update UI status
):
    """
    Performs Excel calculations and generates an Excel file based on provided inputs.
    Updates the UI via a callback function.
    """
    # --- System Input Validation ---
    if system_input != "YES 45TU Front Set(OG)":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "System not matched. Empty file created."
        try:
            wb.save("output.xlsx")
            completion_callback(f"System not matched. Empty 'output.xlsx' created.", "orange")
        except Exception as e:
            completion_callback(f"Error saving empty file: {e}", "red")
        return # Exit the function if system doesn't match

    # --- Proceed with calculations and Excel generation if system matches ---
    try:
        # Create a new Excel workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define input and output headers
        inputs_headers = [
            "System Input", "Elevation Type", "Total Count",
            "# Bays Wide", "# Bays Tall",
            "Opening Width", "Opening Height",
            "Sq Ft per Type", "Total Sq Ft", "Perimeter Ft", "Total Perimeter Ft"
        ]

        outputs_headers = [
            "Total Gasket (Ft)", "End Dam", "Water Deflector", "Assembly Screw",
            "Sill Flash Screw", "End Dam Screw", "Setting Block Chair",
            "Side Block", "Setting Block", "Anti Walk Block Deep Pocket",
            "Anti Walk Block Shallow Pocket", "Setting Block (Int. Horizontal)",
            "Jamb Ft (V)", "Sill Ft (H)", "Flush Filler (V)",
            "Int. Vertical", "OG Int. Horizontal", "OG Head (H)", "Sill Flashing (H)"
        ]

        # Combine input and output headers and write them to the first row
        headers = inputs_headers + outputs_headers
        for idx, header in enumerate(headers):
            ws[f"{get_column_letter(idx + 1)}1"] = header

        # Prepare input values for row 2
        input_values = [
            system_input, elevation_type, total_count,
            bays_wide, bays_tall, opening_width, opening_height,
            sqft_per_type, total_sqft, perimeter_ft, total_perimeter_ft
        ]

        # Calculate output values directly and store them in a list
        output_values = [
            (((bays_wide * 4 * opening_height) + (bays_tall * 4 * opening_width)) * total_count) / 12,  # Total Gasket (Ft)
            2 * total_count,  # End Dam
            2 * bays_wide * total_count,  # Water Deflector
            ((bays_wide * 8) + (((bays_tall - 1) * 6) * bays_wide)) * total_count,  # Assembly Screw
            3 * bays_wide * total_count,  # Sill Flash Screw
            4 * total_count,  # End Dam Screw
            2 * bays_wide,  # Setting Block Chair
            (bays_wide - 1) * bays_tall * total_count,  # Side Block
            2 * bays_wide * total_count,  # Setting Block
            2 * bays_tall * total_count,  # Anti Walk Block Deep Pocket
            (bays_wide - 1) * bays_tall * total_count,  # Anti Walk Block Shallow Pocket
            2 * bays_wide * total_count,  # Setting Block (Int. Horizontal)
            (2 * opening_height) / 12 * total_count,  # Jamb Ft (V)
            (opening_width / 12) * total_count,  # Sill Ft (H)
            ((bays_wide - 1) * total_count * opening_height) / 12,  # Flush Filler (V)
            ((bays_wide - 1) * total_count * opening_height) / 12,  # Int. Vertical (same as Flush Filler)
            (opening_width / 12) * total_count,  # OG Int. Horizontal
            (opening_width / 12) * total_count,  # OG Head (H)
            (opening_width / 12) * total_count   # Sill Flashing (H)
        ]

        # Combine inputs and outputs and write them to row 2
        all_values = input_values + output_values
        for idx, val in enumerate(all_values):
            ws[f"{get_column_letter(idx + 1)}2"] = val

        # Auto-size each column based on max length of header or data
        for idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(idx)
            cell_value = str(ws[f"{col_letter}2"].value) if ws[f"{col_letter}2"].value is not None else ""
            ws.column_dimensions[col_letter].width = max(len(header), len(cell_value)) + 2

        # Save workbook
        wb.save("output.xlsx")
        completion_callback("Excel file 'output.xlsx' generated successfully!", "green")
        print("Excel file saved as 'output.xlsx'") # Also print to console for confirmation

    except Exception as e:
        completion_callback(f"An error occurred during Excel generation: {e}", "red")
        print(f"Error during Excel generation: {e}")

