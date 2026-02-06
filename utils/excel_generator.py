import os
import json
import math
import re
from openpyxl import Workbook
from openpyxl.styles import Font, numbers, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import Counter
import datetime

from utils.pricing import get_price_by_part, reverse_material_impact, load_extra_materials, save_extra_materials, apply_material_impact_to_extra_materials_in_memory, get_unit_price_by_part, parse_length_to_feet, BAY_WIDTH_PARTS, _is_bay_width_part
EPSILON = 1e-9  # Small value for floating point comparisons
from data.part_number import PART_NUMBER_MAP
from data.parts_data import parts_data
from utils.formulas import calculate_door_info

# --- Diagram helpers (doors) ---
def _parse_door_size_inches(size_str):
    """Parse door size string (e.g. \"3' X 7'\") to (width_inches, height_inches)."""
    if not size_str:
        return 36.0, 84.0
    m = re.search(r"(\d+)'\s*[xX]\s*(\d+)'", str(size_str))
    if m:
        return float(m.group(1)) * 12.0, float(m.group(2)) * 12.0
    return 36.0, 84.0

def _door_spans_for_diagram(doors, opening_width):
    """Yield (left_in, right_in, door_h_in) for each door instance for drawing on bay diagram."""
    for door in (doors or []):
        size_str = door.get("size", "")
        count = door.get("count", 1)
        dw, dh = _parse_door_size_inches(size_str)
        xs = []
        if count == 1:
            x_in = door.get("x_in")
            if x_in is not None:
                try:
                    xs.append(float(x_in))
                except (TypeError, ValueError):
                    pass
        else:
            for x in (door.get("x_positions") or [])[:count]:
                if x is not None:
                    try:
                        xs.append(float(x))
                    except (TypeError, ValueError):
                        pass
        for x_center in xs:
            left_in = max(0, min(x_center - dw / 2, opening_width))
            right_in = max(0, min(x_center + dw / 2, opening_width))
            yield left_in, right_in, dh

# --- Helper Functions ---

def _get_multiplier(running_grand_total):
    """Returns multiplier based on running grand total."""
    return 0.614 if running_grand_total < 50000 else 0.572

def _autofit_columns_by_longest_word(ws, start_col, end_col, start_row, end_row):
    """Fit columns to the longest word in any cell. Keeps columns tight so values aren't pushed far right."""
    for col_idx in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_idx)
        max_word_len = 0
        max_full_len = 0
        for r in range(start_row, end_row + 1):
            cell_value = ws.cell(row=r, column=col_idx).value
            if cell_value is not None:
                s = str(cell_value).strip()
                max_full_len = max(max_full_len, len(s))
                for word in s.split():
                    max_word_len = max(max_word_len, len(word))
        if max_word_len > 0:
            # Use longest word + 2, but ensure we don't truncate (use full len for short values)
            width = max(max_word_len + 2, min(max_full_len + 1, 20))
            ws.column_dimensions[col_letter].width = width

def _autofit_columns(ws, start_col, end_col, start_row=1, end_row=None):
    """Autofits columns in the worksheet. Ensures minimum width for currency columns to prevent ######## display."""
    end_row = end_row if end_row is not None else ws.max_row
    for col_idx in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        has_numbers = False
        for r in range(start_row, end_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell_value = cell.value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
                if isinstance(cell_value, (int, float)) or (cell.number_format and '$' in str(cell.number_format).upper()):
                    has_numbers = True
        # Currency/number columns need min width 14 to prevent ######## in Excel
        if has_numbers:
            max_len = max(max_len, 12)
        current_width_obj = ws.column_dimensions[col_letter]
        current_width = current_width_obj.width if current_width_obj.width is not None else 0.0
        if col_idx == 5:  # Column E (Description)
            ws.column_dimensions[col_letter].width = max(max_len, current_width)
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 2, 14 if has_numbers else 0)

def _clean_trailing_blank_rows(ws, start_row):
    """Deletes blank rows from the worksheet starting from a given row."""
    rows_deleted = 0
    current_row = start_row
    while current_row <= ws.max_row:
        if all(ws.cell(row=current_row, column=c).value is None for c in range(1, ws.max_column + 1)):
            ws.delete_rows(current_row, 1)
            rows_deleted += 1
        else: current_row += 1

def _create_cost_pie_chart(material_cost, misc_cost, markup_cost, residual_cost):
    """
    Creates a pie chart showing cost breakdown: Materials, Additional, Markups, and Residual.
    Returns a PIL Image object that can be inserted into Excel.
    """
    try:
        from PIL import Image, ImageDraw, ImageFont
        import io
    except ImportError:
        print("PIL/Pillow not available, skipping pie chart generation")
        return None
    
    # Calculate grand total
    grand_total = material_cost + misc_cost + markup_cost + residual_cost
    
    if grand_total <= 0:
        return None
    
    # Calculate percentages
    material_pct = (material_cost / grand_total * 100) if grand_total > 0 else 0
    misc_pct = (misc_cost / grand_total * 100) if grand_total > 0 else 0
    markup_pct = (markup_cost / grand_total * 100) if grand_total > 0 else 0
    residual_pct = (residual_cost / grand_total * 100) if grand_total > 0 else 0
    
    # Chart dimensions
    chart_width = 420
    chart_height = 400
    center_x = chart_width // 2
    center_y = 160
    radius = 90
    
    # Create image with white background
    img = Image.new('RGB', (chart_width, chart_height), color='white')
    draw = ImageDraw.Draw(img)
    
    # Try to load fonts
    try:
        font_title = ImageFont.truetype("arial.ttf", 14)
        font_label = ImageFont.truetype("arial.ttf", 10)
        font_small = ImageFont.truetype("arial.ttf", 9)
    except:
        try:
            font_title = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 14)
            font_label = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 10)
            font_small = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 9)
        except:
            font_title = ImageFont.load_default()
            font_label = ImageFont.load_default()
            font_small = ImageFont.load_default()
    
    # Draw title
    draw.text((center_x, 15), "Project Cost Breakdown", fill='#333333', anchor='mm', font=font_title)
    
    # Colors for each segment
    material_color = '#4472C4'  # Blue for materials
    misc_color = '#548235'      # Green for addition cost
    markup_color = '#7030A0'    # Purple for markups/profit
    residual_color = '#ED7D31'  # Orange for residual/waste
    
    # Build segments list (only include non-zero values)
    segments = []
    if material_cost > 0:
        segments.append(('Active Materials', material_cost, material_pct, material_color))
    if misc_cost > 0:
        segments.append(('Additional', misc_cost, misc_pct, misc_color))
    if markup_cost > 0:
        segments.append(('Profit/Markups', markup_cost, markup_pct, markup_color))
    if residual_cost > 0:
        segments.append(('Residual/Waste', residual_cost, residual_pct, residual_color))
    
    # Draw pie chart
    start_angle = -90
    for name, cost, pct, color in segments:
        if pct > 0:
            sweep_angle = (pct / 100) * 360
            draw.pieslice(
                [center_x - radius, center_y - radius, center_x + radius, center_y + radius],
                start=start_angle,
                end=start_angle + sweep_angle,
                fill=color,
                outline='white'
            )
            start_angle += sweep_angle
    
    # Draw legend at bottom
    legend_y = 270
    legend_box_size = 12
    legend_spacing = 22
    
    legend_items = [
        ('Active Materials', material_cost, material_pct, material_color),
        ('Additional', misc_cost, misc_pct, misc_color),
        ('Profit/Markups', markup_cost, markup_pct, markup_color),
        ('Residual/Waste', residual_cost, residual_pct, residual_color)
    ]
    
    for i, (name, cost, pct, color) in enumerate(legend_items):
        y_pos = legend_y + (i * legend_spacing)
        draw.rectangle([30, y_pos, 30 + legend_box_size, y_pos + legend_box_size], fill=color, outline='#333333')
        draw.text((50 + legend_box_size, y_pos + legend_box_size // 2), 
                  f"{name}: ${cost:,.2f} ({pct:.1f}%)", 
                  fill='#333333', anchor='lm', font=font_label)
    
    # Grand total at bottom
    draw.text((center_x, chart_height - 15), f"Grand Total: ${grand_total:,.2f}", fill='#333333', anchor='mm', font=font_small)
    
    return img

def _add_pie_chart_to_excel(ws, start_row, start_col, material_cost, misc_cost, markup_cost, residual_cost):
    """Adds a cost distribution pie chart to the Excel worksheet."""
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
        import io
        
        # Create the pie chart
        chart_img = _create_cost_pie_chart(material_cost, misc_cost, markup_cost, residual_cost)
        
        if chart_img:
            # Save to bytes
            img_bytes = io.BytesIO()
            chart_img.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            
            # Add to Excel
            img = OpenpyxlImage(img_bytes)
            # Scale image
            img.width = 380
            img.height = 360
            col_letter = get_column_letter(start_col)
            img.anchor = f'{col_letter}{start_row}'
            ws.add_image(img)
            
            # Return estimated row height for spacing
            estimated_rows = max(22, int(img.height / 15))
            return start_row + estimated_rows + 2
    except Exception as e:
        print(f"Error creating pie chart: {e}")
    
    return start_row + 2

def _create_bay_diagram(bays_wide, bays_tall, opening_width, opening_height, custom_bay_widths=None, custom_bay_heights=None, doors=None):
    """
    Creates a visual blueprint diagram of the bay distribution.
    If doors is provided, draws green door bands (to scale) on the diagram.
    Returns a PIL Image object that can be inserted into Excel.
    """
    try:
        from PIL import Image, ImageDraw, ImageFont
        import io
    except ImportError:
        print("PIL/Pillow not available, skipping diagram generation")
        return None
    
    diagram_width = 400
    diagram_height = 300
    margin = 20  # Reduced for tighter spacing when door is present
    if doors:
        margin = 15  # Even tighter when doors are drawn
    if custom_bay_widths and len(custom_bay_widths) == bays_wide:
        bay_widths = custom_bay_widths
    else:
        bay_widths = [opening_width / bays_wide] * bays_wide if bays_wide > 0 else []
    
    if custom_bay_heights and len(custom_bay_heights) == bays_tall:
        bay_heights = custom_bay_heights
    else:
        bay_heights = [opening_height / bays_tall] * bays_tall if bays_tall > 0 else []
    
    if not bay_widths or not bay_heights:
        return None
    
    img = Image.new('RGB', (diagram_width, diagram_height), color='white')
    draw = ImageDraw.Draw(img)
    
    max_display_width = diagram_width - 2 * margin
    max_display_height = diagram_height - 2 * margin - 60
    
    total_width = sum(bay_widths)
    total_height = sum(bay_heights)
    scale_x = max_display_width / total_width if total_width > 0 else 1
    scale_y = max_display_height / total_height if total_height > 0 else 1
    scale = min(scale_x, scale_y)
    
    scaled_total_width = total_width * scale
    scaled_total_height = total_height * scale
    start_x = margin + (max_display_width - scaled_total_width) / 2
    start_y = margin + 30
    
    try:
        font_large = ImageFont.truetype("arial.ttf", 12)
        font_small = ImageFont.truetype("arial.ttf", 8)
    except:
        try:
            font_large = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 12)
            font_small = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 8)
        except:
            font_large = ImageFont.load_default()
            font_small = ImageFont.load_default()
    
    draw.text((diagram_width // 2, 10), "Bay Distribution Layout", fill='black', anchor='mm', font=font_large)
    
    current_x = start_x
    current_y = start_y
    
    for i, width in enumerate(bay_widths):
        if i > 0:
            draw.line([(current_x, start_y), (current_x, start_y + scaled_total_height)], fill='gray', width=2)
        current_x += width * scale
    
    current_x = start_x
    for i, height in enumerate(bay_heights):
        if i > 0:
            draw.line([(start_x, current_y), (start_x + scaled_total_width, current_y)], fill='gray', width=2)
        current_y += height * scale
    
    draw.rectangle([start_x, start_y, start_x + scaled_total_width, start_y + scaled_total_height], 
                   outline='black', width=3)
    
    current_x = start_x
    current_y = start_y
    bay_num = 1
    
    for row in range(bays_tall):
        current_x = start_x
        for col in range(bays_wide):
            bay_center_x = current_x + (bay_widths[col] * scale) / 2
            bay_center_y = current_y + (bay_heights[row] * scale) / 2
            draw.text((bay_center_x, bay_center_y - 6), f"B{bay_num}", fill='black', anchor='mm', font=font_small)
            dim_text = f"{bay_widths[col]:.1f}\" x {bay_heights[row]:.1f}\""
            draw.text((bay_center_x, bay_center_y + 6), dim_text, fill='black', anchor='mm', font=font_small)
            current_x += bay_widths[col] * scale
            bay_num += 1
        current_y += bay_heights[row] * scale
    
    # Draw door bands (green) when doors provided — use total_width for alignment with bay grid
    width_ref = total_width if total_width > 0 else opening_width
    if doors and width_ref > 0 and opening_height > 0 and scaled_total_width > 0 and scaled_total_height > 0:
        for left_in, right_in, door_h_in in _door_spans_for_diagram(doors, opening_width):
            if right_in <= left_in:
                continue
            px_left = start_x + (left_in / width_ref) * scaled_total_width
            px_right = start_x + (right_in / width_ref) * scaled_total_width
            door_h_px = (door_h_in / opening_height) * scaled_total_height
            px_bottom = start_y + scaled_total_height
            px_top = px_bottom - door_h_px
            draw.rectangle(
                [px_left, px_top, px_right, px_bottom],
                outline='#2E7D32',
                width=2,
                fill='#A5D6A7'
            )
    
    dim_text = f"Total: {opening_width:.1f}\" W x {opening_height:.1f}\" H"
    draw.text((diagram_width // 2, diagram_height - 20), dim_text, fill='black', anchor='mm', font=font_small)
    
    return img

def _add_bay_diagram_to_excel(ws, start_row, bays_wide, bays_tall, opening_width, opening_height, custom_bay_widths=None, custom_bay_heights=None, doors=None):
    """Adds a bay distribution diagram to the Excel worksheet. If doors provided, includes green door bands."""
    if bays_wide == 0 or bays_tall == 0:
        return start_row
    
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
        import io
        
        diagram_img = _create_bay_diagram(bays_wide, bays_tall, opening_width, opening_height, custom_bay_widths, custom_bay_heights, doors=doors)
        
        if diagram_img:
            # Save to bytes
            img_bytes = io.BytesIO()
            diagram_img.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            
            img = OpenpyxlImage(img_bytes)
            original_width = img.width
            original_height = img.height
            img.width = min(original_width, _DIAGRAM_MAX_WIDTH)
            img.height = int(original_height * (img.width / original_width))
            img.anchor = f'A{start_row}'  # Place starting in column A
            ws.add_image(img)
            
            # Return the row after the image (estimate image height)
            # Image height in rows (approximately 1 row per 15 pixels at default row height)
            estimated_rows = max(15, int(img.height / _DIAGRAM_ROW_HEIGHT_PX))
            return start_row + estimated_rows + _EXTRA_ROWS_AFTER_IMAGE
    except Exception as e:
        print(f"Error creating bay diagram: {e}")
        ws.cell(row=start_row, column=3, value="Bay diagram could not be generated")
    
    return start_row + 2

# Consistent diagram sizing for Excel (columns A-C)
_DIAGRAM_MAX_WIDTH = 280  # Slightly smaller
_DIAGRAM_ROW_HEIGHT_PX = 15
_EXTRA_ROWS_AFTER_IMAGE = 1
# Door-only: same column, stacked up/down, as close as possible
_DOOR_DIAGRAM_GAP_ROWS = 0  # Minimal gap between stacked diagrams

def _create_door_only_diagram_single(dw_in, dh_in, label):
    """Create a single door-only diagram image (one door, to scale). Returns PIL Image or None."""
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return None
    diagram_width = 400
    diagram_height = 300
    margin = 50
    max_display_w = diagram_width - 2 * margin
    max_display_h = diagram_height - 2 * margin - 50
    ref_height = 96.0
    scale = max_display_h / ref_height
    if dw_in * scale > max_display_w:
        scale = max_display_w / dw_in
    scaled_w = dw_in * scale
    scaled_h = dh_in * scale
    start_x = margin + (max_display_w - scaled_w) / 2
    start_y = margin + 30
    img = Image.new('RGB', (diagram_width, diagram_height), color='white')
    draw = ImageDraw.Draw(img)
    try:
        font_large = ImageFont.truetype("arial.ttf", 12)
        font_small = ImageFont.truetype("arial.ttf", 8)
    except:
        try:
            font_large = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 12)
            font_small = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 8)
        except:
            font_large = ImageFont.load_default()
            font_small = ImageFont.load_default()
    draw.text((diagram_width // 2, 10), "Door Only — To Scale", fill='black', anchor='mm', font=font_large)
    draw.rectangle(
        [start_x, start_y, start_x + scaled_w, start_y + scaled_h],
        outline='#2E7D32',
        width=2,
        fill='#A5D6A7'
    )
    cx = start_x + scaled_w / 2
    cy = start_y + scaled_h / 2
    dim_txt = f"{dw_in:.0f}\" x {dh_in:.0f}\""
    draw.text((cx, cy - 6), dim_txt, fill='black', anchor='mm', font=font_small)
    draw.text((cx, cy + 6), label, fill='gray', anchor='mm', font=font_small)
    draw.text((diagram_width // 2, diagram_height - 18), dim_txt, fill='black', anchor='mm', font=font_small)
    return img

def _door_only_unique_by_kind(doors):
    """Return one diagram per unique door size (dw, dh). Deduplicates by kind."""
    seen = set()
    out = []
    for door in (doors or []):
        size_str = door.get("size", "")
        dw, dh = _parse_door_size_inches(size_str)
        key = (round(dw, 1), round(dh, 1))
        if key not in seen:
            seen.add(key)
            label = f"{dw:.0f}\" x {dh:.0f}\""
            out.append((dw, dh, label))
    return out

def _add_door_only_diagrams_to_excel(ws, start_row, doors):
    """Add door-only diagram(s): one per unique kind, same column, stacked up/down, as close as possible."""
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
        import io
    except ImportError:
        return start_row + 2
    instances = _door_only_unique_by_kind(doors)
    if not instances:
        return start_row
    current_row = start_row
    for idx, (dw, dh, label) in enumerate(instances):
        if idx > 0:
            current_row += _DOOR_DIAGRAM_GAP_ROWS
        diagram_img = _create_door_only_diagram_single(dw, dh, label)
        if diagram_img:
            img_bytes = io.BytesIO()
            diagram_img.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            img = OpenpyxlImage(img_bytes)
            img.width = min(diagram_img.width, _DIAGRAM_MAX_WIDTH)
            img.height = int(diagram_img.height * (img.width / diagram_img.width))
            img.anchor = f'A{current_row}'
            ws.add_image(img)
            estimated_rows = max(12, int(img.height / _DIAGRAM_ROW_HEIGHT_PX))
            current_row += estimated_rows
        else:
            current_row += 2
    return current_row + _EXTRA_ROWS_AFTER_IMAGE

def _write_output_section(ws, title, items, colE, elevation_finish, system_total_ref, original_system_total_ref, start_output_row, current_extra_materials_state, extra_materials_path, multiplier, show_qty_per_elevation=False, total_count=1, show_total_cost_per_elevation=False, show_discounted_cost_per_elevation=False):
    """Writes a section of calculated outputs to the worksheet."""
    if not items: 
        return start_output_row, [], {'original': 0.0, 'discounted': 0.0}

    current_row = start_output_row
    title_cell = ws.cell(row=current_row, column=colE, value=title)
    title_cell.font = Font(bold=True, size=12)
    # title_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Removed color fill for professional look

    # Build headers based on which optional columns to show
    headers = ["Description", "Part Number", "Total Quantity Required"]
    if show_qty_per_elevation and total_count > 1:
        headers.append("Quantity Per Elevation")
    headers.append("Total List Cost")
    if show_total_cost_per_elevation and total_count > 1:
        headers.append("Total List Cost Per Elevation")
    headers.append("Discounted Total List Cost")
    if show_discounted_cost_per_elevation and total_count > 1:
        headers.append("Discounted Total List Cost Per Elevation")
    
    for i, h in enumerate(headers):
        header_cell = ws.cell(row=current_row + 1, column=colE + i, value=h)
        header_cell.font = Font(bold=True)
        header_cell.border = Border(bottom=Side(style='thin'))
        # header_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid") # Removed color fill for professional look
    current_row += 2

    section_material_impacts = []
    section_original_total = 0.0
    section_discounted_total = 0.0
    section_original_per_elev_total = 0.0  # Sum of actual per-elevation costs
    section_discounted_per_elev_total = 0.0  # Sum of actual per-elevation discounted costs

    for item in items:
        qty_raw = item.get('quantity', 0)
        pn, manual = item.get('part_number'), item.get('manual', False)
        desc = item.get('description', '').strip()
        is_profile = pn in PART_NUMBER_MAP.get('profiles', {})
        is_gasket = "gasket" in desc.lower() or pn in ["E2-0052", "E2-0053", "E2-0065"]
        is_accessory = pn in PART_NUMBER_MAP.get('accessories', {}) or item.get('type', '').lower() == 'accessory'
        is_glass = pn == "GLASS_AREA" or item.get('type', '').lower() == 'glass'

        # Determine if we should process as a group (for optimization)
        is_bay_width_item = _is_bay_width_part(pn, qty_raw, item.get('description', ''))
        is_list = isinstance(qty_raw, list)
        has_multiple_items = is_list and len(qty_raw) > 1
        
        # For profiles/gaskets with a list, ALWAYS process the entire list at once for optimization
        # This ensures proper leftover calculation across multiple cuts
        # CRITICAL: Check this BEFORE creating individual_quantities to avoid splitting the list
        should_process_as_group = (is_profile or is_gasket) and has_multiple_items
        
        # Calculate quantities for display and processing
        # Only split into individual_quantities if NOT processing as group
        if should_process_as_group:
            # Keep as list for group processing
            individual_quantities = qty_raw if is_list else [qty_raw]
            qty_sum = sum(qty_raw) if is_list else qty_raw
        else:
            # Split for individual processing
            individual_quantities = qty_raw if is_list else [qty_raw]
            qty_sum = sum(individual_quantities) if is_list else qty_raw

        unit_type = 'ft' if (is_profile or is_gasket) else 'pcs' if is_accessory else item.get('unit', 'pcs' if not is_glass else 'sqft')
        display_unit = unit_type

        # Format display string
        if is_list:
            if len(qty_raw) > 1 and all(x == qty_raw[0] for x in qty_raw):
                # For profiles, show as "8ft x 3" format, for others use decimal format
                if is_profile:
                    # Check if it's a whole number
                    if qty_raw[0] == int(qty_raw[0]):
                        display_qty_string = f"{int(qty_raw[0])}{display_unit} x {len(qty_raw)}"
                    else:
                        display_qty_string = f"{qty_raw[0]:.2f}{display_unit} x {len(qty_raw)}"
                else:
                    display_qty_string = f"{qty_raw[0]:.2f} {display_unit} x {len(qty_raw)}"
            else:
                # For profiles, show individual cuts without decimals when whole numbers, with decimals otherwise
                if is_profile:
                    display_qty_string = ", ".join([f"{int(q)}{display_unit}" if q == int(q) else f"{q:.2f}{display_unit}" for q in qty_raw])
                else:
                    display_qty_string = ", ".join([f"{q:.2f} {display_unit}" for q in qty_raw])
        else:
            # For profiles, show without decimals when whole number
            if is_profile and qty_raw == int(qty_raw):
                display_qty_string = f"{int(qty_raw)}{display_unit}"
            else:
                display_qty_string = f"{qty_raw:.2f} {display_unit}"

        item_total_cost_for_display = 0.0
        original_item_total_cost = 0.0
        
        # Process as group if it's a profile/gasket with multiple pieces
        if should_process_as_group:
            # Process the entire list as one request for waste optimization
            # Profiles and gaskets are treated as length-based items (sold by length, with leftover tracking)
            print(f"DEBUG: Processing {pn} as GROUP with list {qty_raw} (type: {type(qty_raw)}, is_list: {isinstance(qty_raw, list)})")
            use_group = is_profile or is_gasket
            # CRITICAL: Ensure qty_raw is passed as a list, not converted to a single value
            list_to_process = qty_raw if isinstance(qty_raw, list) else [qty_raw]
            print(f"DEBUG: Calling get_price_by_part with list: {list_to_process}")
            total_price, calculated_unit_type, material_impact_details = \
                get_price_by_part(pn, list_to_process, finish=elevation_finish, current_extra_materials=current_extra_materials_state, extra_materials_file=extra_materials_path, summary=False, group=use_group, description=item.get('description', ''))
            
            item_total_cost_for_display = total_price or 0.0
            original_item_total_cost = total_price or 0.0
            calculated_unit_type = unit_type if (is_profile or is_gasket or is_accessory) else (calculated_unit_type or item.get('unit', 'pcs'))
            
            if material_impact_details:
                leftover_qty = material_impact_details.get('leftover_generated_qty_or_length', 0.0)
                all_leftovers = material_impact_details.get('all_new_leftovers', [])
                print(f"DEBUG: Material impact for {pn}: leftover_qty={leftover_qty}, all_new_leftovers={all_leftovers}")
                material_impact_details['leftover_generated_qty_or_length_display'] = f"{leftover_qty:.2f} {display_unit}"
                section_material_impacts.append(material_impact_details)
                apply_material_impact_to_extra_materials_in_memory(current_extra_materials_state, material_impact_details)
            else:
                print(f"DEBUG: WARNING - No material_impact_details returned for {pn} with list {qty_raw}")
        else:
            # Standard processing: iterate through each quantity (for non-profile/gasket items or single quantities)
            for single_qty_for_calc in individual_quantities:
                total_item_price_single_cut, calculated_unit_type, material_impact_details = 0.0, unit_type, None

                if manual:
                    if pn and pn != "N/A":
                        price_calculated, unit_calculated, material_impact_details = \
                            get_price_by_part(pn, single_qty_for_calc, finish=elevation_finish, current_extra_materials=current_extra_materials_state, extra_materials_file=extra_materials_path, summary=False, group=True)
                        total_item_price_single_cut = (price_calculated if price_calculated is not None else item.get('price', 0.0) * single_qty_for_calc)
                        calculated_unit_type = unit_type if is_profile or is_accessory else (unit_calculated or item.get('unit', 'pcs'))
                    else:
                        total_item_price_single_cut = item.get('price', 0.0) * single_qty_for_calc
                        calculated_unit_type = item.get('unit', 'pcs')
                        material_impact_details = {
                            'part_number': "N/A - Manual", 'requested_qty': single_qty_for_calc, 'purchased_qty_or_length': 0.0,
                            'leftover_generated_qty_or_length': 0.0, 'used_from_leftover_qty_or_length': 0.0,
                            'cost_incurred': total_item_price_single_cut, 'type_processed_as': 'manual_no_pn',
                            'finish': None
                        }
                else:
                    # For profiles/gaskets, use group=True to ensure proper processing
                    # For others, use group only if it's a gasket
                    use_group = (is_profile or is_gasket)
                    total_price, unit_from_pricing, material_impact_details = \
                        get_price_by_part(pn, single_qty_for_calc, finish=elevation_finish, current_extra_materials=current_extra_materials_state, extra_materials_file=extra_materials_path, summary=False, group=use_group)
                    total_item_price_single_cut = total_price or 0.0
                    calculated_unit_type = unit_type if (is_profile or is_gasket or is_accessory) else (unit_from_pricing or item.get('unit', 'pcs'))

                item_total_cost_for_display += total_item_price_single_cut
                original_item_total_cost += total_item_price_single_cut

                if material_impact_details:
                    leftover_qty = material_impact_details.get('leftover_generated_qty_or_length', 0.0)
                    material_impact_details['leftover_generated_qty_or_length_display'] = f"{leftover_qty:.2f} {display_unit}"
                    section_material_impacts.append(material_impact_details)
                    apply_material_impact_to_extra_materials_in_memory(current_extra_materials_state, material_impact_details)

        if is_profile or is_gasket or is_accessory:
            item_total_cost_for_display *= multiplier
            if qty_sum > 0:
                item['price'] = item_total_cost_for_display / qty_sum

        system_total_ref[0] += item_total_cost_for_display
        original_system_total_ref[0] += original_item_total_cost
        section_original_total += original_item_total_cost
        section_discounted_total += item_total_cost_for_display

        # Calculate quantity per elevation if needed - show as "8ft x 2" format (pieces per elevation) for profiles/gaskets only
        qty_per_elev_display = None
        if show_qty_per_elevation and total_count > 1:
            # Only use "x" format for profiles and gaskets (they have specific cut dimensions)
            use_x_format = is_profile or is_gasket
            
            if isinstance(qty_raw, list):
                num_pieces = len(qty_raw)
                pieces_per_elev = num_pieces / total_count
                
                # If all pieces are the same length, show as "8ft x 2" for profiles/gaskets
                if len(qty_raw) > 1 and all(abs(x - qty_raw[0]) < 0.001 for x in qty_raw):
                    piece_length = qty_raw[0]
                    if use_x_format:
                        if is_profile:
                            if piece_length == int(piece_length):
                                qty_per_elev_display = f"{int(piece_length)}{display_unit} x {int(pieces_per_elev)}"
                            else:
                                qty_per_elev_display = f"{piece_length:.2f}{display_unit} x {int(pieces_per_elev)}"
                        else:  # is_gasket
                            qty_per_elev_display = f"{piece_length:.2f} {display_unit} x {int(pieces_per_elev)}"
                    else:
                        # For other items, just show the quantity per elevation without "x"
                        qty_per_elev_display = f"{pieces_per_elev:.2f} {display_unit}"
                else:
                    # Different lengths - group by length and show pieces per elevation for each
                    if use_x_format:
                        length_counts = Counter(qty_raw)
                        parts = []
                        for length, count in sorted(length_counts.items()):
                            pieces_per_length_per_elev = (count / total_count)
                            if is_profile:
                                if length == int(length):
                                    parts.append(f"{int(length)}{display_unit} x {int(pieces_per_length_per_elev)}")
                                else:
                                    parts.append(f"{length:.2f}{display_unit} x {int(pieces_per_length_per_elev)}")
                            else:  # is_gasket
                                parts.append(f"{length:.2f} {display_unit} x {int(pieces_per_length_per_elev)}")
                        qty_per_elev_display = ", ".join(parts)
                    else:
                        # For other items, just show total quantity per elevation
                        qty_per_elev = qty_sum / total_count
                        qty_per_elev_display = f"{qty_per_elev:.2f} {display_unit}"
            else:
                # Single quantity value
                qty_per_elev = qty_raw / total_count
                if use_x_format:
                    # For profiles/gaskets, show as "8ft x 1"
                    if is_profile:
                        if qty_raw == int(qty_raw):
                            qty_per_elev_display = f"{int(qty_raw)}{display_unit} x 1"
                        else:
                            qty_per_elev_display = f"{qty_raw:.2f}{display_unit} x 1"
                    else:  # is_gasket
                        qty_per_elev_display = f"{qty_raw:.2f} {display_unit} x 1"
                else:
                    # For other items, just show the quantity per elevation without "x"
                    qty_per_elev_display = f"{qty_per_elev:.2f} {display_unit}"
        
        # Write data columns
        col_offset = 0
        ws.cell(row=current_row, column=colE + col_offset, value=item.get('description', ''))
        col_offset += 1
        ws.cell(row=current_row, column=colE + col_offset, value=pn or 'N/A')
        col_offset += 1
        ws.cell(row=current_row, column=colE + col_offset, value=display_qty_string)
        col_offset += 1
        
        # Add quantity per elevation column if enabled
        if qty_per_elev_display:
            ws.cell(row=current_row, column=colE + col_offset, value=qty_per_elev_display)
            col_offset += 1
        
        # Total List Cost
        ws.cell(row=current_row, column=colE + col_offset, value=original_item_total_cost).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        col_offset += 1
        
        # Total List Cost Per Elevation (if enabled) - calculate actual purchase cost for per-elevation quantity
        original_cost_per_elev = None
        if show_total_cost_per_elevation and total_count > 1:
            # Calculate the quantity per elevation
            if isinstance(qty_raw, list):
                # For lists, calculate pieces per elevation
                num_pieces = len(qty_raw)
                pieces_per_elev = num_pieces / total_count
                
                # Check if this is a bay width part - if so, keep as list
                is_bay_width_for_per_elev = _is_bay_width_part(pn, qty_raw, item.get('description', ''))
                
                if is_bay_width_for_per_elev:
                    # For bay width parts, take first pieces_per_elev items
                    pieces_per_elev_int = int(pieces_per_elev)
                    if pieces_per_elev_int > 0 and pieces_per_elev_int <= len(qty_raw):
                        qty_per_elev = qty_raw[:pieces_per_elev_int]
                    else:
                        # Fallback: divide each piece equally (not ideal but works)
                        qty_per_elev = [q / total_count for q in qty_raw]
                else:
                    # For non-bay-width parts, if all pieces are the same, use single value
                    if len(qty_raw) > 0 and all(abs(x - qty_raw[0]) < 0.001 for x in qty_raw):
                        # All pieces same length, per elevation is just one piece
                        qty_per_elev = qty_raw[0]
                    else:
                        # Different lengths - sum and divide
                        qty_per_elev = qty_sum / total_count
            else:
                # Single quantity value
                qty_per_elev = qty_raw / total_count
            
            # Calculate actual purchase cost for per-elevation quantity (accounts for minimum purchase lengths)
            if not manual and pn and pn != "N/A":
                use_group_for_per_elev = is_gasket
                
                # Get price for per-elevation quantity
                per_elev_price, _, _ = get_price_by_part(
                    pn, qty_per_elev, 
                    finish=elevation_finish, 
                    current_extra_materials=current_extra_materials_state, 
                    extra_materials_file=extra_materials_path, 
                    summary=True,  # Use summary=True to get price without material impact
                    group=use_group_for_per_elev,
                    description=item.get('description', '')
                )
                original_cost_per_elev = per_elev_price if per_elev_price is not None else (original_item_total_cost / total_count)
            else:
                # For manual items, just divide the cost
                original_cost_per_elev = original_item_total_cost / total_count
            
            ws.cell(row=current_row, column=colE + col_offset, value=original_cost_per_elev).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            section_original_per_elev_total += original_cost_per_elev  # Track sum of per-elevation costs
            col_offset += 1
        
        # Discounted Total List Cost
        ws.cell(row=current_row, column=colE + col_offset, value=item_total_cost_for_display).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        col_offset += 1
        
        # Discounted Total List Cost Per Elevation (if enabled) - apply multiplier to per-elevation cost
        if show_discounted_cost_per_elevation and total_count > 1:
            if original_cost_per_elev is not None:
                # Use the calculated per-elevation cost and apply multiplier
                discounted_cost_per_elev = original_cost_per_elev * multiplier if (is_profile or is_gasket or is_accessory) else original_cost_per_elev
            else:
                # Fallback: recalculate per-elevation quantity and price (shouldn't happen if original_cost_per_elev was calculated)
                if isinstance(qty_raw, list):
                    num_pieces = len(qty_raw)
                    pieces_per_elev = num_pieces / total_count
                    is_bay_width_for_per_elev = _is_bay_width_part(pn, qty_raw, item.get('description', ''))
                    
                    if is_bay_width_for_per_elev:
                        pieces_per_elev_int = int(pieces_per_elev)
                        if pieces_per_elev_int > 0 and pieces_per_elev_int <= len(qty_raw):
                            qty_per_elev = qty_raw[:pieces_per_elev_int]
                        else:
                            qty_per_elev = [q / total_count for q in qty_raw]
                    else:
                        if len(qty_raw) > 0 and all(abs(x - qty_raw[0]) < 0.001 for x in qty_raw):
                            qty_per_elev = qty_raw[0]
                        else:
                            qty_per_elev = qty_sum / total_count
                else:
                    qty_per_elev = qty_raw / total_count
                
                if not manual and pn and pn != "N/A":
                    use_group_for_per_elev = is_gasket
                    per_elev_price, _, _ = get_price_by_part(
                        pn, qty_per_elev,
                        finish=elevation_finish,
                        current_extra_materials=current_extra_materials_state,
                        extra_materials_file=extra_materials_path,
                        summary=True,
                        group=use_group_for_per_elev,
                        description=item.get('description', '')
                    )
                    discounted_cost_per_elev = (per_elev_price * multiplier) if (is_profile or is_gasket or is_accessory) and per_elev_price is not None else (item_total_cost_for_display / total_count)
                else:
                    discounted_cost_per_elev = item_total_cost_for_display / total_count
            ws.cell(row=current_row, column=colE + col_offset, value=discounted_cost_per_elev).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            section_discounted_per_elev_total += discounted_cost_per_elev  # Track sum of per-elevation discounted costs
            col_offset += 1
        
        current_row += 1

    # Add Section Totals
    # Convert title to proper label (e.g., "PROFILES" -> "Profile Cost")
    title_mapping = {
        "PROFILES": "Profile",
        "ACCESSORIES": "Accessory",
        "GASKETS": "Gasket",
        "DOORS": "Door",
        "GLASS": "Glass",
        "LABOR": "Labor"
    }
    title_label = title_mapping.get(title.upper(), title.title())
    total_label = f"Total {title_label} Cost"
    
    # Calculate column offsets based on which optional columns are shown
    total_col_offset = 2  # Start after "Total Quantity Required"
    if show_qty_per_elevation and total_count > 1:
        total_col_offset += 1  # Skip "Quantity Per Elevation" column
    
    # Place total label in the "Total List Cost" column
    ws.cell(row=current_row, column=colE + total_col_offset, value=total_label).font = Font(bold=True)
    total_col_offset += 1
    
    # Total List Cost
    ws.cell(row=current_row, column=colE + total_col_offset, value=section_original_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=current_row, column=colE + total_col_offset).font = Font(bold=True)
    total_col_offset += 1
    
    # Total List Cost Per Elevation (if enabled) - use sum of actual per-elevation costs
    if show_total_cost_per_elevation and total_count > 1:
        ws.cell(row=current_row, column=colE + total_col_offset, value=section_original_per_elev_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=current_row, column=colE + total_col_offset).font = Font(bold=True)
        total_col_offset += 1
    
    # Discounted Total List Cost
    ws.cell(row=current_row, column=colE + total_col_offset, value=section_discounted_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=current_row, column=colE + total_col_offset).font = Font(bold=True)
    total_col_offset += 1
    
    # Discounted Total List Cost Per Elevation (if enabled) - use sum of actual per-elevation discounted costs
    if show_discounted_cost_per_elevation and total_count > 1:
        ws.cell(row=current_row, column=colE + total_col_offset, value=section_discounted_per_elev_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=current_row, column=colE + total_col_offset).font = Font(bold=True)
        total_col_offset += 1
    
    # Add top border for totals row - adjust range based on number of columns
    num_cols = 5  # Base columns: Description, Part Number, Total Quantity Required, Total List Cost, Discounted Total List Cost
    if show_qty_per_elevation and total_count > 1:
        num_cols += 1
    if show_total_cost_per_elevation and total_count > 1:
        num_cols += 1
    if show_discounted_cost_per_elevation and total_count > 1:
        num_cols += 1
    for col in range(colE, colE + num_cols):
        ws.cell(row=current_row, column=col).border = Border(top=Side(style='thin'))

    # Return section totals for summary
    section_totals = {
        'original': section_original_total,
        'discounted': section_discounted_total
    }
    return current_row + 1, section_material_impacts, section_totals


def create_summary_sheet(ws, elevations_json_path, extra_materials_json_path, wb=None, summary_settings_path=None):
    """
    Reads elevation data, aggregates quantities and prices by part number across all elevations,
    and writes a clean summary section into the worksheet, grouped by profiles, accessories, doors, glass, and labor.
    """
    try:
        with open(elevations_json_path, 'r') as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"[WARNING] Could not load elevations JSON: {e}")
        return

    # For summary, use a shared in-memory state that accumulates leftovers across all elevations
    # This allows the summary to utilize waste materials across all elevations
    summary_extra_materials_state = {}
    
    try:
        extra_materials = load_extra_materials(extra_materials_json_path)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"[WARNING] Could not load extra materials JSON: {e}")
        extra_materials = {}

    if not data:
        print("[INFO] No data found, summary cleared if existed.")
        return

    # --- Elevation summary defaults (can be overridden later when settings are loaded) ---
    elevation_summary_settings = {
        "show_elevation_names": False,
        "show_elevation_quantity": False,
        "show_elevation_dimensions": False,
        "show_elevation_sqft": False,
        "show_elevation_perimeter": False,
    }
    elevation_summary_cols = []
    category_start_col = 1  # Profiles section still starts at column 1 by default

    # Step 1: Calculate full_running_grand_total for multiplier
    full_running_grand_total = 0.0
    for elev_key, elev in data.items():
        elevation_finish = elev.get('finish', '').lower()
        for output in elev.get('calculated_outputs', []):
            qty = output.get('quantity', 0)
            qty = sum(qty) if isinstance(qty, list) else qty
            manual = output.get('manual', False)
            part = output.get('part_number', '').strip()
            item_type = output.get('type', '').lower()
            price = 0.0
            if manual or part == "GLASS_AREA" or item_type in ['glass', 'joints_fab_labor', 'door', 'doors']:
                price = output.get('price', 0.0) * qty
            else:
                price, _, _ = get_price_by_part(
                    part,
                    qty,
                    finish=elevation_finish,
                    extra_materials_file=extra_materials_json_path,
                    summary=True,
                    group=True if manual else False
                )
                price = price if price is not None else 0.0
            full_running_grand_total += price

    multiplier = _get_multiplier(full_running_grand_total)

    # Step 2: Aggregate quantities and prices across all elevations, grouped by category
    categories = {
        'PROFILES': [],
        'ACCESSORIES': [],
        'GASKETS': [],
        'DOORS': [],
        'GLASS': [],
        'LABOR': []
    }

    for elev_key, elev in data.items():
        elevation_finish = elev.get('finish', '').lower()
        for output in elev.get('calculated_outputs', []):
            part = output.get('part_number', '').strip()
            desc = output.get('description', '').strip()
            manual = output.get('manual', False)
            qty = output.get('quantity', 0)
            qty_for_aggregation = sum(qty) if isinstance(qty, list) else qty
            is_profile = part in PART_NUMBER_MAP.get('profiles', {})
            is_gasket = "gasket" in desc.lower() or part in ["E2-0052", "E2-0053", "E2-0065"]  # Identify gaskets
            is_accessory = part in PART_NUMBER_MAP.get('accessories', {}) or output.get('type', '').lower() == 'accessory'
            is_glass = part == "GLASS_AREA" or output.get('type', '').lower() == 'glass'
            is_joints_fab_labor = part == "JOINTS_FAB_LABOR" or output.get('type', '').lower() == 'joints_fab_labor' or "joints fabrication" in desc.lower() or "fabrication labor" in desc.lower()
            is_door = output.get('type', '').lower() in ['door', 'doors']

            if is_profile:
                category = 'PROFILES'
            elif is_gasket:
                category = 'GASKETS'
            elif is_accessory:
                category = 'ACCESSORIES'
            elif is_door:
                category = 'DOORS'
            elif is_glass:
                category = 'GLASS'
            elif is_joints_fab_labor:
                category = 'LABOR'
            else:
                continue

            if manual or is_glass or is_joints_fab_labor or is_door:
                if part and part != "N/A":
                    key = f"MANUAL_{part}-{elevation_finish}" if (is_profile or is_gasket or is_joints_fab_labor or is_door or is_glass) and elevation_finish else f"MANUAL_{part}"
                    display = f"{desc} ({part} - {elevation_finish})" if (is_profile or is_gasket or is_joints_fab_labor or is_door or is_glass) and elevation_finish else f"{desc} ({part})"
                else:
                    key = f"MANUAL_NO_PN_{desc}"
                    display = desc
            else:
                if (is_profile or is_gasket or is_joints_fab_labor or is_door or is_glass) and elevation_finish:
                    key = f"{part}-{elevation_finish}"
                    display = f"{part} ({elevation_finish})"
                else:
                    key = part
                    display = part

            categories[category].append({
                'key': key,
                'quantity': qty_for_aggregation,
                'quantity_list': qty if isinstance(qty, list) else [qty],  # Preserve individual quantities
                'description': desc,
                'display': display,
                'part_number': part,
                'manual': manual,
                'unit': 'ft' if (is_profile or is_gasket) else 'pcs' if is_accessory else output.get('unit', 'pcs' if not is_glass else 'sqft'),
                'finish': elevation_finish if (is_profile or is_gasket or is_joints_fab_labor or is_door or is_glass) else '',
                'is_glass': is_glass,
                'is_joints_fab_labor': is_joints_fab_labor,
                'is_door': is_door,
                'is_gasket': is_gasket,
                'price': output.get('price', 0.0) if (manual or is_glass or is_joints_fab_labor or is_door) else 0.0
            })

    # Step 2.5: Aggregate items within each category by key to prevent duplicates across elevations
    for category in categories:
        aggregated_map = {}
        for item in categories[category]:
            k = item['key']
            if k in aggregated_map:
                existing = aggregated_map[k]
                # Preserve description if missing
                if not existing.get('description') and item.get('description'):
                    existing['description'] = item['description']
                # If price is manually set (manual, glass, labor, door), maintain correct total cost by updating unit price
                if existing['manual'] or existing['is_glass'] or existing['is_joints_fab_labor'] or existing['is_door']:
                    # Recalculate weighted average price
                    cost_existing = float(existing.get('price', 0.0)) * float(existing['quantity'])
                    cost_new = float(item.get('price', 0.0)) * float(item['quantity'])
                    total_qty = float(existing['quantity']) + float(item['quantity'])
                    existing['quantity'] = total_qty
                    # Combine quantity lists
                    existing_qty_list = existing.get('quantity_list', [])
                    item_qty_list = item.get('quantity_list', [])
                    if not existing_qty_list:
                        existing_qty_list = [existing['quantity']]
                    if not item_qty_list:
                        item_qty_list = [item['quantity']]
                    existing['quantity_list'] = existing_qty_list + item_qty_list
                    if total_qty > 0:
                        existing['price'] = (cost_existing + cost_new) / total_qty
                else:
                    # For standard parts, just sum quantity; price is re-calculated in Step 3
                    existing['quantity'] = float(existing['quantity']) + float(item['quantity'])
                    # Combine quantity lists
                    existing_qty_list = existing.get('quantity_list', [])
                    item_qty_list = item.get('quantity_list', [])
                    if not existing_qty_list:
                        existing_qty_list = [existing['quantity']]
                    if not item_qty_list:
                        item_qty_list = [item['quantity']]
                    existing['quantity_list'] = existing_qty_list + item_qty_list
            else:
                # Ensure quantity is float
                item['quantity'] = float(item['quantity'])
                if 'quantity_list' not in item:
                    item['quantity_list'] = [item['quantity']]
                # Ensure description exists
                if 'description' not in item or not item.get('description'):
                    item['description'] = item.get('display', '')
                aggregated_map[k] = item
        categories[category] = list(aggregated_map.values())

    # Step 3: Calculate prices for aggregated items and prepare final data
    final_summary_data = []
    total_discounted_price = 0.0
    total_reusable_cost = 0.0
    grand_original_total = 0.0
    grand_discounted_total = 0.0
    grand_residual_total = 0.0

    for category, items in categories.items():
        for item in items:
            key = item['key']
            quantity_aggregated = item['quantity']
            manual = item['manual']
            part = item['part_number']
            display = item['display']
            description = item.get('description', '') or display
            is_profile = part in PART_NUMBER_MAP.get('profiles', {})
            is_accessory = part in PART_NUMBER_MAP.get('accessories', {}) or item.get('type', '').lower() == 'accessory'
            is_gasket = item.get('is_gasket', False) or "gasket" in item.get('description', '').lower() or part in ["E2-0052", "E2-0053", "E2-0065"]
            is_glass = item['is_glass']
            is_joints_fab_labor = item['is_joints_fab_labor']
            is_door = item['is_door']
            item_finish = item['finish']

            display_unit = 'ft' if (is_profile or is_gasket) else 'pcs' if is_accessory else item['unit']
            original_total_cost_for_item = 0.0
            total_cost_for_item = 0.0
            calculated_unit_type = display_unit
            reusable_qty_sum = 0.0
            reusable_pct = 0.0
            reusable_cost = 0.0
            reusable_qty_display_string = "N/A"

            if manual or is_glass or is_joints_fab_labor or is_door:
                price = float(item.get('price', 0.0))
                qty_float = float(quantity_aggregated)
                original_total_cost_for_item = price * qty_float
                calculated_unit_type = item['unit'] or ('sqft' if is_glass else 'pcs')
            else:
                # Gaskets are treated as profiles (sold by length, with leftover tracking)
                use_group = is_gasket
                
                # For profiles and gaskets, use quantity_list if available for optimal cutting
                # This allows the pricing function to optimize cuts across multiple pieces
                quantity_for_pricing = quantity_aggregated
                if (is_profile or is_gasket) and part and part != "N/A":
                    quantity_list = item.get('quantity_list', [])
                    if quantity_list and len(quantity_list) > 1:
                        # Use the list for optimization - filter valid values
                        valid_quantities = [q for q in quantity_list if q is not None and isinstance(q, (int, float)) and q > 0]
                        if valid_quantities:
                            quantity_for_pricing = valid_quantities
                
                # Use shared in-memory state for summary to accumulate leftovers across all elevations
                total_price, unit_type_from_pricing, material_impact = get_price_by_part(
                    part,
                    quantity_for_pricing,
                    finish=item_finish,
                    current_extra_materials=summary_extra_materials_state,
                    extra_materials_file=extra_materials_json_path,
                    summary=False,  # Use False to actually use and track leftovers
                    group=use_group
                )
                # Apply material impact to accumulate leftovers in summary state
                if material_impact:
                    apply_material_impact_to_extra_materials_in_memory(summary_extra_materials_state, material_impact)
                original_total_cost_for_item = total_price if total_price is not None else 0.0
                calculated_unit_type = 'ft' if is_profile else 'pcs' if is_accessory else (unit_type_from_pricing or item['unit'] or 'pcs')

            if is_profile or is_gasket or is_accessory:
                total_cost_for_item = original_total_cost_for_item * multiplier
            else:
                total_cost_for_item = original_total_cost_for_item

            total_discounted_price += total_cost_for_item

            if part and part != "N/A" and (is_profile or is_gasket or is_accessory):
                extra_materials_key_for_reuse = part
                if (is_profile or is_gasket) and item_finish:
                    extra_materials_key_for_reuse = f"{part}-{item_finish}"

                # Use the shared summary state that accumulates leftovers across all elevations
                part_data = summary_extra_materials_state.get(extra_materials_key_for_reuse, {})
                if part_data.get("length_pieces"):
                    # Get min_purchase_length for validation
                    part_info = parts_data.get(part, {})
                    length_str = part_info.get('Length', '')
                    min_purchase_length = parse_length_to_feet(length_str) or 24.0
                    
                    # Filter out invalid leftover pieces (must be > 0 and < min_purchase_length)
                    lengths = [float(x) for x in part_data["length_pieces"] if isinstance(x, (int, float, str))]
                    valid_lengths = [l for l in lengths if l > EPSILON and l < min_purchase_length - EPSILON]
                    reusable_qty_sum = sum(valid_lengths)
                    if valid_lengths:
                        # Count occurrences of each length, using rounded values for grouping
                        counter = Counter([round(l, 2) for l in valid_lengths])
                        reuse_lengths_formatted = []
                        for length_val, count in sorted(counter.items(), key=lambda x: x[0], reverse=True):
                            # Format length: use integer if whole number, otherwise 2 decimals
                            if length_val == int(length_val):
                                length_str = f"{int(length_val)}{display_unit}"
                            else:
                                length_str = f"{length_val:.2f}{display_unit}"
                            # Always show count, even if it's 1
                            reuse_lengths_formatted.append(f"{length_str} x{count}")
                        reusable_qty_display_string = ", ".join(reuse_lengths_formatted)
                    else:
                        reusable_qty_sum = 0.0
                        reusable_qty_display_string = "N/A"
                else:
                    reusable_qty_sum = part_data.get("quantity", 0.0)
                    reusable_qty_display_string = f"{float(reusable_qty_sum):.2f} {display_unit}"

                try:
                    reusable_qty_sum = float(reusable_qty_sum)
                except (TypeError, ValueError):
                    reusable_qty_sum = 0.0

                try:
                    quantity_aggregated_f = float(quantity_aggregated)
                except (TypeError, ValueError):
                    quantity_aggregated_f = 0.0

                if reusable_qty_sum > 0 and quantity_aggregated_f > 0:
                    reusable_pct = min((reusable_qty_sum / (quantity_aggregated_f + reusable_qty_sum)) * 100, 100.0)
                else:
                    reusable_pct = 0.0

                unit_price_for_reuse, unit_type_for_reusable_calc = get_unit_price_by_part(
                    part, finish=item_finish, extra_materials_file=extra_materials_json_path
                )
                reusable_cost = (
                    reusable_qty_sum * unit_price_for_reuse * multiplier
                    if unit_price_for_reuse is not None
                    else 0.0
                )
                total_reusable_cost += reusable_cost
            # Calculate Quantity Req (FT) and Qty Stick/Roll (Req)
            quantity_req_ft = "N/A"
            qty_stick_req = "N/A"
            quantity_display_formatted = f"{quantity_aggregated:.2f} {display_unit}"
            
            if (is_profile or is_gasket) and part and part != "N/A":
                # For profiles and gaskets, quantity is in feet - add units
                quantity_req_ft = f"{quantity_aggregated:.2f} ft"
                # Calculate number of sticks/rolls needed and format with length
                part_data = parts_data.get(part, {})
                length_str = part_data.get('Length', '')
                min_purchase_length = parse_length_to_feet(length_str) or 1.0
                if min_purchase_length > 0:
                    num_units = math.ceil(quantity_aggregated / min_purchase_length)
                    unit_label = "rolls" if is_gasket else "sticks"
                    qty_stick_req = f"{num_units} ({min_purchase_length:.0f}ft per)"
                else:
                    qty_stick_req = "N/A"
                
                # Format quantity_display to show breakdown like "16ft x2, 8ft x1"
                quantity_list = item.get('quantity_list', [quantity_aggregated])
                # Filter out invalid values (None, empty, non-numeric)
                valid_quantities = [q for q in quantity_list if q is not None and (isinstance(q, (int, float)) and q > 0)]
                
                # If no valid quantities, use aggregated quantity
                if not valid_quantities:
                    valid_quantities = [quantity_aggregated] if quantity_aggregated > 0 else []
                
                # Count occurrences of each length
                if valid_quantities:
                    length_counter = Counter([round(q, 2) for q in valid_quantities])
                    if len(length_counter) > 1:
                        # Multiple different lengths - show breakdown
                        length_parts = []
                        for length_val, count in sorted(length_counter.items(), key=lambda x: x[0], reverse=True):
                            if count > 1:
                                length_parts.append(f"{length_val:.0f}ft x{count}")
                            else:
                                length_parts.append(f"{length_val:.0f}ft x1")
                        quantity_display_formatted = ", ".join(length_parts)
                    elif len(length_counter) == 1:
                        # All same length - show total with count if multiple pieces
                        length_val = list(length_counter.keys())[0]
                        count = length_counter[length_val]
                        if count > 1:
                            quantity_display_formatted = f"{length_val:.0f}ft x{count}"
                        else:
                            quantity_display_formatted = f"{length_val:.0f}ft"
                    else:
                        # Empty counter - fallback to default format
                        quantity_display_formatted = f"{quantity_aggregated:.2f} {display_unit}"
                else:
                    # No valid quantities - fallback to default format
                    quantity_display_formatted = f"{quantity_aggregated:.2f} {display_unit}"
                    
            elif is_accessory and part and part != "N/A":
                # For accessories, get bulk order info
                part_data = parts_data.get(part, {})
                units_str = part_data.get('Units', '1 pcs.')
                length_str = part_data.get('Length', '')
                
                # Check if sold by length (feet) or by pieces
                length_ft = parse_length_to_feet(length_str) if length_str else 0.0
                
                if length_ft > 1.0:
                    # Sold by length (e.g., glazing gasket: 500 ft per order)
                    unit_count_per_bundle = length_ft
                    unit_label = "ft per"
                else:
                    # Sold by pieces
                    unit_count_per_bundle = 1
                    if 'pc' in units_str.lower():
                        try:
                            unit_count_per_bundle = int(units_str.lower().split('pc')[0].strip()) or 1
                        except ValueError:
                            unit_count_per_bundle = 1
                    unit_label = "pcs per"
                
                # Column 2: Quantity (total pieces/feet required)
                quantity_req_ft = f"{quantity_aggregated:.2f} {display_unit}"
                
                # Column 3: Quantity Per Order (e.g., "500 ft per" or "20 pcs per")
                qty_stick_req = f"{unit_count_per_bundle:.0f} {unit_label}"
                
                # Column 4: Orders Required (number of orders needed)
                num_orders = math.ceil(quantity_aggregated / unit_count_per_bundle) if unit_count_per_bundle > 0 else 0
                quantity_display_formatted = f"{num_orders} order{'s' if num_orders != 1 else ''}"
            else:
                # For other items (glass, labor, doors), show N/A in first column, unit price in second
                quantity_req_ft = "N/A"
                
                # For the second column, show unit price
                if quantity_aggregated > 0:
                    unit_price = original_total_cost_for_item / quantity_aggregated
                    qty_stick_req = f"${unit_price:.2f}"
                else:
                    qty_stick_req = "$0.00"
            
            final_summary_data.append({
                'category': category,
                'description': description,
                'display': display,
                'quantity_display': quantity_display_formatted,
                'quantity_req_ft': quantity_req_ft,
                'qty_stick_req': qty_stick_req,
                'original_total_cost': original_total_cost_for_item,
                'total_cost': total_cost_for_item,
                'reusable_qty_display': reusable_qty_display_string,
                'reusable_pct': reusable_pct if (is_profile or is_gasket or is_accessory) else "N/A",
                'reusable_cost': reusable_cost if (is_profile or is_gasket or is_accessory) else 0.0,
                'part': part,
                'calculated_unit_type': calculated_unit_type
            })

    # Step 4: Write to worksheet with grouped sections
    start_row = 1
    current_row = start_row
    
    # Define headers based on category
    def get_headers_for_category(category, items_list=None):
        if category == 'PROFILES':
            return [
                "Description", "Project Total Materials", "Total Feet Required", "Sticks Required", "Total Quantity Required", "Total List Cost", "Discounted Total List Cost",
                "Residual Material Quantity", "Residual Waste %", "Residual Material Cost"
            ]
        elif category == 'ACCESSORIES':
            return [
                "Description", "Project Total Materials", "Total Pieces Required", "Quantity Per Order", "Orders Required", "Total List Cost", "Discounted Total List Cost",
                "Residual Material Quantity", "Residual Waste %", "Residual Material Cost"
            ]
        elif category == 'GASKETS':
            return [
                "Description", "Project Total Materials", "Total Feet Required", "Rolls Required", "Total Quantity Required", "Total List Cost", "Discounted Total List Cost",
                "Residual Material Quantity", "Residual Waste %", "Residual Material Cost"
            ]
        elif category == 'GLASS':
            return [
                "Description", "Project Total Materials", "N/A", "Unit Price", "Total Quantity Required", "Total List Cost", "Discounted Total List Cost",
                "Residual Material Quantity", "Residual Waste %", "Residual Material Cost"
            ]
        elif category == 'LABOR':
            return [
                "Description", "Project Total Materials", "N/A", "Unit Price", "Total Quantity Required", "Total List Cost", "Discounted Total List Cost",
                "Residual Material Quantity", "Residual Waste %", "Residual Material Cost"
            ]
        elif category == 'DOORS':
            return [
                "Description", "Project Total Materials", "N/A", "Unit Price", "Total Quantity Required", "Total List Cost", "Discounted Total List Cost",
                "Residual Material Quantity", "Residual Waste %", "Residual Material Cost"
            ]
        else:
            # Default headers
            return [
                "Description", "Project Total Materials", "Quantity Req (FT)", "Qty Stick (Req)", "Total Quantity Required", "Total List Cost", "Discounted Total List Cost",
                "Residual Material Quantity", "Residual Waste %", "Residual Material Cost"
            ]
    
    # Track category totals for markup calculations (initialize before loop)
    category_totals = {
        "PROFILES": {"discounted": 0.0, "residual": 0.0},
        "ACCESSORIES": {"discounted": 0.0, "residual": 0.0},
        "GASKETS": {"discounted": 0.0, "residual": 0.0},
        "DOORS": {"discounted": 0.0, "residual": 0.0},
        "GLASS": {"discounted": 0.0, "residual": 0.0},
        "LABOR": {"discounted": 0.0, "residual": 0.0}
    }

    for category, items in categories.items():
        if not items:
            continue
        headers = get_headers_for_category(category, items)
        header_cell = ws.cell(row=current_row, column=category_start_col, value=category)
        header_cell.font = Font(bold=True, size=12)
        # header_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Removed color fill for professional look
        current_row += 1
        for col, header in enumerate(headers, start=category_start_col):
            header_cell = ws.cell(row=current_row, column=col, value=header)
            header_cell.font = Font(bold=True)
            header_cell.border = Border(bottom=Side(style='thin'))
            # header_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid") # Removed color fill for professional look
        current_row += 1
        
        section_original_total = 0.0
        section_total_cost = 0.0
        section_residual_total = 0.0
        
        for item in final_summary_data:
            if item['category'] == category:
                section_original_total += item['original_total_cost']
                section_total_cost += item['total_cost']
                section_residual_total += item['reusable_cost']
                
                # Get description - it should be in final_summary_data
                description_value = item.get('description', '')
                if not description_value or description_value == '':
                    # Fallback to display if description is missing
                    description_value = item.get('display', '')
                ws.cell(row=current_row, column=category_start_col, value=description_value)
                ws.cell(row=current_row, column=category_start_col + 1, value=item['display'])
                ws.cell(row=current_row, column=category_start_col + 2, value=item['quantity_req_ft'])
                ws.cell(row=current_row, column=category_start_col + 3, value=item['qty_stick_req'])
                ws.cell(row=current_row, column=category_start_col + 4, value=item['quantity_display'])
                ws.cell(row=current_row, column=category_start_col + 5, value=item['original_total_cost']).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                ws.cell(row=current_row, column=category_start_col + 6, value=item['total_cost']).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                ws.cell(row=current_row, column=category_start_col + 7, value=item['reusable_qty_display'])
                ws.cell(row=current_row, column=category_start_col + 8, value=f"{item['reusable_pct']:.2f}%" if isinstance(item['reusable_pct'], (int, float)) else item['reusable_pct'])
                ws.cell(row=current_row, column=category_start_col + 9, value=item['reusable_cost']).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                current_row += 1
        
        grand_original_total += section_original_total
        grand_discounted_total += section_total_cost
        grand_residual_total += section_residual_total
        
        # Track category totals for markup calculations
        if category in category_totals:
            category_totals[category]["discounted"] = section_total_cost
            category_totals[category]["residual"] = section_residual_total

        # Add Section Totals for Summary
        # Convert category to proper label (e.g., "PROFILES" -> "Profile Cost")
        category_mapping = {
            "PROFILES": "Profile",
            "ACCESSORIES": "Accessory",
            "GASKETS": "Gasket",
            "DOORS": "Door",
            "GLASS": "Glass",
            "LABOR": "Labor"
        }
        category_label = category_mapping.get(category.upper(), category.title())
        total_label = f"Total {category_label} Cost"
        ws.cell(row=current_row, column=category_start_col + 4, value=total_label).font = Font(bold=True)
        ws.cell(row=current_row, column=category_start_col + 5, value=section_original_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=current_row, column=category_start_col + 5).font = Font(bold=True)
        ws.cell(row=current_row, column=category_start_col + 6, value=section_total_cost).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=current_row, column=category_start_col + 6).font = Font(bold=True)
        ws.cell(row=current_row, column=category_start_col + 9, value=section_residual_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=current_row, column=category_start_col + 9).font = Font(bold=True)
        
        # Add top border for totals row
        for col in range(category_start_col, category_start_col + 10):
            ws.cell(row=current_row, column=col).border = Border(top=Side(style='thin'))

        current_row += 2
    
    # ============================================================================
    # SUMMARY SECTIONS - Clean Professional Layout
    # ============================================================================
    gt_row = current_row + 2
    
    # Calculate discounted total from column G (adjusted for elevation summary offset)
    discounted_total_col = category_start_col + 6  # Column G was 7, now shifted
    sum_from_column_g = 0.0
    try:
        for row in range(1, gt_row):
            label_cell = ws.cell(row=row, column=category_start_col + 4)  # Column E was 5, now shifted
            value_cell = ws.cell(row=row, column=discounted_total_col)
            if label_cell.value and isinstance(label_cell.value, str):
                if "Total" in label_cell.value and "Cost" in label_cell.value:
                    if value_cell.value is not None:
                        try:
                            sum_from_column_g += float(value_cell.value)
                            print(f"Found section total '{label_cell.value}' in row {row}, column 7: ${value_cell.value}")
                        except (ValueError, TypeError):
                            pass
    except Exception as e:
        print(f"Error reading from column G: {e}")
        sum_from_column_g = 0.0
    
    final_discounted_total = sum_from_column_g if sum_from_column_g > 0 else grand_discounted_total
    if sum_from_column_g > 0:
        print(f"Summary discounted total from column G: ${sum_from_column_g:.2f}, calculated: ${grand_discounted_total:.2f}, using: ${final_discounted_total:.2f}")
    
    reuse_total = total_reusable_cost
    reuse_pct_of_gt = min((total_reusable_cost / total_discounted_price * 100) if total_discounted_price > 0 else 0.0, 100.0)
    
    # ============================================================================
    # COST OVERVIEW BOX - Spans columns (shifted right if elevation summary exists)
    # ============================================================================
    # Add a bit of vertical spacing after the elevation summary / categories
    # so the "TOTAL" cost box doesn't visually overlap the elevation list.
    overview_start_row = gt_row + 4
    overview_start_col = category_start_col
    
    # Header row with background
    ws.cell(row=overview_start_row, column=overview_start_col, value="COST OVERVIEW").font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=overview_start_row, column=overview_start_col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.cell(row=overview_start_row, column=overview_start_col).border = Border(left=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='thin'))
    ws.cell(row=overview_start_row, column=overview_start_col + 1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.cell(row=overview_start_row, column=overview_start_col + 1).border = Border(top=Side(style='medium'), bottom=Side(style='thin'))
    ws.cell(row=overview_start_row, column=overview_start_col + 2).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.cell(row=overview_start_row, column=overview_start_col + 2).border = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='thin'))
    
    # List Price row (ensure values are numeric for display)
    grand_original_safe = grand_original_total if grand_original_total is not None else 0.0
    final_discounted_safe = final_discounted_total if final_discounted_total is not None else 0.0
    ws.cell(row=overview_start_row+1, column=overview_start_col, value="List Price Total:")
    ws.cell(row=overview_start_row+1, column=overview_start_col).border = Border(left=Side(style='medium'))
    ws.cell(row=overview_start_row+1, column=overview_start_col + 2, value=grand_original_safe).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=overview_start_row+1, column=overview_start_col + 2).alignment = Alignment(horizontal='right')
    ws.cell(row=overview_start_row+1, column=overview_start_col + 2).border = Border(right=Side(style='medium'))
    
    # Discounted Total row
    ws.cell(row=overview_start_row+2, column=overview_start_col, value="Discounted Total:")
    ws.cell(row=overview_start_row+2, column=overview_start_col).font = Font(bold=True)
    ws.cell(row=overview_start_row+2, column=overview_start_col).border = Border(left=Side(style='medium'))
    ws.cell(row=overview_start_row+2, column=overview_start_col + 2, value=final_discounted_safe).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=overview_start_row+2, column=overview_start_col + 2).font = Font(bold=True)
    ws.cell(row=overview_start_row+2, column=overview_start_col + 2).alignment = Alignment(horizontal='right')
    ws.cell(row=overview_start_row+2, column=overview_start_col + 2).border = Border(right=Side(style='medium'))
    
    # Residual/Waste Cost row
    ws.cell(row=overview_start_row+3, column=overview_start_col, value="Residual/Waste Cost:")
    ws.cell(row=overview_start_row+3, column=overview_start_col).border = Border(left=Side(style='medium'))
    ws.cell(row=overview_start_row+3, column=overview_start_col + 2, value=reuse_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=overview_start_row+3, column=overview_start_col + 2).alignment = Alignment(horizontal='right')
    ws.cell(row=overview_start_row+3, column=overview_start_col + 2).border = Border(right=Side(style='medium'))
    
    # Waste Percentage row
    ws.cell(row=overview_start_row+4, column=overview_start_col, value="Waste Percentage:")
    ws.cell(row=overview_start_row+4, column=overview_start_col).border = Border(left=Side(style='medium'), bottom=Side(style='medium'))
    ws.cell(row=overview_start_row+4, column=overview_start_col + 1).border = Border(bottom=Side(style='medium'))
    ws.cell(row=overview_start_row+4, column=overview_start_col + 2, value=f"{reuse_pct_of_gt:.2f}%")
    ws.cell(row=overview_start_row+4, column=overview_start_col + 2).alignment = Alignment(horizontal='right')
    ws.cell(row=overview_start_row+4, column=overview_start_col + 2).border = Border(right=Side(style='medium'), bottom=Side(style='medium'))
    
    overview_end_row = overview_start_row + 4
    
    # Ensure value column is wide enough for currency (prevents "########" display)
    value_col_letter = get_column_letter(overview_start_col + 2)
    ws.column_dimensions[value_col_letter].width = max(ws.column_dimensions[value_col_letter].width or 0, 14)
    
    # PIE CHART will be added after addition cost and markup totals are calculated

    # ============================================================================
    # ADDITIONAL COST - Stacked vertically below Cost Overview
    # ============================================================================
    
    # Start with spacing after the cost overview box
    section_start_row = overview_end_row + 2
    misc_start_row = section_start_row
    
    # ADDITIONAL COST Header (shifted right if elevation summary exists)
    ws.cell(row=section_start_row, column=category_start_col, value="ADDITIONAL COSTS").font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=section_start_row, column=category_start_col).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    ws.cell(row=section_start_row, column=category_start_col).border = Border(left=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='thin'))
    ws.cell(row=section_start_row, column=category_start_col + 1).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    ws.cell(row=section_start_row, column=category_start_col + 1).border = Border(top=Side(style='medium'), bottom=Side(style='thin'))
    ws.cell(row=section_start_row, column=category_start_col + 2).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    ws.cell(row=section_start_row, column=category_start_col + 2).border = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='thin'))
    summary_section_row = section_start_row + 1
    
    # Get summary percentages from project settings file
    summary_pcts = {
        "Overhead Materials": 0.0,
        "Overhead Labor": 0.0,
        "Admin and Management": 0.0,
        "Engineering": 0.0,
        "Packaging Materials": 0.0,
        "Shipping and Transport": 0.0,
        "Commissions": 0.0
    }
    
    # Load percentages from settings file
    print(f"[INFO] Attempting to load settings from: {summary_settings_path}")
    print(f"   Elevations path: {elevations_json_path}")
    
    # Always try to construct path from elevations path first (most reliable)
    settings_paths_to_try = []
    
    # 1. Try provided path
    if summary_settings_path:
        settings_paths_to_try.append(summary_settings_path)
        settings_paths_to_try.append(os.path.abspath(summary_settings_path))
    
    # 2. Construct from elevations path (most reliable)
    elev_dir = os.path.dirname(elevations_json_path)
    elev_basename = os.path.basename(elevations_json_path)
    if "_Elevations.json" in elev_basename:
        project_base = elev_basename.replace("_Elevations.json", "")
        constructed_path = os.path.join(elev_dir, f"{project_base}_Settings.json")
        settings_paths_to_try.append(constructed_path)
        print(f"   Constructed from elevations: {constructed_path}")
    
    # 3. Try in same directory as elevations
    if summary_settings_path:
        settings_paths_to_try.append(os.path.join(elev_dir, os.path.basename(summary_settings_path)))
    
    # Remove duplicates while preserving order (rebuild unique_paths here)
    seen = set()
    unique_paths = []
    for path in settings_paths_to_try:
        if path and path not in seen:
            seen.add(path)
            unique_paths.append(path)
    
    settings_loaded = False
    for path_to_try in unique_paths:
        if os.path.exists(path_to_try):
            try:
                print(f"   [OK] Found file, trying to read: {path_to_try}")
                with open(path_to_try, 'r') as f:
                    settings_data = json.load(f)
                    # Additional cost percentages
                    summary_pcts = {
                        "Overhead Materials": settings_data.get("overhead_materials_pct", 0.0),
                        "Overhead Labor": settings_data.get("overhead_labor_pct", 0.0),
                        "Admin and Management": settings_data.get("admin_management_pct", 0.0),
                        "Engineering": settings_data.get("engineering_pct", 0.0),
                        "Packaging Materials": settings_data.get("packaging_materials_pct", 0.0),
                        "Shipping and Transport": settings_data.get("shipping_transport_pct", 0.0),
                        "Commissions": settings_data.get("commissions_pct", 0.0)
                    }
                    print(f"[OK] Loaded summary percentages from {path_to_try}")
                    print(f"   Percentages: {summary_pcts}")
                    # Elevation summary display flags
                    elevation_summary_settings["show_elevation_names"] = settings_data.get("show_elevation_names", False)
                    elevation_summary_settings["show_elevation_quantity"] = settings_data.get("show_elevation_quantity", False)
                    elevation_summary_settings["show_elevation_dimensions"] = settings_data.get("show_elevation_dimensions", False)
                    elevation_summary_settings["show_elevation_sqft"] = settings_data.get("show_elevation_sqft", False)
                    elevation_summary_settings["show_elevation_perimeter"] = settings_data.get("show_elevation_perimeter", False)
                    settings_loaded = True
                    break
            except Exception as e:
                print(f"   [ERROR] Error reading {path_to_try}: {e}")
                import traceback
                traceback.print_exc()
                continue
    
    if not settings_loaded:
        print(f"[WARNING] Could not load settings from any path. Tried:")
        for path_to_try in unique_paths:
            exists = "[EXISTS]" if os.path.exists(path_to_try) else "[NOT FOUND]"
            print(f"   {exists}: {path_to_try}")
        if not summary_settings_path:
            print(f"   [WARNING] No summary settings path was provided to create_summary_sheet")

    # Build elevation summary column definitions based on flags
    elevation_summary_cols = []
    if elevation_summary_settings["show_elevation_names"]:
        elevation_summary_cols.append(("Elevation Name", 1))
    if elevation_summary_settings["show_elevation_quantity"]:
        # Include units in header so summary numbers stay numeric but units are clear
        elevation_summary_cols.append(("Quantity (EA)", len(elevation_summary_cols) + 1))
    if elevation_summary_settings["show_elevation_dimensions"]:
        elevation_summary_cols.append(("Dimensions", len(elevation_summary_cols) + 1))
    if elevation_summary_settings["show_elevation_sqft"]:
        elevation_summary_cols.append(("SQFT Total (SQFT)", len(elevation_summary_cols) + 1))
    if elevation_summary_settings["show_elevation_perimeter"]:
        elevation_summary_cols.append(("Perimeter FT Total (FT)", len(elevation_summary_cols) + 1))
    
    # ============================================================================
    # ELEVATION SUMMARY COLUMNS (if enabled) - Write before categories
    # ============================================================================
    # If elevation summary columns are enabled, write them as their own section
    elevation_summary_start_row = current_row
    if elevation_summary_cols:
        # Colors to match existing section headers (PROJECT TOTAL / GRAND TOTAL style)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        subheader_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

        # Add a clear section title spanning the used columns so it stands out visually
        first_col = elevation_summary_cols[0][1]
        last_col = elevation_summary_cols[-1][1]
        title_cell = ws.cell(row=current_row, column=first_col, value="ELEVATION SUMMARY")
        title_cell.font = header_font
        title_cell.fill = header_fill
        # Apply fill across merged range
        for col in range(first_col, last_col + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = header_fill
        ws.merge_cells(
            start_row=current_row,
            start_column=first_col,
            end_row=current_row,
            end_column=last_col,
        )
        current_row += 1

        # Header row directly under the title
        for _, (header_name, col_num) in enumerate(elevation_summary_cols):
            header_cell = ws.cell(row=current_row, column=col_num, value=header_name)
            header_cell.font = Font(bold=True, size=11)
            header_cell.fill = subheader_fill
            header_cell.border = Border(bottom=Side(style='thin'))
        current_row += 1

        # Totals accumulators
        total_qty = 0.0
        total_sqft = 0.0
        total_perimeter = 0.0

        # Data rows
        for elev_key, elev in data.items():
            col_idx = 0
            if elevation_summary_settings["show_elevation_names"]:
                ws.cell(row=current_row, column=elevation_summary_cols[col_idx][1], value=elev_key)
                col_idx += 1
            if elevation_summary_settings["show_elevation_quantity"]:
                qty = elev.get("total_count", 0) or 0
                ws.cell(row=current_row, column=elevation_summary_cols[col_idx][1], value=qty)
                total_qty += qty
                col_idx += 1
            if elevation_summary_settings["show_elevation_dimensions"]:
                width = elev.get("opening_width_inches", 0) or 0
                height = elev.get("opening_height_inches", 0) or 0
                ws.cell(row=current_row, column=elevation_summary_cols[col_idx][1], value=f"{width}\" x {height}\"")
                col_idx += 1
            if elevation_summary_settings["show_elevation_sqft"]:
                sqft = elev.get("total_sqft", 0) or 0
                ws.cell(row=current_row, column=elevation_summary_cols[col_idx][1], value=sqft)
                total_sqft += sqft
                col_idx += 1
            if elevation_summary_settings["show_elevation_perimeter"]:
                perimeter = elev.get("total_perimeter_ft", 0) or 0
                ws.cell(row=current_row, column=elevation_summary_cols[col_idx][1], value=perimeter)
                total_perimeter += perimeter
                col_idx += 1
            current_row += 1

        # Totals row (styled band across all used columns)
        total_row = current_row
        col_idx = 0

        # Apply a light banded fill across the whole totals row
        for _, (_, col_num) in enumerate(elevation_summary_cols):
            totals_cell = ws.cell(row=total_row, column=col_num)
            totals_cell.fill = subheader_fill

        if elevation_summary_settings["show_elevation_names"]:
            total_label_cell = ws.cell(row=total_row, column=elevation_summary_cols[col_idx][1], value="TOTAL")
            total_label_cell.font = Font(bold=True)
            col_idx += 1
        # Debug: print totals to console
        print(f"[SUMMARY] ELEVATION SUMMARY TOTALS: Qty={total_qty}, SQFT={total_sqft}, Perimeter={total_perimeter}")
        
        if elevation_summary_settings["show_elevation_quantity"]:
            # Write the accumulated total directly
            qty_col_index = elevation_summary_cols[col_idx][1]
            qty_value = int(total_qty) if total_qty == int(total_qty) else total_qty
            ws.cell(row=total_row, column=qty_col_index, value=qty_value)
            ws.cell(row=total_row, column=qty_col_index).font = Font(bold=True)
            ws.cell(row=total_row, column=qty_col_index).alignment = Alignment(horizontal="right")
            print(f"   Writing Qty total {qty_value} to row {total_row}, col {qty_col_index}")
            col_idx += 1
        if elevation_summary_settings["show_elevation_dimensions"]:
            # No total for dimensions; leave blank cell
            col_idx += 1
        if elevation_summary_settings["show_elevation_sqft"]:
            sqft_col_index = elevation_summary_cols[col_idx][1]
            sqft_value = int(total_sqft) if total_sqft == int(total_sqft) else round(total_sqft, 2)
            ws.cell(row=total_row, column=sqft_col_index, value=sqft_value)
            ws.cell(row=total_row, column=sqft_col_index).font = Font(bold=True)
            ws.cell(row=total_row, column=sqft_col_index).alignment = Alignment(horizontal="right")
            print(f"   Writing SQFT total {sqft_value} to row {total_row}, col {sqft_col_index}")
            col_idx += 1
        if elevation_summary_settings["show_elevation_perimeter"]:
            perim_col_index = elevation_summary_cols[col_idx][1]
            perim_value = int(total_perimeter) if total_perimeter == int(total_perimeter) else round(total_perimeter, 2)
            ws.cell(row=total_row, column=perim_col_index, value=perim_value)
            ws.cell(row=total_row, column=perim_col_index).font = Font(bold=True)
            ws.cell(row=total_row, column=perim_col_index).alignment = Alignment(horizontal="right")
            print(f"   Writing Perimeter total {perim_value} to row {total_row}, col {perim_col_index}")

        # Add a thin top border across the used columns to separate totals from data
        for _, (_, col_num) in enumerate(elevation_summary_cols):
            cell = ws.cell(row=total_row, column=col_num)
            cell.border = Border(top=Side(style='thin'))

        current_row += 2  # Spacing after elevation summary section

    # Calculate base amount: use discounted total only
    base_amount = final_discounted_total
    print(f"[SUMMARY] Additional Cost section - Base amount (discounted total): ${base_amount:.2f}")
    
    # ========== STEP 1: Collect all ADDITIONAL COST items ==========
    summary_total = 0.0
    misc_items_list = []
    for label, pct in summary_pcts.items():
        if pct > 0:
            amount = base_amount * (pct / 100.0)
            summary_total += amount
            misc_items_list.append((label, amount))
            print(f"   {label}: {pct}% = ${amount:.2f}")
    
    # ========== STEP 2: Collect all MARKUP items ==========
    # Get markup percentages from project settings file (same file as addition cost)
    markup_pcts = {
        "Profit on Material": 0.0,
        "Profit on Waste": 0.0,
        "Profit on Glass Purchase": 0.0,
        "Profit on Wages": 0.0,
        "Planning / Technical Office": 0.0,
        "Commission": 0.0
    }
    
    # Load markup percentages from settings file (use same file as addition cost settings)
    markup_settings_loaded = False
    for path_to_try_markup in unique_paths:
        if os.path.exists(path_to_try_markup):
            try:
                print(f"   [INFO] Loading markup settings from: {path_to_try_markup}")
                with open(path_to_try_markup, 'r') as f:
                    settings_data = json.load(f)
                    print(f"   [DEBUG] Settings file keys: {list(settings_data.keys())}")
                    markup_pcts = {
                        "Profit on Material": settings_data.get("profit_on_material_pct", 0.0),
                        "Profit on Waste": settings_data.get("profit_on_waste_pct", 0.0),
                        "Profit on Glass Purchase": settings_data.get("profit_on_glass_pct", 0.0),
                        "Profit on Wages": settings_data.get("profit_on_wages_pct", 0.0),
                        "Planning / Technical Office": settings_data.get("planning_technical_pct", 0.0),
                        "Commission": settings_data.get("commission_pct", 0.0)
                    }
                    print(f"[OK] Loaded markup percentages: {markup_pcts}")
                    # Check if any markup percentages are > 0
                    has_markups = any(pct > 0 for pct in markup_pcts.values())
                    print(f"   [DEBUG] Has markups > 0: {has_markups}")
                    markup_settings_loaded = True
                    break
            except Exception as e:
                print(f"   [ERROR] Error reading markup settings from {path_to_try_markup}: {e}")
                import traceback
                traceback.print_exc()
                continue
    
    if not markup_settings_loaded:
        print(f"[WARNING] Could not load markup settings from any path. Tried:")
        for path_to_try_markup in unique_paths:
            exists = "[EXISTS]" if os.path.exists(path_to_try_markup) else "[NOT FOUND]"
            print(f"   {exists}: {path_to_try_markup}")
    
    # Calculate markups based on appropriate bases
    markup_total = 0.0
    markup_items_list = []
    
    # 1. Profit on Material: sum of profiles, accessories, gaskets, and doors
    material_base = (category_totals["PROFILES"]["discounted"] + 
                     category_totals["ACCESSORIES"]["discounted"] + 
                     category_totals["GASKETS"]["discounted"] + 
                     category_totals["DOORS"]["discounted"])
    if markup_pcts["Profit on Material"] > 0 and material_base > 0:
        amount = material_base * (markup_pcts["Profit on Material"] / 100.0)
        markup_total += amount
        markup_items_list.append(("Profit on Material", amount))
        print(f"   Profit on Material: {markup_pcts['Profit on Material']}% of ${material_base:.2f} = ${amount:.2f}")
    
    # 2. Profit on Waste: sum of residual discounted price
    waste_base = grand_residual_total
    if markup_pcts["Profit on Waste"] > 0 and waste_base > 0:
        amount = waste_base * (markup_pcts["Profit on Waste"] / 100.0)
        markup_total += amount
        markup_items_list.append(("Profit on Waste", amount))
        print(f"   Profit on Waste: {markup_pcts['Profit on Waste']}% of ${waste_base:.2f} = ${amount:.2f}")
    
    # 3. Profit on Glass Purchase: sum of glass
    glass_base = category_totals["GLASS"]["discounted"]
    if markup_pcts["Profit on Glass Purchase"] > 0 and glass_base > 0:
        amount = glass_base * (markup_pcts["Profit on Glass Purchase"] / 100.0)
        markup_total += amount
        markup_items_list.append(("Profit on Glass Purchase", amount))
        print(f"   Profit on Glass Purchase: {markup_pcts['Profit on Glass Purchase']}% of ${glass_base:.2f} = ${amount:.2f}")
    
    # 4. Profit on Wages: sum of fabrication (labor)
    wages_base = category_totals["LABOR"]["discounted"]
    if markup_pcts["Profit on Wages"] > 0 and wages_base > 0:
        amount = wages_base * (markup_pcts["Profit on Wages"] / 100.0)
        markup_total += amount
        markup_items_list.append(("Profit on Wages", amount))
        print(f"   Profit on Wages: {markup_pcts['Profit on Wages']}% of ${wages_base:.2f} = ${amount:.2f}")
    
    # 5. Planning / Technical Office: discounted total * percentage
    if markup_pcts["Planning / Technical Office"] > 0 and final_discounted_total > 0:
        amount = final_discounted_total * (markup_pcts["Planning / Technical Office"] / 100.0)
        markup_total += amount
        markup_items_list.append(("Planning / Technical Office", amount))
        print(f"   Planning / Technical Office: {markup_pcts['Planning / Technical Office']}% of ${final_discounted_total:.2f} = ${amount:.2f}")
    
    # 6. Commission: discounted total * percentage
    if markup_pcts["Commission"] > 0 and final_discounted_total > 0:
        amount = final_discounted_total * (markup_pcts["Commission"] / 100.0)
        markup_total += amount
        markup_items_list.append(("Commission", amount))
        print(f"   Commission: {markup_pcts['Commission']}% of ${final_discounted_total:.2f} = ${amount:.2f}")
    
    # ========== STEP 3: Write ADDITIONAL COST items (shifted right if elevation summary exists) ==========
    misc_items_start_row = summary_section_row
    
    for i, (label, amount) in enumerate(misc_items_list):
        row = misc_items_start_row + i
        ws.cell(row=row, column=category_start_col, value=label)
        ws.cell(row=row, column=category_start_col).border = Border(left=Side(style='medium'))
        ws.cell(row=row, column=category_start_col + 2, value=amount).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=row, column=category_start_col + 2).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=category_start_col + 2).border = Border(right=Side(style='medium'))
    
    # Add "(None configured)" if no items
    if len(misc_items_list) == 0:
        ws.cell(row=misc_items_start_row, column=category_start_col, value="(None configured)").font = Font(italic=True)
        ws.cell(row=misc_items_start_row, column=category_start_col).border = Border(left=Side(style='medium'))
        ws.cell(row=misc_items_start_row, column=category_start_col + 2).border = Border(right=Side(style='medium'))
        misc_items_end_row = misc_items_start_row
    else:
        misc_items_end_row = misc_items_start_row + len(misc_items_list) - 1
    
    # Additional cost SUBTOTAL
    misc_subtotal_row = misc_items_end_row + 1
    ws.cell(row=misc_subtotal_row, column=category_start_col, value="SUBTOTAL").font = Font(bold=True)
    ws.cell(row=misc_subtotal_row, column=category_start_col).border = Border(left=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))
    ws.cell(row=misc_subtotal_row, column=category_start_col + 1).border = Border(top=Side(style='thin'), bottom=Side(style='medium'))
    ws.cell(row=misc_subtotal_row, column=category_start_col + 2, value=summary_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=misc_subtotal_row, column=category_start_col + 2).font = Font(bold=True)
    ws.cell(row=misc_subtotal_row, column=category_start_col + 2).alignment = Alignment(horizontal='right')
    ws.cell(row=misc_subtotal_row, column=category_start_col + 2).border = Border(right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))
    
    misc_end_row = misc_subtotal_row
    print(f"[OK] Additional Cost section: {len(misc_items_list)} items, total: ${summary_total:.2f}")
    
    # ========== STEP 4: Write MARKUPS HEADER (below Additional Cost) ==========
    markup_start_row = misc_end_row + 2
    
    ws.cell(row=markup_start_row, column=category_start_col, value="MARKUPS / PROFIT").font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=markup_start_row, column=category_start_col).fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    ws.cell(row=markup_start_row, column=category_start_col).border = Border(left=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='thin'))
    ws.cell(row=markup_start_row, column=category_start_col + 1).fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    ws.cell(row=markup_start_row, column=category_start_col + 1).border = Border(top=Side(style='medium'), bottom=Side(style='thin'))
    ws.cell(row=markup_start_row, column=category_start_col + 2).fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    ws.cell(row=markup_start_row, column=category_start_col + 2).border = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='thin'))
    
    # ========== STEP 5: Write MARKUP items (shifted right if elevation summary exists) ==========
    markup_items_start_row = markup_start_row + 1
    
    for i, (label, amount) in enumerate(markup_items_list):
        row = markup_items_start_row + i
        ws.cell(row=row, column=category_start_col, value=label)
        ws.cell(row=row, column=category_start_col).border = Border(left=Side(style='medium'))
        ws.cell(row=row, column=category_start_col + 2, value=amount).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=row, column=category_start_col + 2).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=category_start_col + 2).border = Border(right=Side(style='medium'))
    
    # Add "(None configured)" if no items
    # Check if markups are configured but just have zero base amounts
    has_configured_markups = any(pct > 0 for pct in markup_pcts.values())
    if len(markup_items_list) == 0:
        if has_configured_markups:
            # Markups are configured but bases are zero - show configured markups with $0.00
            configured_markups = []
            for label, pct_key in [("Profit on Material", "Profit on Material"),
                                    ("Profit on Waste", "Profit on Waste"),
                                    ("Profit on Glass Purchase", "Profit on Glass Purchase"),
                                    ("Profit on Wages", "Profit on Wages"),
                                    ("Planning / Technical Office", "Planning / Technical Office"),
                                    ("Commission", "Commission")]:
                if markup_pcts[pct_key] > 0:
                    configured_markups.append((label, 0.0))
            
            for i, (label, amount) in enumerate(configured_markups):
                row = markup_items_start_row + i
                ws.cell(row=row, column=category_start_col, value=label)
                ws.cell(row=row, column=category_start_col).border = Border(left=Side(style='medium'))
                ws.cell(row=row, column=category_start_col + 2, value=amount).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                ws.cell(row=row, column=category_start_col + 2).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=category_start_col + 2).border = Border(right=Side(style='medium'))
                markup_items_list.append((label, amount))
            
            if len(markup_items_list) > 0:
                markup_items_end_row = markup_items_start_row + len(markup_items_list) - 1
            else:
                ws.cell(row=markup_items_start_row, column=category_start_col, value="(None configured)").font = Font(italic=True)
                ws.cell(row=markup_items_start_row, column=category_start_col).border = Border(left=Side(style='medium'))
                ws.cell(row=markup_items_start_row, column=category_start_col + 2).border = Border(right=Side(style='medium'))
                markup_items_end_row = markup_items_start_row
        else:
            # No markups configured at all
            ws.cell(row=markup_items_start_row, column=category_start_col, value="(None configured)").font = Font(italic=True)
            ws.cell(row=markup_items_start_row, column=category_start_col).border = Border(left=Side(style='medium'))
            ws.cell(row=markup_items_start_row, column=category_start_col + 2).border = Border(right=Side(style='medium'))
            markup_items_end_row = markup_items_start_row
    else:
        markup_items_end_row = markup_items_start_row + len(markup_items_list) - 1
    
    # Markup SUBTOTAL
    markup_subtotal_row = markup_items_end_row + 1
    ws.cell(row=markup_subtotal_row, column=category_start_col, value="SUBTOTAL").font = Font(bold=True)
    ws.cell(row=markup_subtotal_row, column=category_start_col).border = Border(left=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))
    ws.cell(row=markup_subtotal_row, column=category_start_col + 1).border = Border(top=Side(style='thin'), bottom=Side(style='medium'))
    ws.cell(row=markup_subtotal_row, column=category_start_col + 2, value=markup_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=markup_subtotal_row, column=category_start_col + 2).font = Font(bold=True)
    ws.cell(row=markup_subtotal_row, column=category_start_col + 2).alignment = Alignment(horizontal='right')
    ws.cell(row=markup_subtotal_row, column=category_start_col + 2).border = Border(right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))
    
    markup_end_row = markup_subtotal_row
    print(f"[OK] Markup section: {len(markup_items_list)} items, total: ${markup_total:.2f}")
    
    # ============================================================================
    # FINAL TOTAL - Below markup section with spacing
    # ============================================================================
    
    final_total_row = markup_end_row + 2
    final_total_amount = final_discounted_safe + summary_total + markup_total
    
    # Header row with dark background (shifted right if elevation summary exists)
    ws.cell(row=final_total_row, column=category_start_col, value="PROJECT TOTAL").font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=final_total_row, column=category_start_col).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ws.cell(row=final_total_row, column=category_start_col).border = Border(left=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='thin'))
    ws.cell(row=final_total_row, column=category_start_col + 1).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ws.cell(row=final_total_row, column=category_start_col + 1).border = Border(top=Side(style='medium'), bottom=Side(style='thin'))
    ws.cell(row=final_total_row, column=category_start_col + 2).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ws.cell(row=final_total_row, column=category_start_col + 2).border = Border(right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='thin'))
    
    # Light fill for alternating rows
    light_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    
    # Discounted Total row
    final_total_row += 1
    ws.cell(row=final_total_row, column=category_start_col, value="Discounted Total:")
    ws.cell(row=final_total_row, column=category_start_col).fill = light_fill
    ws.cell(row=final_total_row, column=category_start_col).border = Border(left=Side(style='medium'))
    ws.cell(row=final_total_row, column=category_start_col + 1).fill = light_fill
    ws.cell(row=final_total_row, column=category_start_col + 2, value=final_discounted_safe).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=final_total_row, column=category_start_col + 2).alignment = Alignment(horizontal='right')
    ws.cell(row=final_total_row, column=category_start_col + 2).fill = light_fill
    ws.cell(row=final_total_row, column=category_start_col + 2).border = Border(right=Side(style='medium'))
    
    # Additional cost Total row
    final_total_row += 1
    ws.cell(row=final_total_row, column=category_start_col, value="+ Additional:")
    ws.cell(row=final_total_row, column=category_start_col).border = Border(left=Side(style='medium'))
    ws.cell(row=final_total_row, column=category_start_col + 2, value=summary_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=final_total_row, column=category_start_col + 2).alignment = Alignment(horizontal='right')
    ws.cell(row=final_total_row, column=category_start_col + 2).border = Border(right=Side(style='medium'))
    
    # Markup Total row
    final_total_row += 1
    ws.cell(row=final_total_row, column=category_start_col, value="+ Markups:")
    ws.cell(row=final_total_row, column=category_start_col).border = Border(left=Side(style='medium'))
    ws.cell(row=final_total_row, column=category_start_col + 2, value=markup_total).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=final_total_row, column=category_start_col + 2).alignment = Alignment(horizontal='right')
    ws.cell(row=final_total_row, column=category_start_col + 2).border = Border(right=Side(style='medium'))
    
    # Grand Total row (bold, highlighted with dark background)
    final_total_row += 1
    ws.cell(row=final_total_row, column=category_start_col, value="GRAND TOTAL:").font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=final_total_row, column=category_start_col).fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    ws.cell(row=final_total_row, column=category_start_col).border = Border(left=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))
    ws.cell(row=final_total_row, column=category_start_col + 1).fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    ws.cell(row=final_total_row, column=category_start_col + 1).border = Border(top=Side(style='thin'), bottom=Side(style='medium'))
    ws.cell(row=final_total_row, column=category_start_col + 2, value=final_total_amount).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws.cell(row=final_total_row, column=category_start_col + 2).font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=final_total_row, column=category_start_col + 2).alignment = Alignment(horizontal='right')
    ws.cell(row=final_total_row, column=category_start_col + 2).fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    ws.cell(row=final_total_row, column=category_start_col + 2).border = Border(right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))
    
    # Ensure value column is wide enough for currency (prevents "########" display)
    total_value_col = get_column_letter(category_start_col + 2)
    ws.column_dimensions[total_value_col].width = max(ws.column_dimensions[total_value_col].width or 0, 14)
    
    # ============================================================================
    # PIE CHART - Cost Distribution (placed at specific location)
    # ============================================================================
    # Add pie chart showing full cost breakdown: Materials, Additional, Markups, Residual
    # Place it at row 40, column G (column 7)
    pie_chart_col = 7  # Column G
    pie_chart_row = 40  # Row 40
    
    # Calculate active material cost (total minus residual to avoid double-counting)
    active_material_cost = max(0, final_discounted_total - reuse_total)
    
    try:
        _add_pie_chart_to_excel(ws, pie_chart_row, pie_chart_col, active_material_cost, summary_total, markup_total, reuse_total)
        print(f"[OK] Added pie chart at row {pie_chart_row}, column {pie_chart_col}")
        print(f"   Active Materials: ${active_material_cost:.2f}, Additional: ${summary_total:.2f}, Markups: ${markup_total:.2f}, Residual: ${reuse_total:.2f}")
    except Exception as e:
        print(f"[WARNING] Could not add pie chart: {e}")
    
    print(f"[SUMMARY] Final Total: ${final_discounted_total:.2f} (discounted) + ${summary_total:.2f} (addition) + ${markup_total:.2f} (markups) = ${final_total_amount:.2f}")
    print(f"[INFO] Markup section written to rows {markup_start_row} to {markup_end_row if 'markup_end_row' in locals() else markup_section_row}")
    print(f"[INFO] Final total written to row {final_total_row}")
    
    _autofit_columns(ws, 1, 10, start_row, final_total_row)
    _clean_trailing_blank_rows(ws, 1)

    print(f"[OK] Summary updated with grouped sections: Profiles, Accessories, Doors, Glass, Labor.")

def _format_door_summary(calculated_outputs):
    if not calculated_outputs:
        return ""
    door_lines = []
    for item in calculated_outputs:
        if item.get("type", "").lower() in ["door", "doors"] and item.get('manual', False):
            quantity = item.get("quantity", 1)
            style = item.get("Style", "").strip()
            price = item.get("price", 0.0)
            hardware = item.get("hardware", {})
            if not (isinstance(quantity, (int, float)) and quantity > 0):
                continue
            if not style or style.lower() == "unknown":
                continue
            enabled_hw = [hw for hw, enabled in hardware.items() if enabled]
            if enabled_hw:
                door_lines.append(f"{quantity} x {style} Door (${price:,.2f})\n  with: {', '.join(enabled_hw)}")
            else:
                door_lines.append(f"{quantity} x {style} Door (${price:,.2f})")
    return "; \n".join(door_lines) if door_lines else "None"

def generate_excel_report(
    excel_path, elevations_json_path, extra_materials_json_path,
    system_input, finish_input, elevation_type, total_count,
    bays_wide, bays_tall, opening_width, opening_height,
    sqft_per_type, total_sqft, perimeter_ft, total_perimeter_ft,
    calculated_outputs, completion_callback=None, reset=False, delete_elevation_type=None,
    doors=None, mode=None, custom_bay_widths=None, custom_bay_heights=None, summary_settings_path=None,
    door_only=False
):
    """Generates or updates an Excel report with detailed elevation inputs and calculated outputs."""
    COL_A, COL_B, COL_E, PRICE_COL = 1, 2, 5, 9

    project_root = os.getcwd()
    private_projects_dir = os.path.abspath(os.path.join(project_root, '.files'))
    public_reports_dir = os.path.abspath(os.path.join(project_root, 'reports'))

    os.makedirs(private_projects_dir, exist_ok=True)
    os.makedirs(public_reports_dir, exist_ok=True)
    
    # Normalize paths - use provided path if it exists and is in .files, otherwise construct it
    elevations_path_abs = os.path.abspath(elevations_json_path)
    if os.path.dirname(elevations_path_abs) == private_projects_dir and os.path.exists(elevations_path_abs):
        private_elevations_path = elevations_path_abs
    else:
        private_elevations_path = os.path.join(private_projects_dir, os.path.basename(elevations_json_path))
    
    materials_path_abs = os.path.abspath(extra_materials_json_path)
    if os.path.dirname(materials_path_abs) == private_projects_dir and os.path.exists(materials_path_abs):
        private_extra_materials_path = materials_path_abs
    else:
        private_extra_materials_path = os.path.join(private_projects_dir, os.path.basename(extra_materials_json_path))
    
    excel_path_abs = os.path.abspath(excel_path)
    if os.path.dirname(excel_path_abs) == private_projects_dir:
        private_excel_path = excel_path_abs
    else:
        private_excel_path = os.path.join(private_projects_dir, os.path.basename(excel_path))
    
    # Debug logging
    print(f"[PATHS] Using paths:")
    print(f"   Elevations: {private_elevations_path}")
    print(f"   Materials: {private_extra_materials_path}")
    print(f"   Excel: {private_excel_path}")
    
    output_excel_path = public_reports_dir if mode == "export_all" else private_excel_path
    
    current_saved_elevations = {}
    if os.path.exists(private_elevations_path):
        try:
            with open(private_elevations_path, 'r') as f:
                current_saved_elevations = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"Error loading {private_elevations_path}: {e}. Starting with empty elevations in memory.")

    if delete_elevation_type:
        elevation_to_delete_data = current_saved_elevations.get(delete_elevation_type)
        if elevation_to_delete_data and 'material_impact' in elevation_to_delete_data:
            reverse_material_impact(elevation_to_delete_data['material_impact'], extra_materials_file=private_extra_materials_path)

        if delete_elevation_type in current_saved_elevations:
            del current_saved_elevations[delete_elevation_type]

        try:
            with open(private_elevations_path, 'w') as f:
                json.dump(current_saved_elevations, f, indent=4)
        except IOError as e:
            print(f"Error saving updated {private_elevations_path} during delete: {e}")
            if completion_callback: completion_callback(f"Error saving updated elevations after delete: {e}")

    # Handle regeneration/cleanup case
    if delete_elevation_type and system_input == "": 
        pass # Don't add a new empty elevation if we are just deleting
    elif mode == "export_all":
        pass
    else:
        if elevation_type in current_saved_elevations and not reset:
            old_elevation_data = current_saved_elevations[elevation_type]
            if 'material_impact' in old_elevation_data:
                reverse_material_impact(old_elevation_data['material_impact'], extra_materials_file=private_extra_materials_path)

        # Build door items from UI doors, but avoid double-adding if base outputs already include doors
        base_outputs = list(calculated_outputs or [])
        
        # Remove any existing door entries from base_outputs to prevent duplication/stale data
        base_outputs = [item for item in base_outputs if not (item.get('type', '').lower() in ['door', 'doors'] and item.get('manual', False))]
        
        # Recalculate door items fresh from current inputs
        door_items = calculate_door_info(doors, finish_input, total_count) if doors else []
        
        # Avoid mutating the incoming list reference
        elevation_outputs = base_outputs + door_items

        # Preserve column display preferences from old elevation data if they exist
        old_show_qty_per_elev = False
        old_show_total_cost_per_elev = False
        old_show_discounted_cost_per_elev = False
        if elevation_type in current_saved_elevations:
            old_show_qty_per_elev = current_saved_elevations[elevation_type].get('show_qty_per_elevation', False)
            old_show_total_cost_per_elev = current_saved_elevations[elevation_type].get('show_total_cost_per_elevation', False)
            old_show_discounted_cost_per_elev = current_saved_elevations[elevation_type].get('show_discounted_cost_per_elevation', False)

        current_saved_elevations[elevation_type] = {
            "system": "Door Only" if door_only else system_input, "finish": finish_input, "total_count": total_count,
            "bays_wide": bays_wide, "bays_tall": bays_tall, "opening_width_inches": opening_width,
            "opening_height_inches": opening_height, "sqft_per_type": sqft_per_type, "total_sqft": total_sqft,
            "perimeter_ft": perimeter_ft, "total_perimeter_ft": total_perimeter_ft,
            "calculated_outputs": elevation_outputs,
            "material_impact": [],
            "custom_bay_widths": custom_bay_widths or [],
            "custom_bay_heights": custom_bay_heights or [],
            "show_qty_per_elevation": old_show_qty_per_elev,
            "show_total_cost_per_elevation": old_show_total_cost_per_elev,
            "show_discounted_cost_per_elevation": old_show_discounted_cost_per_elev,
            "door_only": door_only
        }

        try:
            with open(private_elevations_path, 'w') as f:
                json.dump(current_saved_elevations, f, indent=4)
        except IOError as e:
            print(f"Error saving elevation to {private_elevations_path}: {e}")
            if completion_callback: completion_callback(f"Error saving elevation: {e}")
            return

    wb = Workbook()
    # Remove the default "Sheet" immediately so it doesn't end up in the final report
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    save_extra_materials({}, private_extra_materials_path)
    overall_current_extra_materials_state = load_extra_materials(private_extra_materials_path)

    full_running_grand_total = 0.0
    for elev_name in current_saved_elevations:
        elev_data = current_saved_elevations[elev_name]
        for item in elev_data.get('calculated_outputs', []):
            pn = item.get('part_number')
            qty_raw = item.get('quantity', 0)
            individual_quantities = qty_raw if isinstance(qty_raw, list) else [qty_raw]
            manual = item.get('manual', False)
            price_sum = 0.0
            for single_qty in individual_quantities:
                if manual:
                    if pn and pn != "N/A":
                        p, _, _ = get_price_by_part(pn, single_qty, finish=elev_data.get('finish'), summary=True, group=True)
                        price_sum += p if p is not None else item.get('price', 0.0) * single_qty
                    else:
                        price_sum += item.get('price', 0.0) * single_qty
                else:
                    p, _, _ = get_price_by_part(pn, single_qty, finish=elev_data.get('finish'), summary=True)
                    price_sum += p if p is not None else 0.0
            full_running_grand_total += price_sum

    multiplier = _get_multiplier(full_running_grand_total)

    sorted_elev_names = sorted(current_saved_elevations.keys())

    if not sorted_elev_names:
        pass
    else:
        for elev_name in sorted_elev_names:
            ws = wb.create_sheet(title=elev_name)
            elev_data = current_saved_elevations[elev_name]
            
            # Create a fresh extra_materials state for this elevation (no leftovers from other elevations)
            # This ensures each elevation is calculated independently
            elevation_extra_materials_state = {}

            # Format custom bay dimensions for display
            custom_bay_widths_str = ", ".join([f"{w:.2f} in" for w in elev_data.get('custom_bay_widths', [])]) if elev_data.get('custom_bay_widths') else "Equal distribution"
            custom_bay_heights_str = ", ".join([f"{h:.2f} in" for h in elev_data.get('custom_bay_heights', [])]) if elev_data.get('custom_bay_heights') else "Equal distribution"

            # Safe formatting for potentially missing values (e.g. door-only elevations)
            _ow = elev_data.get('opening_width_inches')
            _oh = elev_data.get('opening_height_inches')
            _sqft = elev_data.get('sqft_per_type')
            _tsqft = elev_data.get('total_sqft')
            _perim = elev_data.get('perimeter_ft')
            _tperim = elev_data.get('total_perimeter_ft')
            system_display = "Door Only" if elev_data.get("door_only") else (elev_data.get("system") or "N/A")
            input_data = [
                ("System Input", system_display),
                ("Finish", elev_data.get("finish") or "N/A"),
                ("Elevation Type", elev_name),
                ("Total Count", elev_data.get("total_count")),
                ("Bays Wide", elev_data.get("bays_wide")),
                ("Bays Tall", elev_data.get("bays_tall")),
                ("Custom Bay Widths", custom_bay_widths_str),
                ("Custom Bay Heights", custom_bay_heights_str),
                ("Opening Width", f"{_ow:.2f} in" if _ow is not None else "N/A"),
                ("Opening Height", f"{_oh:.2f} in" if _oh is not None else "N/A"),
                ("Sq Ft per Type", f"{_sqft:.2f} sqft" if _sqft is not None else "N/A"),
                ("Total Sq Ft", f"{_tsqft:.2f} sqft" if _tsqft is not None else "N/A"),
                ("Perimeter Ft", f"{_perim:.2f} ft" if _perim is not None else "N/A"),
                ("Total Perimeter Ft", f"{_tperim:.2f} ft" if _tperim is not None else "N/A"),
                ("Doors", _format_door_summary(elev_data.get("calculated_outputs", [])))
            ]

            current_excel_row = 1
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

            for i, (header, value) in enumerate(input_data):
                header_cell = ws.cell(row=current_excel_row + i, column=COL_A, value=header)
                header_cell.font = Font(bold=True)
                # header_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Removed color fill for professional look
                header_cell.border = thin_border 

                value_cell = ws.cell(row=current_excel_row + i, column=COL_B, value=value)
                value_cell.border = thin_border
                if header in ["Total Count", "Bays Wide", "Bays Tall"]:
                    value_cell.alignment = Alignment(horizontal='left')
            
            # Add bay distribution diagram if custom bays are used
            if elev_data.get("bays_wide") and elev_data.get("bays_tall"):
                diagram_row = current_excel_row + len(input_data) + 2  # Tighter spacing
                custom_widths = elev_data.get('custom_bay_widths', [])
                custom_heights = elev_data.get('custom_bay_heights', [])
                
                # Add diagram label - with 1 more row separation before the picture
                label_cell = ws.cell(row=diagram_row - 2, column=COL_A, value="Bay Distribution Diagram")
                label_cell.font = Font(bold=True, size=12)
                # Picture will be at diagram_row, so there's now 2 rows between label and picture
                
                # Add note in column B (next to bay diagram)
                note_cell = ws.cell(row=diagram_row - 2, column=COL_B, value="*Note - C/L Dimensions")
                note_cell.font = Font(size=12)
                note_cell.alignment = Alignment(horizontal='left', vertical='top')
                
                # Add the diagram (include green door bands when doors present)
                _add_bay_diagram_to_excel(
                    ws, 
                    diagram_row,
                    elev_data.get("bays_wide", 0),
                    elev_data.get("bays_tall", 0),
                    elev_data.get('opening_width_inches', 0),
                    elev_data.get('opening_height_inches', 0),
                    custom_widths if custom_widths else None,
                    custom_heights if custom_heights else None,
                    doors=elev_data.get("doors")
                )
            elif elev_data.get("door_only") and elev_data.get("doors"):
                # Door-only elevation: add diagram(s) — adjacent (side-by-side) in same row
                diagram_row = current_excel_row + len(input_data) + 2  # Tighter spacing
                label_cell = ws.cell(row=diagram_row - 2, column=COL_A, value="Door Only Diagram(s)")
                label_cell.font = Font(bold=True, size=12)
                _add_door_only_diagrams_to_excel(ws, diagram_row, elev_data.get("doors"))
            
            output_section_current_row = 1
            profiles_for_section, accessories_for_section, gaskets_for_section, doors_for_section, other_items_for_section = [], [], [], [], []

            current_elevation_finish = elev_data.get("finish")

            # First pass: collect all items and combine same part numbers for profiles/gaskets
            items_by_part = {}  # Key: (part_number, finish), Value: list of items
            
            for item in elev_data.get('calculated_outputs', []):
                pn, manual = item.get('part_number'), item.get('manual', False)
                desc = item.get('description', '').strip()
                is_gasket = "gasket" in desc.lower() or pn in ["E2-0052", "E2-0053", "E2-0065"]
                is_profile = pn in PART_NUMBER_MAP.get("profiles", {})
                is_door = item.get('type', '').lower() in ['door', 'doors']
                
                # For profiles and gaskets, combine items with same part number
                if (is_profile or is_gasket) and pn and pn != "N/A" and not manual:
                    key = (pn, current_elevation_finish)
                    if key not in items_by_part:
                        items_by_part[key] = []
                    items_by_part[key].append(item)
                else:
                    # For other items, add directly to appropriate section
                    if pn and pn != "N/A":
                        if manual and is_door:
                            doors_for_section.append(item)
                        elif manual:
                            other_items_for_section.append(item)
                        elif is_profile:
                            profiles_for_section.append(item)
                        elif is_gasket:
                            gaskets_for_section.append(item)
                        elif pn in PART_NUMBER_MAP.get("accessories", {}) or item.get('type', '').lower() == 'accessory':
                            accessories_for_section.append(item)
                        else:
                            other_items_for_section.append(item)
                    else:
                        if is_door:
                            doors_for_section.append(item)
                        else:
                            other_items_for_section.append(item)
            
            # Combine items with same part number for profiles/gaskets
            for (pn, finish), items_list in items_by_part.items():
                if len(items_list) > 1:
                    # Combine quantities into one list
                    combined_quantities = []
                    combined_descriptions = []
                    total_qty_sum = 0.0
                    total_cost = 0.0
                    
                    for item in items_list:
                        qty = item.get('quantity', 0)
                        if isinstance(qty, list):
                            combined_quantities.extend(qty)
                            qty_sum = sum(qty)
                        else:
                            combined_quantities.append(qty)
                            qty_sum = qty
                        
                        combined_descriptions.append(item.get('description', ''))
                        total_qty_sum += qty_sum
                        # Price is per unit, so multiply by quantity
                        total_cost += item.get('price', 0.0) * qty_sum
                    
                    # Create combined item with weighted average price
                    combined_item = {
                        'description': ' / '.join(combined_descriptions) if len(set(combined_descriptions)) > 1 else combined_descriptions[0],
                        'quantity': combined_quantities,
                        'part_number': pn,
                        'type': items_list[0].get('type', 'profiles'),
                        'price': total_cost / total_qty_sum if total_qty_sum > 0 else 0.0
                    }
                    
                    print(f"DEBUG: Combined {len(items_list)} items for {pn}: {combined_quantities}")
                    
                    # Add to appropriate section
                    if pn in PART_NUMBER_MAP.get("profiles", {}):
                        profiles_for_section.append(combined_item)
                    elif "gasket" in combined_item['description'].lower() or pn in ["E2-0052", "E2-0053", "E2-0065"]:
                        gaskets_for_section.append(combined_item)
                else:
                    # Single item, add directly
                    item = items_list[0]
                    if pn in PART_NUMBER_MAP.get("profiles", {}):
                        profiles_for_section.append(item)
                    elif "gasket" in item.get('description', '').lower() or pn in ["E2-0052", "E2-0053", "E2-0065"]:
                        gaskets_for_section.append(item)

            system_total_for_this_block = [0.0]
            original_system_total_for_this_block = [0.0]
            newly_calculated_material_impacts_for_this_elevation = []
            
            # Get column display preferences and total_count from elevation data
            show_qty_per_elev = elev_data.get('show_qty_per_elevation', False)
            show_total_cost_per_elev = elev_data.get('show_total_cost_per_elevation', False)
            show_discounted_cost_per_elev = elev_data.get('show_discounted_cost_per_elevation', False)
            elev_total_count = elev_data.get('total_count', 1)
            
            # Track totals row numbers for reading from Excel columns
            profile_totals_row = None
            accessory_totals_row = None
            gasket_totals_row = None
            door_totals_row = None

            next_row_after_profiles, impacts_p, profile_totals = _write_output_section(
                ws, "PROFILES", profiles_for_section, COL_E, current_elevation_finish,
                system_total_for_this_block, original_system_total_for_this_block, output_section_current_row,
                elevation_extra_materials_state, private_extra_materials_path, multiplier,
                show_qty_per_elevation=show_qty_per_elev, total_count=elev_total_count,
                show_total_cost_per_elevation=show_total_cost_per_elev, show_discounted_cost_per_elevation=show_discounted_cost_per_elev
            )
            profile_totals_row = (next_row_after_profiles - 1) if profiles_for_section else None

            next_row_after_accessories, impacts_a, accessory_totals = _write_output_section(
                ws, "ACCESSORIES", accessories_for_section, COL_E, current_elevation_finish,
                system_total_for_this_block, original_system_total_for_this_block, next_row_after_profiles,
                elevation_extra_materials_state, private_extra_materials_path, multiplier,
                show_qty_per_elevation=show_qty_per_elev, total_count=elev_total_count,
                show_total_cost_per_elevation=show_total_cost_per_elev, show_discounted_cost_per_elevation=show_discounted_cost_per_elev
            )
            accessory_totals_row = (next_row_after_accessories - 1) if accessories_for_section else None

            next_row_after_gaskets, impacts_g, gasket_totals = _write_output_section(
                ws, "GASKETS", gaskets_for_section, COL_E, current_elevation_finish,
                system_total_for_this_block, original_system_total_for_this_block, next_row_after_accessories,
                elevation_extra_materials_state, private_extra_materials_path, multiplier,
                show_qty_per_elevation=show_qty_per_elev, total_count=elev_total_count,
                show_total_cost_per_elevation=show_total_cost_per_elev, show_discounted_cost_per_elevation=show_discounted_cost_per_elev
            )
            gasket_totals_row = (next_row_after_gaskets - 1) if gaskets_for_section else None

            newly_calculated_material_impacts_for_this_elevation.extend(impacts_p)
            newly_calculated_material_impacts_for_this_elevation.extend(impacts_a)
            newly_calculated_material_impacts_for_this_elevation.extend(impacts_g)

            # Process doors section
            next_row_after_doors, impacts_d, door_totals = _write_output_section(
                ws, "DOORS", doors_for_section, COL_E, current_elevation_finish,
                system_total_for_this_block, original_system_total_for_this_block, next_row_after_gaskets,
                elevation_extra_materials_state, private_extra_materials_path, multiplier,
                show_qty_per_elevation=show_qty_per_elev, total_count=elev_total_count,
                show_total_cost_per_elevation=show_total_cost_per_elev, show_discounted_cost_per_elevation=show_discounted_cost_per_elev
            )
            # Only set door_totals_row if there are actually doors
            door_totals_row = (next_row_after_doors - 1) if doors_for_section else None
            newly_calculated_material_impacts_for_this_elevation.extend(impacts_d)

            current_section_row = next_row_after_doors
            grouped_other_misc = {}
            glass_totals_rows = []
            fabrication_totals_rows = []

            for item in other_items_for_section:
                item_type = item.get('type', 'ADDITIONAL ITEMS').upper()
                grouped_other_misc.setdefault(item_type, []).append(item)

            for grp_title, grp_items in grouped_other_misc.items():
                next_row_after_group, impacts_g, group_totals = _write_output_section(
                    ws, grp_title, grp_items, COL_E, None,
                    system_total_for_this_block, original_system_total_for_this_block, current_section_row,
                    elevation_extra_materials_state, private_extra_materials_path, multiplier,
                    show_qty_per_elevation=show_qty_per_elev, total_count=elev_total_count,
                    show_total_cost_per_elevation=show_total_cost_per_elev, show_discounted_cost_per_elevation=show_discounted_cost_per_elev
                )
                newly_calculated_material_impacts_for_this_elevation.extend(impacts_g)
                
                # Track totals row for reading from Excel
                group_totals_row = next_row_after_group - 1
                
                # Categorize other items into Glass or Fabrication
                if grp_title == "GLASS" or any(item.get('part_number') == "GLASS_AREA" or item.get('type', '').lower() == 'glass' for item in grp_items):
                    glass_totals_rows.append(group_totals_row)
                else:
                    # Treat other items as fabrication costs
                    fabrication_totals_rows.append(group_totals_row)
                
                current_section_row = next_row_after_group
                print(f"Section '{grp_title}' ended at row {current_section_row}")

            current_saved_elevations[elev_name]['material_impact'] = newly_calculated_material_impacts_for_this_elevation

            # Merge elevation's extra materials state into overall state
            # This accumulates leftovers from all elevations
            print(f"DEBUG: Merging elevation state for {elev_name}, keys: {list(elevation_extra_materials_state.keys())}")
            for key, value in elevation_extra_materials_state.items():
                print(f"DEBUG: Processing key {key}, value: {value}")
                if key not in overall_current_extra_materials_state:
                    overall_current_extra_materials_state[key] = {'quantity': 0, 'length_pieces': []}
                
                # Merge length_pieces (for profiles/gaskets)
                if 'length_pieces' in value and isinstance(value['length_pieces'], list):
                    existing_pieces = overall_current_extra_materials_state[key].get('length_pieces', [])
                    if not isinstance(existing_pieces, list):
                        existing_pieces = []
                    # Combine and sort
                    combined = existing_pieces + value['length_pieces']
                    overall_current_extra_materials_state[key]['length_pieces'] = sorted(combined)
                    print(f"DEBUG: Merged length_pieces for {key}: {combined}")
                
                # Merge quantity (for accessories)
                if 'quantity' in value:
                    overall_current_extra_materials_state[key]['quantity'] = (
                        overall_current_extra_materials_state[key].get('quantity', 0) + value.get('quantity', 0)
                    )
                    print(f"DEBUG: Merged quantity for {key}: {overall_current_extra_materials_state[key]['quantity']}")

            # Create cost breakdown summary table
            # Use the explicitly tracked row position after the last section
            # Add spacing: explicitly create blank rows
            spacing_rows = 1  # Number of blank rows before cost summary (row 39 -> row 43 for headers)
            
            # Explicitly create blank rows by writing empty cells to ensure rows exist
            for blank_row in range(1, spacing_rows + 1):
                # Write empty cell to ensure row exists in Excel
                ws.cell(row=current_section_row + blank_row, column=COL_A, value="")
            
            # Headers start after all spacing rows, with one additional blank row
            cost_summary_row = current_section_row + spacing_rows + 1
            
            # Calculate column numbers for reading from Excel
            # Use the same logic as _write_output_section uses for writing totals row
            # total_col_offset starts at 2 (after Description, Part Number, Total Quantity Required)
            # Then: +1 if qty_per_elev, +1 (label), +1 (Total List Cost), +1 if total_cost_per_elev, then Discounted Total List Cost
            total_col_offset = 2  # Start after "Total Quantity Required"
            if show_qty_per_elev and elev_total_count > 1:
                total_col_offset += 1  # Skip "Quantity Per Elevation" column
            total_col_offset += 1  # Skip to "Total List Cost" column (where label is written)
            total_col_offset += 1  # Skip "Total List Cost" value column
            if show_total_cost_per_elev and elev_total_count > 1:
                total_col_offset += 1  # Skip "Total List Cost Per Elevation" column
            # Now total_col_offset points to "Discounted Total List Cost" column
            
            col_k = COL_E + total_col_offset  # Discounted Total List Cost (Column K)
            col_l = col_k + 1 if (show_discounted_cost_per_elev and elev_total_count > 1) else None  # Discounted Total List Cost Per Elevation (Column L)
            
            # Set cost summary columns - use col_k for total elevation cost column
            header_col = PRICE_COL - 2
            cost_per_elev_col = col_l if col_l else col_k  # Use column L if available, otherwise K
            total_elev_cost_col = col_k  # Use the actual Discounted Total List Cost column
            
            print(f"Column calculation: show_qty_per_elev={show_qty_per_elev}, show_total_cost_per_elev={show_total_cost_per_elev}, show_discounted_cost_per_elev={show_discounted_cost_per_elev}")
            print(f"Total col offset={total_col_offset}, Column K={col_k}, Column L={col_l}")
            
            total_count = elev_data.get("total_count", 1)
            
            # Debug: print row numbers to verify
            print(f"Cost summary for '{elev_name}': Last section row={current_section_row}, Spacing={spacing_rows} rows, Cost summary starts at row={cost_summary_row}")
            print(f"Column K={col_k}, Column L={col_l}")
            
            # Headers on second row
            ws.cell(row=cost_summary_row, column=header_col, value="COST/ELEVATION").font = Font(bold=True)
            ws.cell(row=cost_summary_row, column=cost_per_elev_col, value="COST/ELEVATION").font = Font(bold=True)
            ws.cell(row=cost_summary_row, column=total_elev_cost_col, value="TOTAL ELEVATION COST").font = Font(bold=True)
            
            # Add borders to headers
            for col in [header_col, cost_per_elev_col, total_elev_cost_col]:
                ws.cell(row=cost_summary_row, column=col).border = Border(bottom=Side(style='thin'))
            
            cost_summary_row += 1
            
            # Helper function to read values from Excel cells
            def read_cell_value(row, col):
                """Read cell value, return 0 if None or empty"""
                cell = ws.cell(row=row, column=col)
                return cell.value if cell.value is not None else 0.0
            
            # Profile Costs - read from column L (per elevation) and column K (total). Use 0 for door-only (no profiles).
            profile_cost_per_elev = read_cell_value(profile_totals_row, col_l) if (profile_totals_row and col_l) else (read_cell_value(profile_totals_row, col_k) / total_count if profile_totals_row else 0.0)
            profile_total_cost = read_cell_value(profile_totals_row, col_k) if profile_totals_row else 0.0
            ws.cell(row=cost_summary_row, column=header_col, value="PROFILE COSTS")
            ws.cell(row=cost_summary_row, column=cost_per_elev_col, value=profile_cost_per_elev).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws.cell(row=cost_summary_row, column=total_elev_cost_col, value=profile_total_cost).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            cost_summary_row += 1
            
            # Accessory Costs - Use 0 for door-only (no accessories).
            accessory_cost_per_elev = read_cell_value(accessory_totals_row, col_l) if (accessory_totals_row and col_l) else (read_cell_value(accessory_totals_row, col_k) / total_count if accessory_totals_row else 0.0)
            accessory_total_cost = read_cell_value(accessory_totals_row, col_k) if accessory_totals_row else 0.0
            ws.cell(row=cost_summary_row, column=header_col, value="ACCESSORY COSTS")
            ws.cell(row=cost_summary_row, column=cost_per_elev_col, value=accessory_cost_per_elev).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws.cell(row=cost_summary_row, column=total_elev_cost_col, value=accessory_total_cost).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            cost_summary_row += 1
            
            # Gasket Costs - Use 0 for door-only (no gaskets).
            gasket_cost_per_elev = read_cell_value(gasket_totals_row, col_l) if (gasket_totals_row and col_l) else (read_cell_value(gasket_totals_row, col_k) / total_count if gasket_totals_row else 0.0)
            gasket_total_cost = read_cell_value(gasket_totals_row, col_k) if gasket_totals_row else 0.0
            ws.cell(row=cost_summary_row, column=header_col, value="GASKET COSTS")
            ws.cell(row=cost_summary_row, column=cost_per_elev_col, value=gasket_cost_per_elev).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws.cell(row=cost_summary_row, column=total_elev_cost_col, value=gasket_total_cost).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            cost_summary_row += 1
            
            # Door Costs - only show if there are actually doors
            door_cost_per_elev = read_cell_value(door_totals_row, col_l) if (door_totals_row and col_l) else (read_cell_value(door_totals_row, col_k) / total_count if door_totals_row else 0.0)
            door_total_cost = read_cell_value(door_totals_row, col_k) if door_totals_row else 0.0
            
            # Display door costs when we have doors (door-only or elevation with doors)
            if door_totals_row:
                ws.cell(row=cost_summary_row, column=header_col, value="DOOR COSTS")
                ws.cell(row=cost_summary_row, column=cost_per_elev_col, value=door_cost_per_elev).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                ws.cell(row=cost_summary_row, column=total_elev_cost_col, value=door_total_cost).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                cost_summary_row += 1
            else:
                # No doors, set to 0 for calculations but don't display
                door_cost_per_elev = 0.0
                door_total_cost = 0.0
            
            # Glass Costs - sum from all glass sections
            glass_cost_per_elev = sum(read_cell_value(row, col_l) if col_l else read_cell_value(row, col_k) / total_count for row in glass_totals_rows) if glass_totals_rows else 0.0
            glass_total_cost = sum(read_cell_value(row, col_k) for row in glass_totals_rows) if glass_totals_rows else 0.0
            ws.cell(row=cost_summary_row, column=header_col, value="GLASS COSTS")
            ws.cell(row=cost_summary_row, column=cost_per_elev_col, value=glass_cost_per_elev).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws.cell(row=cost_summary_row, column=total_elev_cost_col, value=glass_total_cost).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            cost_summary_row += 1
            
            # Fabrication Costs - sum from all fabrication sections
            fabrication_cost_per_elev = sum(read_cell_value(row, col_l) if col_l else read_cell_value(row, col_k) / total_count for row in fabrication_totals_rows) if fabrication_totals_rows else 0.0
            fabrication_total_cost = sum(read_cell_value(row, col_k) for row in fabrication_totals_rows) if fabrication_totals_rows else 0.0
            ws.cell(row=cost_summary_row, column=header_col, value="FABRICATION COSTS")
            ws.cell(row=cost_summary_row, column=cost_per_elev_col, value=fabrication_cost_per_elev).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws.cell(row=cost_summary_row, column=total_elev_cost_col, value=fabrication_total_cost).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            cost_summary_row += 1
            
            # Separator line
            for col in [header_col, cost_per_elev_col, total_elev_cost_col]:
                ws.cell(row=cost_summary_row, column=col).border = Border(top=Side(style='thin'))
            cost_summary_row += 1
            
            # Total Costs - sum from column L (per elevation) and column K (total)
            # Use door costs when doors exist (door-only or elevation with doors)
            door_cost_for_total = door_cost_per_elev if door_totals_row else 0.0
            door_total_for_total = door_total_cost if door_totals_row else 0.0
            
            total_cost_per_elev = (profile_cost_per_elev + accessory_cost_per_elev + gasket_cost_per_elev + 
                                   door_cost_for_total + glass_cost_per_elev + fabrication_cost_per_elev)
            total_elevation_cost = (profile_total_cost + accessory_total_cost + gasket_total_cost + 
                                   door_total_for_total + glass_total_cost + fabrication_total_cost)
            
            ws.cell(row=cost_summary_row, column=header_col, value=f"{elev_name} TOTAL COSTS").font = Font(bold=True)
            ws.cell(row=cost_summary_row, column=cost_per_elev_col, value=total_cost_per_elev).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws.cell(row=cost_summary_row, column=cost_per_elev_col).font = Font(bold=True)
            ws.cell(row=cost_summary_row, column=total_elev_cost_col, value=total_elevation_cost).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws.cell(row=cost_summary_row, column=total_elev_cost_col).font = Font(bold=True)
            cost_summary_row += 1
            
            # Note
            note_cell = ws.cell(row=cost_summary_row, column=header_col, value="*Note - Elevation costs based on discounted material costs")
            note_cell.font = Font(italic=True, size=10)
            
            print(f"Rebuilt System Total for '{elev_name}': ${system_total_for_this_block[0]:.2f}")

            # Calculate maximum column based on which optional columns are enabled
            # Column structure: Description (5), Part Number (6), Total Quantity Required (7),
            # [Quantity Per Elevation (8) - optional], Total List Cost (9),
            # [Total List Cost Per Elevation (10) - optional], Discounted Total List Cost (11),
            # [Discounted Total List Cost Per Elevation (12) - optional]
            # Base last column is 11 (Discounted Total List Cost)
            max_col = 11  # Start with base last column (Discounted Total List Cost)
            if show_total_cost_per_elev and elev_total_count > 1:
                max_col += 1  # Total List Cost Per Elevation (column 10)
            if show_discounted_cost_per_elev and elev_total_count > 1:
                max_col += 1  # Discounted Total List Cost Per Elevation (column 12)
            
            _autofit_columns(ws, COL_A, max_col, 1, ws.max_row)
            # Keep columns A,B tight (fit by longest word) so values aren't pushed far right
            _autofit_columns_by_longest_word(ws, COL_A, COL_B, 1, min(25, ws.max_row))
            _clean_trailing_blank_rows(ws, 1)

    save_extra_materials(overall_current_extra_materials_state, private_extra_materials_path)

    summary_ws = wb.create_sheet(title="Summary")
    # Get summary settings path - always construct from elevations path to ensure consistency
    private_summary_settings_path = None
    if summary_settings_path:
        # Extract project base name from elevations path
        elev_basename = os.path.basename(private_elevations_path)
        if "_Elevations.json" in elev_basename:
            project_base = elev_basename.replace("_Elevations.json", "")
            # Construct settings path in the same directory as elevations
            private_summary_settings_path = os.path.join(private_projects_dir, f"{project_base}_Settings.json")
            print(f"[INFO] Constructed settings path from elevations: {private_summary_settings_path}")
        else:
            # Fallback to provided path
            summary_settings_path_abs = os.path.abspath(summary_settings_path)
            if os.path.exists(summary_settings_path_abs):
                private_summary_settings_path = summary_settings_path_abs
                print(f"[INFO] Using provided settings path: {private_summary_settings_path}")
            else:
                # Try in private projects dir
                private_summary_settings_path = os.path.join(private_projects_dir, os.path.basename(summary_settings_path))
                print(f"[INFO] Trying constructed path: {private_summary_settings_path}")
    else:
        # Try to construct from elevations path even if not provided
        elev_basename = os.path.basename(private_elevations_path)
        if "_Elevations.json" in elev_basename:
            project_base = elev_basename.replace("_Elevations.json", "")
            private_summary_settings_path = os.path.join(private_projects_dir, f"{project_base}_Settings.json")
            print(f"[INFO] No settings path provided, constructing from elevations: {private_summary_settings_path}")
    
    if private_summary_settings_path and os.path.exists(private_summary_settings_path):
        print(f"[OK] Settings file found: {private_summary_settings_path}")
    elif private_summary_settings_path:
        print(f"[WARNING] Settings file not found: {private_summary_settings_path}")
    
    create_summary_sheet(summary_ws, private_elevations_path, private_extra_materials_path, wb, summary_settings_path=private_summary_settings_path)
    
    final_save_path = os.path.join(public_reports_dir, os.path.basename(excel_path)) if mode == "export_all" else private_excel_path
    
    try:
        wb.save(final_save_path)
        print(f"Excel report '{final_save_path}' fully rebuilt with separate tabs.")
    except Exception as save_err:
        print(f"[ERROR] Error saving Excel report during full rebuild: {save_err}")
        if completion_callback: completion_callback(f"Error saving report: {save_err}")
        return

    if mode != "export_all":
        try:
            with open(private_elevations_path, 'w') as f:
                json.dump(current_saved_elevations, f, indent=4)
        except IOError as e:
            print(f"Error saving all elevations to {private_elevations_path} after rebuild: {e}")

    if completion_callback:
        completion_callback()