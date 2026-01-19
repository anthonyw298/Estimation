"""
PDF Export functionality for project reports.
Converts Excel sheets to PDF following specific structure.
"""
import os
import io
from datetime import datetime
from openpyxl import load_workbook

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

def get_logo_path():
    """Get the path to the company logo."""
    possible_paths = [
        os.path.join("assets", "logo.png"),
        os.path.join("assets", "logo.jpg"),
        os.path.join("assets", "logo.jpeg"),
        os.path.join("assets", "company_logo.png"),
        os.path.join("assets", "company_logo.jpg"),
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            return path
    
    return None

def format_cell_value(value, cell=None):
    """Format a cell value for PDF display."""
    if value is None:
        return ""
    
    if isinstance(value, (int, float)):
        if cell and cell.number_format:
            fmt_str = str(cell.number_format).upper()
            if '$' in fmt_str or 'CURRENCY' in fmt_str or 'ACCOUNTING' in fmt_str:
                return f"${value:,.2f}"
            elif '%' in fmt_str:
                return f"{value:.2f}%"
        
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        elif abs(value) >= 1:
            return f"{value:,.2f}"
        else:
            return f"{value:.4f}"
    
    return str(value).strip()

def extract_images_from_excel(ws):
    """Extract images from Excel worksheet."""
    images = {}
    try:
        if hasattr(ws, '_images') and ws._images:
            for img in ws._images:
                try:
                    if hasattr(img, 'anchor'):
                        anchor = img.anchor
                        if hasattr(anchor, '_from') and hasattr(anchor._from, 'row'):
                            row = anchor._from.row + 1
                        elif isinstance(anchor, str):
                            row = int(''.join(filter(str.isdigit, anchor)))
                        else:
                            continue
                        
                        if hasattr(img, '_data'):
                            images[row] = img._data()
                except:
                    continue
    except:
        pass
    return images

def is_section_header(cell_value, cell=None):
    """Determine if a cell is a section header."""
    if not cell_value:
        return False
    
    cell_str = str(cell_value).strip().upper()
    
    is_bold = cell and cell.font and cell.font.bold
    is_colored = cell and cell.fill and cell.fill.start_color and cell.fill.start_color.rgb != 'FFFFFFFF'
    
    # Exclude items that should NEVER be section headers (these are regular data rows)
    excluded_data_items = [
        'OVERHEAD MATERIALS', 'OVERHEAD LABOR', 'OVERHEAD', 
        'ADMIN AND MANAGEMENT', 'ENGINEERING', 'PACKAGING MATERIALS',
        'SHIPPING AND TRANSPORT', 'COMMISSIONS', 'COMMISSION',
        'PROFIT ON MATERIAL', 'PROFIT ON WASTE', 'PROFIT ON GLASS PURCHASE',
        'PROFIT ON WAGES', 'PLANNING / TECHNICAL OFFICE', 'PLANNING/TECHNICAL OFFICE',
        'JOINTS FABRICATION LABOR', 'FABRICATION LABOR',
        'LIST PRICE TOTAL', 'DISCOUNTED TOTAL', 'RESIDUAL/WASTE COST',
        'WASTE PERCENTAGE', 'MISCELLANEOUS', 'MARKUPS',
        'GLASS AREA', 'GLASS AREA (ADJUSTED)'
    ]
    
    # If it's one of these excluded items, it's NOT a section header (even if bold)
    for excluded in excluded_data_items:
        if excluded in cell_str or cell_str.startswith(excluded):
            return False
    
    # Special section headers that should always be treated as headers (must have colored background)
    special_section_headers = ['GRAND TOTAL', 'COST OVERVIEW', 'MISCELLANEOUS COSTS', 
                               'MARKUPS / PROFIT', 'MARKUPS/PROFIT', 'PROJECT TOTAL']
    
    # Check if this is a special section header with bold + colored styling
    for special in special_section_headers:
        if special in cell_str and is_bold and is_colored:
            return True
    
    # Exclude common column headers and totals that should never be section headers
    excluded_keywords = ['DESCRIPTION', 'PART NUMBER', 'TOTAL QUANTITY', 'QUANTITY', 
                        'TOTAL LIST COST', 'DISCOUNTED', 'ITEM', 'VALUE', 'AMOUNT',
                        'SUBTOTAL', 'RESIDUAL/WASTE', 'RESIDUAL',
                        'WASTE PERCENTAGE', 'WASTE COST']
    # Check for exact matches or if excluded keyword is the main content of the cell
    for excluded in excluded_keywords:
        if excluded in cell_str and len(cell_str.split()) <= 3:
            return False
        # Also check if cell_str starts with excluded keyword
        if cell_str.startswith(excluded):
            return False
    
    section_keywords = [
        'PROFILES', 'ACCESSORIES', 'GASKETS', 'GLASS', 'LABOR', 'DOORS',
        'COST OVERVIEW', 'PROJECT TOTAL',
        'ELEVATION SUMMARY', 'FABRICATION', 'COST/ELEVATION', 'SUMMARY',
        'SYSTEM INPUT', 'BAY DISTRIBUTION'
    ]
    
    # Only treat as section header if it matches a section keyword AND has colored background
    # This prevents items like "MISCELLANEOUS" or "MARKUPS" from being detected as headers
    if any(kw in cell_str for kw in section_keywords) and len(str(cell_value)) < 60 and is_colored:
        return True
    
    # Bold and colored cells are section headers only if they're not excluded items
    if is_bold and is_colored and len(cell_str.split()) <= 2:
        # Double-check it's not an excluded data item
        if not any(excluded in cell_str for excluded in excluded_data_items):
            return True
    
    return False

def is_table_header_row(row_data):
    """Determine if a row contains table headers."""
    if not row_data or len(row_data) < 2:
        return False
    
    row_text = ' '.join([str(v).upper() for v in row_data if v])
    header_keywords = [
        'DESCRIPTION', 'PART NUMBER', 'TOTAL QUANTITY', 'TOTAL LIST COST',
        'DISCOUNTED', 'QUANTITY REQUIRED', 'MATERIALS', 'REQUIRED',
        'PROJECT', 'FINISH', 'ELEVATION', 'NAME', 'DIMENSIONS', 'SQFT',
        'PERIMETER', 'ITEM', 'AMOUNT', 'VALUE'
    ]
    
    keyword_count = sum(1 for kw in header_keywords if kw in row_text)
    return keyword_count >= 2

def excel_to_pdf(excel_path, pdf_path, include_logo=True):
    """Convert entire Excel sheet to PDF following specific structure."""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is not installed. Install with: pip install reportlab")
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    # Load Excel workbook
    wb = load_workbook(excel_path, data_only=True)
    
    # Identify all sheets: elevation sheets and Summary sheet
    all_sheets = wb.sheetnames
    elevation_sheets = [s for s in all_sheets if s.upper() != 'SUMMARY']
    summary_sheet_name = None
    if 'Summary' in all_sheets:
        summary_sheet_name = 'Summary'
    elif 'SUMMARY' in all_sheets:
        summary_sheet_name = 'SUMMARY'
    
    print(f"[INFO] Found {len(elevation_sheets)} elevation sheet(s): {elevation_sheets}")
    if summary_sheet_name:
        print(f"[INFO] Found Summary sheet: {summary_sheet_name}")
    
    # Create PDF document
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=letter,
        rightMargin=0.4*inch,
        leftMargin=0.4*inch,
        topMargin=0.75*inch,
        bottomMargin=0.5*inch
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1A1A1A'),
        spaceAfter=15,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    section_style = ParagraphStyle(
        'SectionStyle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#1A1A1A'),
        spaceAfter=8,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#333333'),
        spaceAfter=3,
        fontName='Helvetica'
    )
    
    # Add company logo if available
    if include_logo:
        logo_path = get_logo_path()
        if logo_path:
            try:
                logo = Image(logo_path, width=2.5*inch, height=1*inch)
                logo.hAlign = 'CENTER'
                story.append(logo)
                story.append(Spacer(1, 0.15*inch))
            except:
                pass
    
    # Add title
    project_name = os.path.basename(excel_path).replace('.xlsx', '').replace('_', ' ').title()
    story.append(Paragraph("PROJECT ESTIMATION REPORT", title_style))
    story.append(Paragraph(f"<b>Project:</b> {project_name}", normal_style))
    story.append(Paragraph(f"<b>Date:</b> {datetime.now().strftime('%B %d, %Y')}", normal_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Helper function to process a single elevation sheet
    def process_elevation_sheet(ws, sheet_name, story, section_style, normal_style):
        """Process a single elevation sheet."""
        excel_images = extract_images_from_excel(ws)
        bay_diagrams = []
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"[INFO] Processing sheet '{sheet_name}': {max_row} rows x {max_col} columns")
        
        # Add elevation sheet title
        story.append(Paragraph(f"<b>{sheet_name}</b>", section_style))
        story.append(Spacer(1, 0.2*inch))
        
        # STEP 1: Process System Input (Column A-B, Rows 1-15)
        print("[INFO] Processing System Input (Columns A-B, Rows 1-15)")
        input_table = []
        for row_idx in range(1, 16):  # Rows 1-15
            label = format_cell_value(ws.cell(row=row_idx, column=1).value)
            value = format_cell_value(ws.cell(row=row_idx, column=2).value)
            if label or value:
                input_table.append([label, value])
        
        if input_table:
            story.append(Paragraph("<b>System Input</b>", section_style))
            story.append(Spacer(1, 0.1*inch))
            story.extend(create_table(input_table, ['Item', 'Value'], normal_style))
        
        # STEP 2: Process sections starting at Column E (Profiles, Accessories, etc.)
        print("[INFO] Processing sections from Column E")
        
        # Find where sections start (look for section headers in column E)
        sections_found = []
        fabrication_row = None
        
        for row_idx in range(1, max_row + 1):
            cell_e = ws.cell(row=row_idx, column=5)  # Column E
            cell_value = cell_e.value
            
            # Skip if this is clearly a table header row (not a section header)
            if cell_value:
                # Read the row to check if it's a table header
                row_data = []
                for col_idx in range(5, min(max_col + 1, 15)):
                    row_data.append(format_cell_value(ws.cell(row=row_idx, column=col_idx).value))
                if is_table_header_row(row_data):
                    continue  # Skip table headers, they're not section headers
            
            if cell_value and is_section_header(cell_value, cell_e):
                section_name = str(cell_value)
                sections_found.append((row_idx, section_name))
                if 'FABRICATION' in section_name.upper():
                    fabrication_row = row_idx
        
        # Process each section from Column E
        current_table = []
        current_headers = None
        last_header_row = None
        
        for section_idx, (start_row, section_name) in enumerate(sections_found):
            # Determine end row (next section or end of sheet)
            if section_idx + 1 < len(sections_found):
                end_row = sections_found[section_idx + 1][0] - 1
            else:
                end_row = max_row
            
            print(f"[INFO] Processing section: {section_name} (rows {start_row}-{end_row})")
            
            # Add section header
            story.append(Paragraph(f"<b>{section_name}</b>", section_style))
            story.append(Spacer(1, 0.1*inch))
            
            # Process rows in this section
            for row_idx in range(start_row, end_row + 1):
                # Check for images
                if row_idx in excel_images:
                    img_data = excel_images[row_idx]
                    # Bay diagrams are usually in elevation sections
                    if 'BAY' in str(img_data) or row_idx < (start_row + end_row) / 2:
                        bay_diagrams.append((row_idx, img_data))
                    continue
                
                # Read row from Column E onwards
                row_data = []
                row_cells = []
                for col_idx in range(5, min(max_col + 1, 15)):  # Start at column E (5)
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = format_cell_value(cell.value, cell)
                    row_data.append(value)
                    row_cells.append(cell)
                
                if not any([str(v).strip() for v in row_data if v]):
                    if current_table and current_headers:
                        story.extend(create_table(current_table, current_headers, normal_style))
                        current_table = []
                        current_headers = None
                    continue
                
                # Check for table header
                if is_table_header_row(row_data):
                    if current_table and current_headers:
                        story.extend(create_table(current_table, current_headers, normal_style))
                        current_table = []
                    new_headers = [h.strip() for h in row_data if h and str(h).strip()]
                    if new_headers != current_headers:
                        current_headers = new_headers
                        last_header_row = row_idx
                    continue
                
                # Skip if this row was already used as a header
                if row_idx == last_header_row:
                    continue
                
                # Add data row
                if current_headers:
                    clean_row = row_data[:len(current_headers)] if len(row_data) >= len(current_headers) else row_data
                    while len(clean_row) < len(current_headers):
                        clean_row.append("")
                    if len([v for v in clean_row if v and str(v).strip()]) >= 1:
                        current_table.append(clean_row)
            
            # Add table for this section
            if current_table and current_headers:
                story.extend(create_table(current_table, current_headers, normal_style))
                current_table = []
                current_headers = None
        
        # Add bay diagrams after sections
        for row_idx, img_data in bay_diagrams:
            try:
                if isinstance(img_data, bytes):
                    img_bytes = io.BytesIO(img_data)
                    img = Image(img_bytes, width=5*inch, height=3.75*inch)
                    img.hAlign = 'CENTER'
                    story.append(img)
                    story.append(Spacer(1, 0.2*inch))
            except:
                pass
        
        story.append(Spacer(1, 0.3*inch))  # Space between elevation sheets
    
    # Process all elevation sheets
    bay_diagrams = []
    pie_chart = None
    
    for sheet_name in elevation_sheets:
        ws = wb[sheet_name]
        process_elevation_sheet(ws, sheet_name, story, section_style, normal_style)
    
    # STEP 3: Process Summary sheet separately if it exists
    if summary_sheet_name:
        ws_summary = wb[summary_sheet_name]
        print(f"[INFO] Processing Summary sheet separately: {summary_sheet_name}")
        story.append(PageBreak())
        story.append(Paragraph("<b>SUMMARY</b>", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        excel_images_summary = extract_images_from_excel(ws_summary)
        max_row_summary = ws_summary.max_row
        max_col_summary = ws_summary.max_column
        
        # Find pie chart in summary
        for row_idx in range(1, max_row_summary + 1):
            if row_idx in excel_images_summary:
                if pie_chart is None:
                    pie_chart = (row_idx, excel_images_summary[row_idx])
        
        # Process summary content
        current_table = []
        current_headers = None
        last_header_row = None  # Track to avoid processing header row as data
        
        for row_idx in range(1, max_row_summary + 1):
            if row_idx in excel_images_summary and pie_chart and row_idx == pie_chart[0]:
                continue
            
            row_data = []
            row_cells = []
            row_bold_flags = []  # Track which cells are bold
            # Read more columns to ensure we get column C (index 2) for MISCELLANEOUS COSTS
            for col_idx in range(1, min(max_col_summary + 1, 10)):
                cell = ws_summary.cell(row=row_idx, column=col_idx)
                value = format_cell_value(cell.value, cell)
                is_bold = cell.font and cell.font.bold if cell.font else False
                row_data.append(value)
                row_cells.append(cell)
                row_bold_flags.append(is_bold)
            
            if not any([str(v).strip() for v in row_data if v]):
                # Don't close table on empty rows - just skip them
                # Tables should only close when we hit a new section header
                continue
            
            first_cell = row_cells[0] if row_cells else None
            first_value = row_data[0] if row_data else None
            
            if first_value and is_section_header(first_value, first_cell):
                # Close previous table if exists
                if current_table and current_headers:
                    story.extend(create_table(current_table, current_headers, normal_style))
                    current_table = []
                    current_headers = None
                    last_header_row = None
                
                # Start new section
                story.append(Paragraph(f"<b>{first_value}</b>", section_style))
                story.append(Spacer(1, 0.1*inch))
                
                # Special handling for GRAND TOTAL - value is in same row (column C)
                cell_str = str(first_value).strip().upper()
                if 'GRAND TOTAL' in cell_str:
                    # Check if value is in column C of same row
                    if len(row_data) > 2 and row_data[2] and str(row_data[2]).strip():
                        grand_total_value = row_data[2]
                        grand_total_bold = row_bold_flags[2] if len(row_bold_flags) > 2 else False
                        # Display as a simple Item/Value table with just one row
                        grand_table = [[(first_value, True), (grand_total_value, grand_total_bold)]]
                        story.extend(create_table(grand_table, ['Item', 'Value'], normal_style))
                        current_table = []
                        current_headers = None
                        last_header_row = None
                        continue
                
                # For summary sections (COST OVERVIEW, MISCELLANEOUS, etc.), set up Item/Value headers
                # Check if next row is a table header, if not, default to Item/Value
                if row_idx + 1 <= max_row_summary:
                    next_row_data = []
                    for col_idx in range(1, min(max_col_summary + 1, 15)):
                        next_row_data.append(format_cell_value(ws_summary.cell(row=row_idx + 1, column=col_idx).value))
                    if not is_table_header_row(next_row_data):
                        # This section uses Item/Value format
                        current_headers = ['Item', 'Value']
                        last_header_row = None
                continue
            
            # For Summary sheet, only treat as table header if we're NOT in an Item/Value section
            # Item/Value sections (COST OVERVIEW, MISCELLANEOUS, etc.) already have headers set
            # and should continue until the next section header
            is_item_value_section = (current_headers and len(current_headers) == 2 and 
                                     current_headers[0].upper() == 'ITEM' and 
                                     current_headers[1].upper() == 'VALUE')
            
            if is_table_header_row(row_data) and not is_item_value_section:
                if current_table and current_headers:
                    story.extend(create_table(current_table, current_headers, normal_style))
                    current_table = []
                # Only update headers if they're different (avoid duplicates)
                new_headers = [h.strip() for h in row_data if h and str(h).strip()]
                if new_headers != current_headers:
                    current_headers = new_headers
                    last_header_row = row_idx
                continue
            
            # Skip if this row was already used as a header
            if row_idx == last_header_row:
                continue
            
            if current_headers:
                # For Item/Value format, handle values that might be in column C instead of B
                if len(current_headers) == 2 and current_headers[0].upper() == 'ITEM' and current_headers[1].upper() == 'VALUE':
                    label = row_data[0] if len(row_data) > 0 else ""
                    # Check actual cell for bold (column 1 = column A)
                    try:
                        label_cell = ws_summary.cell(row=row_idx, column=1)
                        label_bold = bool(label_cell.font and label_cell.font.bold) if label_cell.font else False
                    except:
                        label_bold = False
                    
                    value = ""
                    value_bold = False
                    value_excel_col = -1
                    # Try column C (3) first (MISCELLANEOUS COSTS stores values in column C)
                    # Check row_data first (simpler, already formatted)
                    if len(row_data) > 2 and row_data[2] and str(row_data[2]).strip():
                        value = row_data[2]
                        value_excel_col = 3
                        # Get bold flag from row_bold_flags if available
                        if len(row_bold_flags) > 2:
                            value_bold = row_bold_flags[2]
                    # If not in row_data, try reading from row_cells directly
                    elif len(row_cells) > 2:
                        cell_c = row_cells[2]
                        if cell_c and cell_c.value is not None:
                            val_str = format_cell_value(cell_c.value, cell_c)
                            if val_str and str(val_str).strip():
                                value = val_str
                                value_excel_col = 3
                                value_bold = bool(cell_c.font and cell_c.font.bold) if cell_c.font else False
                    # If not found in C, try column B (2) as fallback
                    if not value:
                        if len(row_data) > 1 and row_data[1] and str(row_data[1]).strip():
                            value = row_data[1]
                            value_excel_col = 2
                            if len(row_bold_flags) > 1:
                                value_bold = row_bold_flags[1]
                        elif len(row_cells) > 1:
                            cell_b = row_cells[1]
                            if cell_b and cell_b.value is not None:
                                val_str = format_cell_value(cell_b.value, cell_b)
                                if val_str and str(val_str).strip():
                                    value = val_str
                                    value_excel_col = 2
                                    value_bold = bool(cell_b.font and cell_b.font.bold) if cell_b.font else False
                    
                    # Get bold flag from the actual cell we found the value in
                    if value_excel_col > 0:
                        try:
                            value_cell = ws_summary.cell(row=row_idx, column=value_excel_col)
                            value_bold = bool(value_cell.font and value_cell.font.bold) if value_cell.font else False
                        except:
                            # Try to get bold from row_cells if we found the value there
                            col_idx = value_excel_col - 1  # Convert to 0-based
                            if col_idx < len(row_bold_flags):
                                value_bold = row_bold_flags[col_idx]
                            else:
                                value_bold = False
                    
                    # Always add row if we have a label (even if value is empty, it's still a valid row)
                    # This ensures all rows in a section are captured (e.g., all 7 MISCELLANEOUS items, all 4 COST OVERVIEW rows)
                    if label and str(label).strip():  
                        current_table.append([(label, label_bold), (value, value_bold)])
                else:
                    # For other header formats, use standard logic with bold formatting
                    formatted_row = []
                    for col_idx in range(len(current_headers)):
                        if col_idx < len(row_data):
                            value = row_data[col_idx]
                            # Check actual Excel cell for bold (col_idx + 1 because Excel columns are 1-indexed)
                            excel_col = col_idx + 1
                            try:
                                cell = ws_summary.cell(row=row_idx, column=excel_col)
                                is_bold = bool(cell.font and cell.font.bold) if cell.font else False
                            except:
                                is_bold = False
                            formatted_row.append((value, is_bold))
                        else:
                            formatted_row.append(("", False))
                    if len([v for v in formatted_row if v[0] and str(v[0]).strip()]) >= 1:
                        current_table.append(formatted_row)
            elif len([v for v in row_data if v]) >= 1:
                # For Item/Value format, find the value - it might be in column C (index 2) instead of B (index 1)
                label = row_data[0] if len(row_data) > 0 else ""
                # Check actual cell for bold (column 1 = column A)
                label_cell = ws_summary.cell(row=row_idx, column=1)
                label_bold = label_cell.font and label_cell.font.bold if label_cell.font else False
                
                value = ""
                value_bold = False
                value_excel_col = -1
                
                # Try column C (3) first (MISCELLANEOUS COSTS stores values in column C)
                # Check row_data first (simpler, already formatted)
                if len(row_data) > 2 and row_data[2] and str(row_data[2]).strip():
                    value = row_data[2]
                    value_excel_col = 3
                    # Get bold flag from row_bold_flags if available
                    if len(row_bold_flags) > 2:
                        value_bold = row_bold_flags[2]
                # If not in row_data, try reading from row_cells directly
                elif len(row_cells) > 2:
                    cell_c = row_cells[2]
                    if cell_c and cell_c.value is not None:
                        val_str = format_cell_value(cell_c.value, cell_c)
                        if val_str and str(val_str).strip():
                            value = val_str
                            value_excel_col = 3
                            value_bold = bool(cell_c.font and cell_c.font.bold) if cell_c.font else False
                # If not found in C, try column B (2) as fallback
                if not value:
                    if len(row_data) > 1 and row_data[1] and str(row_data[1]).strip():
                        value = row_data[1]
                        value_excel_col = 2
                        if len(row_bold_flags) > 1:
                            value_bold = row_bold_flags[1]
                    elif len(row_cells) > 1:
                        cell_b = row_cells[1]
                        if cell_b and cell_b.value is not None:
                            val_str = format_cell_value(cell_b.value, cell_b)
                            if val_str and str(val_str).strip():
                                value = val_str
                                value_excel_col = 2
                                value_bold = bool(cell_b.font and cell_b.font.bold) if cell_b.font else False
                
                # Fallback: check actual cell for bold if we found a value
                if value_excel_col > 0 and not value_bold:
                    try:
                        value_cell = ws_summary.cell(row=row_idx, column=value_excel_col)
                        value_bold = bool(value_cell.font and value_cell.font.bold) if value_cell.font else False
                    except:
                        pass
                
                # Always add row if we have a label (even if value is empty)
                # This ensures all rows in a section are captured until next section header
                if label and str(label).strip():
                    # Only create Item/Value headers if we don't already have headers set
                    if not current_headers:
                        current_headers = ['Item', 'Value']
                        last_header_row = None  # No actual header row, just default headers
                    current_table.append([(label, label_bold), (value, value_bold)])
        
        # Close final table
        if current_table and current_headers:
            story.extend(create_table(current_table, current_headers, normal_style))
    
    # Add pie chart at the very end
    if pie_chart:
        story.append(Spacer(1, 0.4*inch))
        story.append(Paragraph("<b>Project Cost Breakdown</b>", section_style))
        story.append(Spacer(1, 0.2*inch))
        try:
            row_idx, img_data = pie_chart
            if isinstance(img_data, bytes):
                img_bytes = io.BytesIO(img_data)
                img = Image(img_bytes, width=5*inch, height=3.75*inch)
                img.hAlign = 'CENTER'
                story.append(img)
        except Exception as e:
            print(f"[WARNING] Could not add pie chart: {e}")
    
    # Build PDF
    doc.build(story)
    print(f"[OK] PDF exported to: {pdf_path}")

def create_table(table_data, headers, normal_style):
    """Create a formatted table from data."""
    elements = []
    
    if not table_data or not headers:
        return elements
    
    # Build table
    full_data = []
    
    # Add headers with wrapping
    wrapped_headers = []
    for h in headers:
        if h:
            if len(h) > 20:
                words = h.split()
                lines = []
                current = ""
                for word in words:
                    test = current + " " + word if current else word
                    if len(test) <= 20:
                        current = test
                    else:
                        if current:
                            lines.append(current)
                        current = word
                if current:
                    lines.append(current)
                wrapped_headers.append(Paragraph("<br/>".join(lines), normal_style))
            else:
                wrapped_headers.append(Paragraph(h, normal_style))
        else:
            wrapped_headers.append("")
    
    full_data.append(wrapped_headers)
    
    # Add data rows
    for row in table_data:
        formatted = []
        for i, cell in enumerate(row):
            if i < len(headers):
                # Handle both regular strings and tuples of (value, is_bold)
                cell_value = cell
                is_bold = False
                if isinstance(cell, tuple) and len(cell) == 2:
                    cell_value, is_bold = cell
                
                if cell_value is not None:
                    cell_str = str(cell_value).strip()
                    if cell_str:
                        # Apply bold formatting if needed
                        if is_bold:
                            formatted.append(Paragraph(f"<b>{cell_str}</b>", normal_style))
                        else:
                            formatted.append(Paragraph(cell_str, normal_style))
                    else:
                        formatted.append("")
                else:
                    formatted.append("")
        while len(formatted) < len(headers):
            formatted.append("")
        full_data.append(formatted[:len(headers)])
    
    if not full_data:
        return elements
    
    # Calculate column widths
    num_cols = len(headers)
    total_width = 7.2 * inch
    
    if num_cols <= 3:
        col_widths = [total_width / num_cols] * num_cols
    elif num_cols <= 5:
        desc_width = total_width * 0.3
        other_width = (total_width - desc_width) / (num_cols - 1)
        col_widths = [desc_width] + [other_width] * (num_cols - 1)
    else:
        desc_width = total_width * 0.22
        other_width = (total_width - desc_width) / (num_cols - 1)
        col_widths = [desc_width] + [other_width] * (num_cols - 1)
    
    # Create table with page break support
    table = Table(full_data, colWidths=col_widths[:num_cols], repeatRows=1, splitByRow=1)
    
    # Table style
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8E8E8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1A1A1A')),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('LEFTPADDING', (0, 0), (-1, 0), 4),
        ('RIGHTPADDING', (0, 0), (-1, 0), 4),
        ('BOTTOMBORDER', (0, 0), (-1, 0), 1, colors.HexColor('#666666')),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('LEFTPADDING', (0, 1), (-1, -1), 4),
        ('RIGHTPADDING', (0, 1), (-1, -1), 4),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
    ]
    
    table.setStyle(TableStyle(table_style))
    elements.append(table)
    elements.append(Spacer(1, 0.2*inch))
    
    return elements

def export_project_to_pdf(project_name, excel_path=None, output_dir="reports", include_logo=True):
    """Export a project to PDF."""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is not installed. Install with: pip install reportlab")
    
    # Find Excel file if not provided
    if excel_path is None:
        excel_path = os.path.join(".files", project_name, f"{project_name}_Report.xlsx")
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel report not found for project: {project_name}")
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate PDF filename
    pdf_filename = f"{project_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf_path = os.path.join(output_dir, pdf_filename)
    
    # Generate PDF
    excel_to_pdf(excel_path, pdf_path, include_logo=include_logo)
    
    return pdf_path
