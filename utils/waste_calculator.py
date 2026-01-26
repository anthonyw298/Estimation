"""
Waste Calculator Enhancement Module
Provides visual waste percentage impact, waste breakdown by material type, and optimization suggestions.
"""
import os
import json
import math
from typing import Dict, List, Tuple, Optional
from data.parts_data import parts_data
from utils.pricing import parse_length_to_feet, get_unit_price_by_part, load_extra_materials
from collections import Counter

def read_waste_from_excel(excel_path: str) -> Dict:
    """
    Read waste statistics directly from Excel Summary sheet.
    
    Returns:
        Dictionary with waste statistics read from Excel:
        - total_waste_cost: Residual/Waste Cost from Excel
        - total_material_cost: Discounted Total from Excel
        - overall_waste_percentage: Waste Percentage from Excel
    """
    try:
        from openpyxl import load_workbook
        
        if not os.path.exists(excel_path):
            return None
        
        wb = load_workbook(excel_path, data_only=True)
        
        # Check if Summary sheet exists
        if "Summary" not in wb.sheetnames:
            if "SUMMARY" in wb.sheetnames:
                ws = wb["SUMMARY"]
            else:
                return None
        else:
            ws = wb["Summary"]
        
        # Search for "COST OVERVIEW" to find the start position
        overview_row = None
        overview_col = None
        
        for row in range(1, min(ws.max_row + 1, 100)):  # Search first 100 rows
            for col in range(1, min(ws.max_column + 1, 10)):  # Search first 10 columns
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and "COST OVERVIEW" in str(cell_value).upper():
                    overview_row = row
                    overview_col = col
                    break
            if overview_row:
                break
        
        if not overview_row:
            return None
        
        # Read values from relative positions
        # Discounted Total is at overview_row+2, overview_col+2
        # Residual/Waste Cost is at overview_row+3, overview_col+2
        # Waste Percentage is at overview_row+4, overview_col+2
        
        discounted_total_cell = ws.cell(row=overview_row + 2, column=overview_col + 2)
        waste_cost_cell = ws.cell(row=overview_row + 3, column=overview_col + 2)
        waste_percentage_cell = ws.cell(row=overview_row + 4, column=overview_col + 2)
        
        # Extract values
        total_material_cost = float(discounted_total_cell.value) if discounted_total_cell.value else 0.0
        total_waste_cost = float(waste_cost_cell.value) if waste_cost_cell.value else 0.0
        
        # Extract waste percentage (remove % sign if present)
        waste_pct_str = str(waste_percentage_cell.value) if waste_percentage_cell.value else "0"
        waste_pct_str = waste_pct_str.replace("%", "").strip()
        overall_waste_percentage = float(waste_pct_str) if waste_pct_str else 0.0
        
        return {
            "total_waste_cost": total_waste_cost,
            "total_material_cost": total_material_cost,
            "overall_waste_percentage": overall_waste_percentage
        }
    except Exception as e:
        print(f"[WARNING] Error reading waste from Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

def calculate_waste_statistics(project_path: str = None, extra_materials_path: str = None, excel_path: Optional[str] = None, elevations_data: Optional[Dict] = None, extra_materials: Optional[Dict] = None) -> Dict:
    """
    Calculate comprehensive waste statistics for a project.
    
    Can accept either file paths OR data dictionaries directly.
    If data dictionaries are provided, they take precedence over file paths.
    
    Args:
        project_path: Path to elevations JSON file (optional if elevations_data provided)
        extra_materials_path: Path to extra materials JSON file (optional if extra_materials provided)
        excel_path: Optional path to Excel file for reading waste stats
        elevations_data: Optional dictionary of elevations data (takes precedence over project_path)
        extra_materials: Optional dictionary of extra materials (takes precedence over extra_materials_path)
    
    Returns:
        Dictionary with waste statistics including:
        - total_waste_cost: Total cost of waste materials
        - total_material_cost: Total cost of materials
        - overall_waste_percentage: Overall waste percentage
        - material_breakdown: List of waste data per material
        - suggestions: List of optimization suggestions
    """
    # Load data from dictionaries if provided, otherwise from files
    if elevations_data is None:
        if project_path and os.path.exists(project_path):
            try:
                with open(project_path, 'r') as f:
                    elevations_data = json.load(f)
            except Exception as e:
                print(f"[ERROR] Error loading elevations from file: {e}")
                elevations_data = {}
        else:
            elevations_data = {}
    
    if extra_materials is None:
        if extra_materials_path and os.path.exists(extra_materials_path):
            try:
                extra_materials = load_extra_materials(extra_materials_path)
            except Exception as e:
                print(f"[ERROR] Error loading extra materials from file: {e}")
                extra_materials = {}
        else:
            extra_materials = {}
    
    if not elevations_data and not extra_materials:
        return {
            "total_waste_cost": 0.0,
            "total_material_cost": 0.0,
            "overall_waste_percentage": 0.0,
            "material_breakdown": [],
            "suggestions": []
        }
    
    try:
        # Debug output
        print(f"[Waste Calculator] Loaded {len(elevations_data)} elevations, {len(extra_materials)} extra materials")
    except Exception as e:
        print(f"[ERROR] Error loading waste data: {e}")
        import traceback
        traceback.print_exc()
        return {
            "total_waste_cost": 0.0,
            "total_material_cost": 0.0,
            "overall_waste_percentage": 0.0,
            "material_breakdown": [],
            "suggestions": []
        }
    
    # Calculate waste per material
    material_breakdown = []
    total_waste_cost = 0.0
    total_material_cost = 0.0
    
    # Process each material in extra_materials (these are the waste/leftovers)
    for material_key, material_data in extra_materials.items():
        if not material_data:
            continue
        
        # Extract part number and finish from key (format: "BE9-2513-clear" or "E1-0199" or "PM-1006-SS")
        parts = material_key.split('-')
        # Common finish codes
        finish_codes = ['clear', 'bronze', 'grey', 'black', 'white']
        
        if len(parts) >= 3:
            # Check if last part is a finish code
            if parts[-1].lower() in finish_codes:
                # Format: "BE9-2513-clear" -> part_number = "BE9-2513", finish = "clear"
                part_number = '-'.join(parts[:-1])
                finish = parts[-1].lower()
            else:
                # Format: "PM-1006-SS" -> part_number = "PM-1006-SS", finish = ""
                part_number = material_key
                finish = ""
        elif len(parts) == 2:
            # Could be "E1-0199" (no finish) or "BE9-clear" (finish)
            if parts[1].lower() in finish_codes:
                part_number = parts[0]
                finish = parts[1].lower()
            else:
                # No finish code, entire key is part number
                part_number = material_key
                finish = ""
        else:
            part_number = material_key
            finish = ""
        
        # Get material info
        part_info = parts_data.get(part_number, {})
        description = part_info.get('Description', part_number)
        
        # Calculate waste quantity and cost
        length_pieces = material_data.get('length_pieces', [])
        quantity = material_data.get('quantity', 0.0)
        
        # Calculate waste quantity (from leftover pieces or quantity)
        waste_qty = 0.0
        if length_pieces and len(length_pieces) > 0:
            # For profiles/gaskets, waste is in length_pieces
            length_str = part_info.get('Length', '')
            min_purchase_length = parse_length_to_feet(length_str) or 24.0
            valid_lengths = []
            for l in length_pieces:
                try:
                    l_float = float(l)
                    # Valid leftover: > 0 and < min_purchase_length
                    if 0 < l_float < min_purchase_length:
                        valid_lengths.append(l_float)
                except (TypeError, ValueError):
                    pass
            waste_qty = sum(valid_lengths)
        elif quantity and float(quantity) > 0:
            # For accessories, waste is in quantity field
            try:
                waste_qty = float(quantity)
            except (TypeError, ValueError):
                waste_qty = 0.0
        
        if waste_qty <= 0:
            continue
        
        # Calculate material cost (using unit price)
        unit_price, _ = get_unit_price_by_part(
            part_number, 
            finish=finish, 
            extra_materials_file=extra_materials_path
        )
        
        # Apply multiplier based on total cost (same logic as Excel generator)
        # Get total project cost to determine multiplier
        total_project_cost = 0.0
        for elev_key, elev in elevations_data.items():
            for output in elev.get('calculated_outputs', []):
                qty = output.get('quantity', 0)
                if isinstance(qty, list):
                    qty_val = sum([float(x) for x in qty if x])
                else:
                    qty_val = float(qty) if qty else 0.0
                price = output.get('price', 0.0)
                total_project_cost += qty_val * float(price) if price else 0.0
        
        # Use same multiplier logic as Excel generator (0.614 if < 50000, 0.572 if >= 50000)
        multiplier = 0.614 if total_project_cost < 50000 else 0.572
        waste_cost = waste_qty * unit_price * multiplier if unit_price else 0.0
        
        # Calculate total material quantity used for this part (sum across all elevations)
        # This is the quantity that was actually used (before waste)
        total_qty_used = 0.0
        for elev_key, elev in elevations_data.items():
            elev_finish = elev.get('finish', '').lower()
            for output in elev.get('calculated_outputs', []):
                output_part = output.get('part_number', '').strip()
                # Match part number (exact match or match without finish suffix)
                part_matches = False
                if output_part == part_number:
                    part_matches = True
                elif not finish and output_part.startswith(part_number):
                    # If no finish in key, try partial match (e.g., "PM-1006-SS" matches "PM-1006-SS")
                    part_matches = True
                
                if part_matches:
                    # If material key has finish, check if elevation finish matches
                    if finish and elev_finish != finish.lower():
                        continue
                    qty = output.get('quantity', 0)
                    if isinstance(qty, list):
                        total_qty_used += sum([float(x) for x in qty if x and x > 0])
                    else:
                        try:
                            total_qty_used += float(qty) if qty else 0.0
                        except (TypeError, ValueError):
                            pass
        
        # If we don't have usage data, try to estimate from purchased quantity
        # Purchased quantity = used + waste
        total_qty_purchased = total_qty_used + waste_qty
        
        # Skip if no waste and no usage
        if waste_qty <= 0 and total_qty_used <= 0:
            continue
        
        # If we have waste but no usage data, debug the matching issue
        if total_qty_used <= 0 and waste_qty > 0:
            # Try more lenient matching - check if any part_number starts with our part
            for elev_key, elev in elevations_data.items():
                for output in elev.get('calculated_outputs', []):
                    output_part = output.get('part_number', '').strip()
                    # Try exact match, or partial match (for cases like "PM-1006-SS")
                    if output_part == part_number or output_part == material_key:
                        qty = output.get('quantity', 0)
                        if isinstance(qty, list):
                            total_qty_used += sum([float(x) for x in qty if x and x > 0])
                        else:
                            try:
                                total_qty_used += float(qty) if qty else 0.0
                            except (TypeError, ValueError):
                                pass
            
            # If still no usage found, skip to avoid 100% waste calculation
            if total_qty_used <= 0:
                print(f"[WARNING] Waste Calculator: Could not match usage data for {material_key} (waste: {waste_qty}, tried part_number: {part_number})")
                continue
            
            # Recalculate purchased quantity with found usage
            total_qty_purchased = total_qty_used + waste_qty
        
        # Calculate waste percentage for this material
        # Waste percentage = waste / (used + waste) * 100
        if total_qty_purchased > 0:
            waste_percentage = (waste_qty / total_qty_purchased * 100)
        else:
            waste_percentage = 0.0
        
        # Calculate material cost for used materials only (not including waste)
        # This matches the Excel calculation: waste_cost / total_discounted_price * 100
        # where total_discounted_price is the cost of USED materials only
        used_material_cost = total_qty_used * unit_price * multiplier if unit_price else 0.0
        
        material_breakdown.append({
            "part_number": part_number,
            "description": description,
            "finish": finish,
            "total_quantity": total_qty_used,
            "waste_quantity": waste_qty,
            "waste_percentage": waste_percentage,
            "waste_cost": waste_cost,
            "material_cost": used_material_cost,
            "unit": "ft" if length_pieces and len(length_pieces) > 0 else "pcs"
        })
        
        total_waste_cost += waste_cost
        total_material_cost += used_material_cost  # Only count used materials, not waste
    
    # Try to read waste statistics directly from Excel Summary sheet (more accurate)
    excel_waste_data = None
    if excel_path and os.path.exists(excel_path):
        excel_waste_data = read_waste_from_excel(excel_path)
    
    # Use Excel values if available, otherwise calculate from data
    if excel_waste_data:
        overall_waste_percentage = excel_waste_data["overall_waste_percentage"]
        total_waste_cost = excel_waste_data["total_waste_cost"]
        total_material_cost = excel_waste_data["total_material_cost"]
        print(f"[OK] Waste Calculator: Using values from Excel (waste: {overall_waste_percentage:.2f}%, cost: ${total_waste_cost:.2f})")
    else:
        # Calculate overall waste percentage to match Excel report
        # Excel uses: waste_cost / total_discounted_price * 100
        # where total_discounted_price is the cost of USED materials (not including waste)
        overall_waste_percentage = (total_waste_cost / total_material_cost * 100) if total_material_cost > 0 else 0.0
        print(f"[WARNING] Waste Calculator: Excel not available, calculated waste: {overall_waste_percentage:.2f}%")
    
    # Generate optimization suggestions with elevation context
    suggestions = generate_optimization_suggestions(
        material_breakdown, 
        overall_waste_percentage, 
        elevations_data=elevations_data,
        extra_materials=extra_materials
    )
    
    return {
        "total_waste_cost": total_waste_cost,
        "total_material_cost": total_material_cost,
        "overall_waste_percentage": overall_waste_percentage,
        "material_breakdown": sorted(material_breakdown, key=lambda x: x['waste_cost'], reverse=True),
        "suggestions": suggestions
    }

def _generate_bay_width_suggestion(material: Dict, elevations_data: Dict, stock_length_ft: float, part_number: str) -> Optional[Dict]:
    """Generate specific bay width optimization suggestions for horizontal parts."""
    if not elevations_data:
        return None
    
    # Find elevations using this material
    elevs_with_material = []
    for elev_name, elev_data in elevations_data.items():
        finish = elev_data.get('finish', '').lower()
        if material.get('finish', '').lower() != finish:
            continue
        
        # Check if this elevation uses the part
        for output in elev_data.get('calculated_outputs', []):
            if output.get('part_number', '').strip() == part_number:
                bays_wide = elev_data.get('bays_wide', 0)
                custom_bay_widths = elev_data.get('custom_bay_widths', [])
                opening_width = elev_data.get('opening_width_inches', 0)
                
                if bays_wide > 0 and opening_width > 0:
                    elevs_with_material.append({
                        'name': elev_name,
                        'bays_wide': bays_wide,
                        'custom_bay_widths': custom_bay_widths,
                        'opening_width': opening_width,
                        'quantity': output.get('quantity', [])
                    })
                break
    
    if not elevs_with_material:
        return None
    
    # Analyze bay widths for optimization opportunities
    top_elev = elevs_with_material[0]
    waste_qty = material.get('waste_quantity', 0)
    waste_pct = material.get('waste_percentage', 0)
    waste_cost = material.get('waste_cost', 0)
    finish_text = f" ({material.get('finish', '').capitalize()})" if material.get('finish') else ""
    
    if top_elev['custom_bay_widths'] and len(top_elev['custom_bay_widths']) == top_elev['bays_wide']:
        # Has custom bay widths - analyze for optimization
        bay_widths_inches = top_elev['custom_bay_widths']
        bay_widths_ft = [w / 12.0 for w in bay_widths_inches]
        total_width_ft = sum(bay_widths_ft)
        
        # Check if adjusting bay widths could reduce waste
        avg_waste_per_bay = waste_qty / len(bay_widths_ft) if bay_widths_ft else 0
        
        if waste_pct > 20 and avg_waste_per_bay > 0.5:
            # Suggest specific bay width adjustments
            suggestion = {
                "priority": "high" if waste_pct > 30 else "medium",
                "category": "Bay Width Optimization",
                "message": f"{material.get('description', 'Unknown')}{finish_text} in '{top_elev['name']}' has {waste_pct:.1f}% waste (${waste_cost:.2f}). Current bay widths: {', '.join([f'{w:.1f}\"' for w in bay_widths_inches[:5]])}. Average waste per bay: {avg_waste_per_bay:.2f}ft. Consider adjusting bay widths to better utilize {stock_length_ft:.0f}ft stock lengths. For example, try distributing widths more evenly or adjusting individual bays by +/-2-3\" to reduce leftover pieces.",
                "estimated_savings": waste_cost * 0.25
            }
            return suggestion
    
    return None

def _generate_height_suggestion(material: Dict, elevations_data: Dict, stock_length_ft: float, part_number: str) -> Optional[Dict]:
    """Generate specific height optimization suggestions for vertical parts."""
    if not elevations_data:
        return None
    
    # Find elevations using this material
    elevs_with_material = []
    for elev_name, elev_data in elevations_data.items():
        finish = elev_data.get('finish', '').lower()
        if material.get('finish', '').lower() != finish:
            continue
        
        for output in elev_data.get('calculated_outputs', []):
            if output.get('part_number', '').strip() == part_number:
                opening_height = elev_data.get('opening_height_inches', 0)
                bays_tall = elev_data.get('bays_tall', 0)
                
                if opening_height > 0:
                    elevs_with_material.append({
                        'name': elev_name,
                        'opening_height': opening_height,
                        'bays_tall': bays_tall
                    })
                break
    
    if not elevs_with_material:
        return None
    
    top_elev = elevs_with_material[0]
    waste_pct = material.get('waste_percentage', 0)
    waste_cost = material.get('waste_cost', 0)
    finish_text = f" ({material.get('finish', '').capitalize()})" if material.get('finish') else ""
    height_ft = top_elev['opening_height'] / 12.0
    
    if waste_pct > 20 and height_ft > 0:
        return {
            "priority": "high" if waste_pct > 30 else "medium",
            "category": "Height Optimization",
            "message": f"{material.get('description', 'Unknown')}{finish_text} in '{top_elev['name']}' has {waste_pct:.1f}% waste (${waste_cost:.2f}). Opening height is {top_elev['opening_height']:.1f}\" ({height_ft:.2f}ft). With {stock_length_ft:.0f}ft stock lengths, consider adjusting the opening height or bay height configuration to better utilize full stock lengths and reduce waste.",
            "estimated_savings": waste_cost * 0.2
        }
    
    return None

def _calculate_optimal_cuts_for_leftover(leftover_ft: float, stock_length_ft: float = 24.0, min_cut_ft: float = 0.5) -> List[Dict]:
    """
    Calculate specific cut dimensions that would FIT INSIDE a leftover piece.
    Returns list of dimension combinations that would use the leftover efficiently.
    For example: If leftover is 23.75ft, suggest "23.5ft (282\")" or "12ft + 11.5ft = 23.5ft"
    """
    optimal_cuts = []
    
    if leftover_ft < min_cut_ft:
        return optimal_cuts
    
    leftover_inches = leftover_ft * 12
    tolerance = 0.5 * 12  # 0.5ft tolerance in inches
    
    # Generate cut combinations that would fit inside the leftover
    # Try various combinations: single cut, two cuts, three cuts, etc.
    
    # Single cut - use most of the leftover
    single_cut_ft = leftover_ft - 0.05  # Leave 0.05ft margin
    if single_cut_ft >= min_cut_ft:
        optimal_cuts.append({
            'cuts': [single_cut_ft],
            'cuts_inches': [single_cut_ft * 12],
            'total_ft': single_cut_ft,
            'total_inches': single_cut_ft * 12,
            'waste': leftover_ft - single_cut_ft,
            'utilization': (single_cut_ft / leftover_ft) * 100,
            'description': f'Single cut of {single_cut_ft:.2f}ft ({single_cut_ft*12:.1f}")'
        })
    
    # Two cuts that sum close to leftover - dynamic calculation
    # Try various ratios dynamically
    for ratio in [0.5, 0.55, 0.6, 0.65, 0.7]:  # Various split ratios
        cut1_ft = leftover_ft * ratio - 0.02
        cut2_ft = leftover_ft * (1 - ratio) - 0.02
        
        # Round to nearest 0.5 inch for practicality
        cut1_ft = round(cut1_ft * 24) / 24
        cut2_ft = round(cut2_ft * 24) / 24
        
        # Recalculate total after rounding
        total = cut1_ft + cut2_ft
        waste = leftover_ft - total
        
        if cut1_ft >= min_cut_ft and cut2_ft >= min_cut_ft and waste >= 0 and waste < 0.5:
            optimal_cuts.append({
                'cuts': [cut1_ft, cut2_ft],
                'cuts_inches': [cut1_ft * 12, cut2_ft * 12],
                'total_ft': total,
                'total_inches': total * 12,
                'waste': waste,
                'utilization': (total / leftover_ft) * 100,
                'description': f'Two cuts: {cut1_ft:.2f}ft ({cut1_ft*12:.1f}") + {cut2_ft:.2f}ft ({cut2_ft*12:.1f}") = {total:.2f}ft'
            })
    
    # Three cuts - dynamic calculation
    if leftover_ft >= 3 * min_cut_ft:
        for num_equal in [2, 3]:  # 2 equal + 1 different, or 3 equal
            if num_equal == 3:
                cut_size = (leftover_ft - 0.04) / 3
                cut_size = round(cut_size * 24) / 24
                cuts_3 = [cut_size, cut_size, leftover_ft - 2*cut_size - 0.02]
            else:
                # 2 equal + 1 different
                equal_size = (leftover_ft - 0.04) / 2.5
                equal_size = round(equal_size * 24) / 24
                diff_size = leftover_ft - 2*equal_size - 0.02
                cuts_3 = [equal_size, equal_size, diff_size]
            
            # Validate all cuts
            if all(c >= min_cut_ft for c in cuts_3):
                total = sum(cuts_3)
                waste = leftover_ft - total
                
                if waste >= 0 and waste < 0.5:
                    optimal_cuts.append({
                        'cuts': cuts_3,
                        'cuts_inches': [c * 12 for c in cuts_3],
                        'total_ft': total,
                        'total_inches': total * 12,
                        'waste': waste,
                        'utilization': (total / leftover_ft) * 100,
                        'description': f'Three cuts: {" + ".join([f"{c:.2f}ft ({c*12:.1f}\")" for c in cuts_3])} = {total:.2f}ft'
                    })
    
    # Four cuts - dynamic calculation
    if leftover_ft >= 4 * min_cut_ft:
        for num_equal in [2, 4]:  # 2 equal pairs or 4 equal
            if num_equal == 4:
                cut_size = (leftover_ft - 0.06) / 4
                cut_size = round(cut_size * 24) / 24
                cuts_4 = [cut_size] * 3 + [leftover_ft - 3*cut_size - 0.02]
            else:
                # 2 pairs of equal sizes
                pair1_size = (leftover_ft - 0.06) / 2.1
                pair1_size = round(pair1_size * 24) / 24
                pair2_size = (leftover_ft - 2*pair1_size - 0.02) / 2
                pair2_size = round(pair2_size * 24) / 24
                cuts_4 = [pair1_size, pair1_size, pair2_size, leftover_ft - 2*pair1_size - pair2_size - 0.02]
            
            if all(c >= min_cut_ft for c in cuts_4):
                total = sum(cuts_4)
                waste = leftover_ft - total
                
                if waste >= 0 and waste < 0.5:
                    optimal_cuts.append({
                        'cuts': cuts_4,
                        'cuts_inches': [c * 12 for c in cuts_4],
                        'total_ft': total,
                        'total_inches': total * 12,
                        'waste': waste,
                        'utilization': (total / leftover_ft) * 100,
                        'description': f'Four cuts: {" + ".join([f"{c:.2f}ft ({c*12:.1f}\")" for c in cuts_4[:2]])} + ... = {total:.2f}ft'
                    })
    
    # Sort by utilization (best first) and limit to top 5
    optimal_cuts.sort(key=lambda x: x['utilization'], reverse=True)
    return optimal_cuts[:5]

def _find_leftover_reuse_opportunities(extra_materials: Dict, elevations_data: Dict) -> List[Dict]:
    """
    Analyze leftover pieces and suggest optimal dimension combinations that would utilize them.
    Returns list of specific cut suggestions based on leftover dimensions.
    """
    reuse_suggestions = []
    
    if not extra_materials:
        return reuse_suggestions
    
    finish_codes = ['clear', 'bronze', 'grey', 'black', 'white']
    
    # Analyze each leftover material
    for material_key, material_data in extra_materials.items():
        length_pieces = material_data.get('length_pieces', [])
        if not length_pieces:
            # Check quantity for accessories
            quantity = material_data.get('quantity', 0)
            if quantity > 0:
                # Accessories - suggest reuse in future projects
                parts = material_key.split('-')
                if len(parts) >= 3 and parts[-1].lower() in finish_codes:
                    part_number = '-'.join(parts[:-1])
                    finish = parts[-1].lower()
                else:
                    part_number = material_key
                    finish = ""
                
                part_info = parts_data.get(part_number, {})
                description = part_info.get('Description', part_number)
                finish_text = f" ({finish.capitalize()})" if finish else ""
                
                if quantity > 20:  # Significant leftover quantity
                    reuse_suggestions.append({
                        'priority': 'medium',
                        'category': 'Leftover Reuse',
                        'message': f"You have {int(quantity)} leftover pieces of {description}{finish_text} (PN: {part_number}). Plan future projects to utilize these pieces to avoid purchasing new stock.",
                        'estimated_savings': quantity * 0.5,  # Rough estimate
                        'leftover_ft': quantity,
                        'part_number': part_number
                    })
            continue
        
        # Parse part number and finish from key
        parts = material_key.split('-')
        if len(parts) >= 3:
            if parts[-1].lower() in finish_codes:
                part_number = '-'.join(parts[:-1])
                finish = parts[-1].lower()
            else:
                part_number = material_key
                finish = ""
        elif len(parts) == 2:
            if parts[1].lower() in finish_codes:
                part_number = parts[0]
                finish = parts[1].lower()
            else:
                part_number = material_key
                finish = ""
        else:
            part_number = material_key
            finish = ""
        
        part_info = parts_data.get(part_number, {})
        description = part_info.get('Description', part_number)
        stock_length_ft = parse_length_to_feet(part_info.get('Length', '')) or 24.0
        
        # Get unit price for savings calculation
        unit_price, _ = get_unit_price_by_part(part_number, finish=finish)
        if not unit_price:
            continue
        
        # Sum up all leftover pieces for this material
        total_leftover_ft = sum(length_pieces)
        finish_text = f" ({finish.capitalize()})" if finish else ""
        
        # Calculate specific cuts that would FIT INSIDE each leftover piece
        for leftover_ft in sorted(length_pieces, reverse=True):
            if leftover_ft < 1.0:  # Skip very small pieces
                continue
            
            # Calculate specific cut dimensions that would fit inside this leftover
            optimal_cuts = _calculate_optimal_cuts_for_leftover(leftover_ft, stock_length_ft)
            
            if optimal_cuts:
                # Get best pattern (highest utilization)
                best_pattern = optimal_cuts[0]
                
                savings = leftover_ft * unit_price * 0.6
                
                # Determine if this is a horizontal (bay width) or vertical part
                is_horizontal = part_number in ["BE9-2513", "BE9-2514", "BE9-2515", "E9-2519"]
                is_vertical = part_number in ["E9-2512", "BE9-2511"]
                
                dimension_type = "bay widths" if is_horizontal else "bay heights" if is_vertical else "cut lengths"
                
                # Build message with specific dimension examples that FIT INSIDE the leftover
                if len(best_pattern['cuts']) == 1:
                    # Single cut
                    cut_inches = best_pattern['cuts_inches'][0]
                    message = f"You have a {leftover_ft:.2f}ft ({leftover_ft*12:.1f}\") leftover piece of {description}{finish_text} (PN: {part_number}). Use this leftover by setting {dimension_type} to {cut_inches:.1f}\" ({best_pattern['cuts'][0]:.2f}ft). This would use {best_pattern['utilization']:.1f}% of the leftover with only {best_pattern['waste']*12:.1f}\" waste."
                
                elif len(best_pattern['cuts']) == 2:
                    # Two cuts
                    cut1_inches, cut2_inches = best_pattern['cuts_inches']
                    total_inches = best_pattern['total_inches']
                    message = f"You have a {leftover_ft:.2f}ft ({leftover_ft*12:.1f}\") leftover piece of {description}{finish_text} (PN: {part_number}). Use this leftover by setting {dimension_type} to {cut1_inches:.1f}\" + {cut2_inches:.1f}\" = {total_inches:.1f}\" ({best_pattern['total_ft']:.2f}ft total). This would use {best_pattern['utilization']:.1f}% of the leftover with only {best_pattern['waste']*12:.1f}\" waste."
                
                else:
                    # Multiple cuts
                    cuts_inches_str = " + ".join([f"{c:.1f}\"" for c in best_pattern['cuts_inches']])
                    total_inches = best_pattern['total_inches']
                    message = f"You have a {leftover_ft:.2f}ft ({leftover_ft*12:.1f}\") leftover piece of {description}{finish_text} (PN: {part_number}). Use this leftover by setting {dimension_type} to {cuts_inches_str} = {total_inches:.1f}\" ({best_pattern['total_ft']:.2f}ft total). This would use {best_pattern['utilization']:.1f}% of the leftover with only {best_pattern['waste']*12:.1f}\" waste."
                
                # Add savings info
                message += f" This saves ${savings:.2f} by avoiding a new {stock_length_ft:.0f}ft stock purchase."
                
                reuse_suggestions.append({
                    'priority': 'high' if leftover_ft > 10 else 'medium',
                    'category': 'Leftover Reuse - Specific Dimensions',
                    'message': message,
                    'estimated_savings': savings,
                    'leftover_ft': leftover_ft,
                    'part_number': part_number,
                    'optimal_pattern': best_pattern
                })
    
    return reuse_suggestions

def generate_optimization_suggestions(
    material_breakdown: List[Dict], 
    overall_waste_percentage: float,
    elevations_data: Optional[Dict] = None,
    extra_materials: Optional[Dict] = None
) -> List[Dict]:
    """
    Generate specific, actionable optimization suggestions based on waste statistics and project data.
    """
    suggestions = []
    
    # Sort materials by waste cost for prioritization
    sorted_materials = sorted(material_breakdown, key=lambda x: x.get('waste_cost', 0), reverse=True)
    
    # NEW: Analyze leftover pieces for specific reuse opportunities (highest priority)
    if extra_materials and elevations_data:
        leftover_suggestions = _find_leftover_reuse_opportunities(extra_materials, elevations_data)
        if leftover_suggestions:
            # Sort by estimated savings
            leftover_suggestions.sort(key=lambda x: x.get('estimated_savings', 0), reverse=True)
            suggestions.extend(leftover_suggestions[:5])  # Top 5 leftover reuse suggestions
    
    # 1. High overall waste percentage with specific recommendations
    if overall_waste_percentage > 15:
        # Find specific materials causing high waste
        top_3_materials = sorted_materials[:3]
        material_list = ", ".join([f"{m['description']} ({m['waste_percentage']:.1f}%)" for m in top_3_materials])
        
        suggestions.append({
            "priority": "high",
            "category": "Overall Waste Reduction",
            "message": f"Overall waste is {overall_waste_percentage:.1f}%, significantly above the 10% target. Top contributors: {material_list}. Focus optimization efforts on these materials first, then consolidate orders across elevations to reduce total waste.",
            "estimated_savings": overall_waste_percentage * 100  # Rough estimate: $100 per percentage point
        })
    elif overall_waste_percentage > 10:
        suggestions.append({
            "priority": "medium",
            "category": "Overall Waste Reduction",
            "message": f"Overall waste is {overall_waste_percentage:.1f}%, slightly above the 10% target. Review top waste materials and consider order consolidation or bay configuration adjustments.",
            "estimated_savings": (overall_waste_percentage - 10) * 50  # Rough estimate
        })
    
    # 2. Specific material waste analysis with actionable suggestions
    for material in sorted_materials[:5]:  # Analyze top 5 waste materials
        waste_pct = material.get('waste_percentage', 0)
        waste_cost = material.get('waste_cost', 0)
        waste_qty = material.get('waste_quantity', 0)
        part_number = material.get('part_number', '')
        description = material.get('description', 'Unknown')
        finish = material.get('finish', '')
        total_qty = material.get('total_quantity', 0)
        unit = material.get('unit', 'ft')
        
        # Skip if already covered by high overall waste
        if waste_pct > 20 and waste_cost > 100:
            # Get part info for stock length
            part_info = parts_data.get(part_number, {})
            stock_length_ft = parse_length_to_feet(part_info.get('Length', '')) or 24.0
            
            # Calculate specific optimization opportunities
            avg_waste_per_stock = (waste_qty / (total_qty + waste_qty)) * stock_length_ft if (total_qty + waste_qty) > 0 else 0
            
            # Generate specific suggestion based on material type
            if part_number in ["BE9-2513", "BE9-2514", "BE9-2515", "E9-2519"] and elevations_data:
                # These are horizontal bay-width parts - suggest bay width adjustments
                suggestion = _generate_bay_width_suggestion(
                    material, elevations_data, stock_length_ft, part_number
                )
                if suggestion:
                    suggestions.append(suggestion)
                    continue  # Skip generic suggestion
            elif part_number in ["E9-2512", "BE9-2511"] and elevations_data:
                # Vertical parts - suggest height adjustments
                suggestion = _generate_height_suggestion(
                    material, elevations_data, stock_length_ft, part_number
                )
                if suggestion:
                    suggestions.append(suggestion)
                    continue  # Skip generic suggestion
            
            # Generic profile optimization if no specific suggestion generated
            if avg_waste_per_stock > stock_length_ft * 0.2:  # More than 20% waste per stock
                finish_text = f" ({finish.capitalize()})" if finish else ""
                potential_savings = waste_cost * 0.3  # Estimate 30% reduction possible
                
                suggestions.append({
                    "priority": "high" if waste_pct > 30 else "medium",
                    "category": "Cutting Strategy",
                    "message": f"{description}{finish_text} (PN: {part_number}) has {waste_pct:.1f}% waste generating ${waste_cost:.2f} in waste cost. With {waste_qty:.2f} {unit} of waste from {total_qty:.2f} {unit} used, average waste is {avg_waste_per_stock:.2f}ft per {stock_length_ft:.0f}ft stock length. Consider combining cuts across elevations or adjusting bay configurations to better utilize full stock lengths.",
                    "estimated_savings": potential_savings
                })
    
    # 3. High cost waste materials with specific actions
    high_cost_waste = [m for m in sorted_materials if m.get('waste_cost', 0) > 500]
    if high_cost_waste:
        for material in high_cost_waste[:3]:  # Top 3 high-cost waste materials
            waste_cost = material.get('waste_cost', 0)
            if waste_cost > 500:
                finish_text = f" ({material.get('finish', '').capitalize()})" if material.get('finish') else ""
                
                suggestions.append({
                    "priority": "high",
                    "category": "Cost Impact",
                    "message": f"{material.get('description', 'Unknown')}{finish_text} waste cost is ${waste_cost:.2f}, representing {waste_cost / (overall_waste_percentage * 10 + 1) * 100:.1f}% of total waste. This is a high-priority optimization target. Review cut patterns, consider consolidating with future projects, or adjust configurations to minimize this specific material waste.",
                    "estimated_savings": waste_cost * 0.25  # Estimate 25% reduction
                })
    
    # 4. Multiple small leftover pieces - suggest consolidation
    small_waste_materials = [m for m in material_breakdown 
                            if 0 < m.get('waste_quantity', 0) < 2 and m.get('waste_percentage', 0) > 5]
    if len(small_waste_materials) > 3:
        total_small_waste_cost = sum(m.get('waste_cost', 0) for m in small_waste_materials)
        part_numbers = [m.get('part_number', '') for m in small_waste_materials[:5]]
        
        suggestions.append({
            "priority": "medium",
            "category": "Leftover Consolidation",
            "message": f"Found {len(small_waste_materials)} materials with small leftover pieces (< 2{small_waste_materials[0].get('unit', 'ft')}), totaling ${total_small_waste_cost:.2f} in waste. Materials: {', '.join(part_numbers[:3])}. These small pieces may be usable in future projects. Consider maintaining a leftover inventory or adjusting future project configurations to utilize these pieces.",
            "estimated_savings": total_small_waste_cost * 0.5  # Could save 50% by reusing
        })
    
    # 5. Stock length optimization for profiles
    profile_materials = [m for m in sorted_materials if m.get('unit') == 'ft' and m.get('waste_percentage', 0) > 15]
    for material in profile_materials[:3]:
        part_number = material.get('part_number', '')
        part_info = parts_data.get(part_number, {})
        stock_length_ft = parse_length_to_feet(part_info.get('Length', '')) or 24.0
        total_needed = material.get('total_quantity', 0) + material.get('waste_quantity', 0)
        
        if stock_length_ft > 0 and total_needed > stock_length_ft * 2:
            # Calculate optimal stock usage
            current_stocks = math.ceil(total_needed / stock_length_ft)
            waste_per_stock = material.get('waste_quantity', 0) / max(1, current_stocks)
            
            if waste_per_stock > stock_length_ft * 0.15:  # More than 15% waste per stock
                finish_text = f" ({material.get('finish', '').capitalize()})" if material.get('finish') else ""
                
                suggestions.append({
                    "priority": "medium",
                    "category": "Stock Length Optimization",
                    "message": f"{material.get('description', 'Unknown')}{finish_text} averages {waste_per_stock:.2f}ft waste per {stock_length_ft:.0f}ft stock length ({waste_per_stock/stock_length_ft*100:.1f}% per stick). Total of {total_needed:.1f}ft needed across {current_stocks} stock(s). Consider adjusting bay configurations to create cut lengths that better fill {stock_length_ft:.0f}ft stocks, or combine with other elevations/projects using the same material.",
                    "estimated_savings": material.get('waste_cost', 0) * 0.2
                })
    
    # 6. Finish-specific consolidation suggestions
    if elevations_data:
        finish_breakdown = {}
        for elev_name, elev_data in elevations_data.items():
            finish = elev_data.get('finish', '').lower()
            if finish:
                if finish not in finish_breakdown:
                    finish_breakdown[finish] = []
                finish_breakdown[finish].append(elev_name)
        
        # Check for same finish with waste
        for finish, elev_names in finish_breakdown.items():
            if len(elev_names) > 1:
                finish_waste = [m for m in sorted_materials 
                              if m.get('finish', '').lower() == finish and m.get('waste_cost', 0) > 100]
                if finish_waste:
                    total_finish_waste = sum(m.get('waste_cost', 0) for m in finish_waste)
                    suggestions.append({
                        "priority": "medium",
                        "category": "Finish Consolidation",
                        "message": f"Multiple elevations use {finish.capitalize()} finish ({', '.join(elev_names[:3])}). Total waste cost for {finish.capitalize()} materials is ${total_finish_waste:.2f}. Consider consolidating orders for all {finish.capitalize()} materials across these elevations to better optimize cut lengths and reduce overall waste.",
                        "estimated_savings": total_finish_waste * 0.15
                    })
    
    # 7. Cross-elevation optimization
    if elevations_data and len(elevations_data) > 1:
        # Check if same materials are used in multiple elevations
        material_usage = {}
        for elev_name, elev_data in elevations_data.items():
            for output in elev_data.get('calculated_outputs', []):
                part = output.get('part_number', '').strip()
                finish = elev_data.get('finish', '').lower()
                key = f"{part}-{finish}" if finish else part
                if part and part != "N/A":
                    if key not in material_usage:
                        material_usage[key] = []
                    material_usage[key].append(elev_name)
        
        # Find materials used in multiple elevations with waste
        multi_elev_materials = {k: v for k, v in material_usage.items() if len(v) > 1}
        if multi_elev_materials:
            # Check if any of these have waste
            finish_codes = ['clear', 'bronze', 'grey', 'black', 'white']
            for key, elev_names in list(multi_elev_materials.items())[:3]:
                # Parse key to extract part number and finish
                # Handle keys like "BE9-2513-clear" or "PM-1006-SS" or "E1-0199"
                if '-' in key:
                    parts = key.split('-')
                    # Check if last part is a finish code
                    if parts[-1].lower() in finish_codes:
                        part = '-'.join(parts[:-1])
                        finish = parts[-1].lower()
                    else:
                        # No finish code, entire key is part number
                        part = key
                        finish = ''
                else:
                    part = key
                    finish = ''
                
                material = next((m for m in sorted_materials 
                               if m.get('part_number') == part and m.get('finish', '').lower() == finish.lower()), None)
                if material and material.get('waste_cost', 0) > 200:
                    suggestions.append({
                        "priority": "medium",
                        "category": "Cross-Elevation Optimization",
                        "message": f"{material.get('description', 'Unknown')} is used in {len(elev_names)} elevations ({', '.join(elev_names[:3])}) with ${material.get('waste_cost', 0):.2f} waste cost. Consider planning all {material.get('part_number', '')} cuts across these elevations together to optimize stock usage and reduce waste through better cut sequencing.",
                        "estimated_savings": material.get('waste_cost', 0) * 0.2
                    })
    
    # No high-priority suggestions
    if not suggestions:
        suggestions.append({
            "priority": "low",
            "category": "General",
            "message": f"Waste levels are acceptable at {overall_waste_percentage:.1f}%. Continue current optimization practices. Monitor waste trends to identify optimization opportunities early.",
            "estimated_savings": None
        })
    
    # Remove duplicates and limit suggestions
    unique_suggestions = []
    seen_messages = set()
    seen_leftovers = set()  # Track leftover pieces already suggested
    
    for suggestion in suggestions:
        msg_key = suggestion.get('message', '')[:100]  # First 100 chars as key
        
        # For leftover reuse suggestions, also check if we've already suggested this leftover
        if suggestion.get('category') == 'Leftover Reuse':
            leftover_key = f"{suggestion.get('part_number', '')}-{suggestion.get('leftover_ft', 0):.2f}"
            if leftover_key in seen_leftovers:
                continue
            seen_leftovers.add(leftover_key)
        
        if msg_key not in seen_messages:
            seen_messages.add(msg_key)
            unique_suggestions.append(suggestion)
    
    # Sort by priority (high first), then estimated savings (highest first)
    # Prioritize leftover reuse suggestions
    priority_order = {"high": 0, "medium": 1, "low": 2}
    unique_suggestions.sort(key=lambda x: (
        -1 if x.get('category') == 'Leftover Reuse' else 0,  # Leftover reuse first
        priority_order.get(x.get('priority', 'low'), 2),
        -(x.get('estimated_savings') or 0)
    ))
    
    # Return top suggestions, prioritizing leftover reuse
    return unique_suggestions[:10]

def get_waste_percentage_color(waste_percentage: float) -> str:
    """
    Get color code based on waste percentage (green < 10%, yellow 10-15%, red > 15%).
    """
    if waste_percentage < 10:
        return "#4CAF50"  # Green
    elif waste_percentage < 15:
        return "#FFC107"  # Yellow/Orange
    else:
        return "#F44336"  # Red
